"""
LBO (Leveraged Buyout) Model Generator
Generates institutional-grade LBO Excel models with debt schedules, returns analysis
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Border, Side, Alignment, NamedStyle,
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.workbook.defined_name import DefinedName
from typing import Dict, Any, List, Optional
from datetime import datetime
import os
import logging

logger = logging.getLogger(__name__)


def add_named_range(wb: Workbook, name: str, reference: str):
    """Helper to properly add a named range to a workbook"""
    try:
        defn = DefinedName(name, attr_text=reference)
        wb.defined_names.add(defn)
    except Exception as e:
        logger.debug(f"Could not add named range {name}: {e}")


class LBOStyler:
    """Professional LBO-specific Excel styles"""
    
    COLORS = {
        'primary': '1F4E79',
        'secondary': '2E75B6',
        'header': '1F4E79',
        'input': 'FFF2CC',
        'output': 'E2EFDA',
        'highlight': 'DCE6F1',
        'debt_senior': '4472C4',
        'debt_mezz': '7030A0',
        'debt_sub': 'C00000',
        'equity': '00B050',
        'white': 'FFFFFF',
        'black': '000000',
        'grey': 'D9D9D9',
    }
    
    @classmethod
    def create_styles(cls, wb: Workbook):
        """Create named styles for LBO model"""
        # Header style
        header = NamedStyle(name='lbo_header')
        header.font = Font(bold=True, color=cls.COLORS['white'], size=11)
        header.fill = PatternFill('solid', fgColor=cls.COLORS['header'])
        header.alignment = Alignment(horizontal='center', vertical='center')
        header.border = Border(
            bottom=Side(style='thin', color=cls.COLORS['black'])
        )
        wb.add_named_style(header)
        
        # Input cell style
        input_style = NamedStyle(name='lbo_input')
        input_style.fill = PatternFill('solid', fgColor=cls.COLORS['input'])
        input_style.font = Font(color=cls.COLORS['primary'])
        input_style.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        wb.add_named_style(input_style)
        
        # Output cell style
        output_style = NamedStyle(name='lbo_output')
        output_style.fill = PatternFill('solid', fgColor=cls.COLORS['output'])
        output_style.font = Font(bold=True)
        wb.add_named_style(output_style)
        
        # Returns highlight style
        returns_style = NamedStyle(name='lbo_returns')
        returns_style.fill = PatternFill('solid', fgColor=cls.COLORS['equity'])
        returns_style.font = Font(bold=True, color=cls.COLORS['white'], size=12)
        returns_style.alignment = Alignment(horizontal='center')
        wb.add_named_style(returns_style)


class LBOModelGenerator:
    """Generates production-grade LBO Excel models"""
    
    FORMATS = {
        'number': '#,##0',
        'decimal': '#,##0.0',
        'currency': '₹#,##0',
        'currency_millions': '₹#,##0.0,,"M"',
        'percent': '0.0%',
        'multiple': '0.00x',
        'years': '0.0',
        'irr': '0.0%',
    }
    
    def __init__(
        self,
        company_name: str,
        financial_data: Dict[str, Any],
        lbo_assumptions: Dict[str, Any],
        industry_info: Dict[str, Any],
    ):
        self.company_name = company_name
        self.financial_data = financial_data
        self.assumptions = lbo_assumptions
        self.industry_info = industry_info
        self.wb = Workbook()
        self.current_year = datetime.now().year
        
        # Default LBO assumptions if not provided
        self.lbo_defaults = {
            'entry_multiple': 8.0,
            'exit_multiple': 8.0,
            'holding_period': 5,
            'senior_debt_multiple': 3.0,
            'senior_interest_rate': 0.08,
            'senior_amort_years': 7,
            'mezz_debt_multiple': 1.5,
            'mezz_interest_rate': 0.12,
            'mezz_pik_rate': 0.02,
            'sub_debt_multiple': 0.5,
            'sub_interest_rate': 0.14,
            'min_cash_balance': 50,
            'revolver_size': 100,
            'revolver_rate': 0.07,
            'transaction_fees_pct': 0.02,
            'financing_fees_pct': 0.03,
            'management_rollover_pct': 0.10,
        }
        
        # Merge with provided assumptions
        for key, value in self.lbo_defaults.items():
            if key not in self.assumptions:
                self.assumptions[key] = value
    
    def generate(self, output_path: str) -> str:
        """Generate complete LBO Excel model"""
        try:
            # Remove default sheet
            if 'Sheet' in self.wb.sheetnames:
                del self.wb['Sheet']
            
            # Create styles
            LBOStyler.create_styles(self.wb)
            
            # Create all sheets
            self._create_cover_sheet()
            self._create_assumptions_sheet()
            self._create_sources_uses()
            self._create_operating_model()
            self._create_debt_schedule()
            self._create_cash_flow()
            self._create_returns_analysis()
            self._create_sensitivity()
            
            # Save workbook
            self.wb.save(output_path)
            logger.info(f"LBO model saved to {output_path}")
            return output_path
            
        except Exception as e:
            logger.error(f"Error generating LBO model: {e}")
            raise
    
    def _setup_sheet(self, name: str, title: str):
        """Setup a sheet with headers"""
        ws = self.wb.create_sheet(name)
        ws.merge_cells('A1:H1')
        ws['A1'] = title
        ws['A1'].font = Font(bold=True, size=14, color='1F4E79')
        ws['A1'].alignment = Alignment(horizontal='left')
        ws.row_dimensions[1].height = 25
        return ws
    
    def _create_cover_sheet(self):
        """Create LBO cover page"""
        ws = self._setup_sheet('Cover', f'{self.company_name} - LBO Analysis')
        
        # Company info
        info_data = [
            ('Company', self.company_name),
            ('Industry', self.industry_info.get('industry', 'General')),
            ('Analysis Date', datetime.now().strftime('%B %d, %Y')),
            ('Holding Period', f"{self.assumptions['holding_period']} Years"),
            ('Entry Multiple', f"{self.assumptions['entry_multiple']:.1f}x EV/EBITDA"),
            ('Exit Multiple', f"{self.assumptions['exit_multiple']:.1f}x EV/EBITDA"),
        ]
        
        row = 4
        for label, value in info_data:
            ws[f'B{row}'] = label
            ws[f'B{row}'].font = Font(bold=True)
            ws[f'C{row}'] = value
            row += 1
        
        # Model contents
        row += 2
        ws[f'B{row}'] = 'Model Contents'
        ws[f'B{row}'].font = Font(bold=True, size=12, color='1F4E79')
        row += 1
        
        sheets = [
            'Assumptions - Transaction & Operating Assumptions',
            'Sources & Uses - Transaction Funding Structure',
            'Operating Model - Revenue & EBITDA Projections',
            'Debt Schedule - Senior, Mezz, Subordinated Tranches',
            'Cash Flow - Free Cash Flow & Debt Paydown',
            'Returns - IRR, MoIC, Payback Analysis',
            'Sensitivity - Returns by Exit Multiple & Leverage',
        ]
        
        for sheet in sheets:
            ws[f'C{row}'] = f'• {sheet}'
            row += 1
        
        # Set column widths
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 50
    
    def _create_assumptions_sheet(self):
        """Create assumptions sheet with all LBO inputs"""
        ws = self._setup_sheet('Assumptions', 'LBO Assumptions')
        
        # Transaction Assumptions
        row = 4
        ws[f'B{row}'] = 'Transaction Assumptions'
        ws[f'B{row}'].font = Font(bold=True, size=11, color='1F4E79')
        row += 1
        
        trans_assumptions = [
            ('Entry EV/EBITDA Multiple', self.assumptions['entry_multiple'], 'x', 'entry_multiple'),
            ('Exit EV/EBITDA Multiple', self.assumptions['exit_multiple'], 'x', 'exit_multiple'),
            ('Holding Period (Years)', self.assumptions['holding_period'], '', 'holding_period'),
            ('Transaction Fees (% of EV)', self.assumptions['transaction_fees_pct'], '%', 'transaction_fees'),
            ('Financing Fees (% of Debt)', self.assumptions['financing_fees_pct'], '%', 'financing_fees'),
            ('Management Rollover (%)', self.assumptions['management_rollover_pct'], '%', 'mgmt_rollover'),
        ]
        
        for label, value, fmt, name in trans_assumptions:
            ws[f'B{row}'] = label
            ws[f'C{row}'] = value
            if fmt == '%':
                ws[f'C{row}'].number_format = self.FORMATS['percent']
            elif fmt == 'x':
                ws[f'C{row}'].number_format = self.FORMATS['multiple']
            ws[f'C{row}'].style = 'lbo_input'
            # Create named range
            add_named_range(self.wb, name, f"Assumptions!$C${row}")
            row += 1
        
        # Debt Structure
        row += 2
        ws[f'B{row}'] = 'Debt Structure'
        ws[f'B{row}'].font = Font(bold=True, size=11, color='1F4E79')
        row += 1
        
        debt_assumptions = [
            ('Senior Debt Multiple (x EBITDA)', self.assumptions['senior_debt_multiple'], 'x', 'senior_multiple'),
            ('Senior Interest Rate', self.assumptions['senior_interest_rate'], '%', 'senior_rate'),
            ('Senior Amortization (Years)', self.assumptions['senior_amort_years'], '', 'senior_amort'),
            ('Mezzanine Debt Multiple (x EBITDA)', self.assumptions['mezz_debt_multiple'], 'x', 'mezz_multiple'),
            ('Mezzanine Cash Interest', self.assumptions['mezz_interest_rate'], '%', 'mezz_rate'),
            ('Mezzanine PIK Interest', self.assumptions['mezz_pik_rate'], '%', 'mezz_pik'),
            ('Subordinated Debt Multiple (x EBITDA)', self.assumptions['sub_debt_multiple'], 'x', 'sub_multiple'),
            ('Subordinated Interest Rate', self.assumptions['sub_interest_rate'], '%', 'sub_rate'),
        ]
        
        for label, value, fmt, name in debt_assumptions:
            ws[f'B{row}'] = label
            ws[f'C{row}'] = value
            if fmt == '%':
                ws[f'C{row}'].number_format = self.FORMATS['percent']
            elif fmt == 'x':
                ws[f'C{row}'].number_format = self.FORMATS['multiple']
            ws[f'C{row}'].style = 'lbo_input'
            add_named_range(self.wb, name, f"Assumptions!$C${row}")
            row += 1
        
        # Operating Assumptions
        row += 2
        ws[f'B{row}'] = 'Operating Assumptions'
        ws[f'B{row}'].font = Font(bold=True, size=11, color='1F4E79')
        row += 1
        
        op_assumptions = [
            ('Revenue Growth Rate', self.assumptions.get('revenue_growth', 0.08), '%', 'rev_growth'),
            ('EBITDA Margin', self.assumptions.get('ebitda_margin', 0.25), '%', 'ebitda_margin'),
            ('D&A (% of Revenue)', self.assumptions.get('da_pct', 0.05), '%', 'da_pct'),
            ('CapEx (% of Revenue)', self.assumptions.get('capex_pct', 0.04), '%', 'capex_pct'),
            ('Working Capital (% of Revenue)', self.assumptions.get('nwc_pct', 0.10), '%', 'nwc_pct'),
            ('Tax Rate', self.assumptions.get('tax_rate', 0.25), '%', 'tax_rate'),
        ]
        
        for label, value, fmt, name in op_assumptions:
            ws[f'B{row}'] = label
            ws[f'C{row}'] = value
            ws[f'C{row}'].number_format = self.FORMATS['percent']
            ws[f'C{row}'].style = 'lbo_input'
            add_named_range(self.wb, name, f"Assumptions!$C${row}")
            row += 1
        
        # Set column widths
        ws.column_dimensions['B'].width = 35
        ws.column_dimensions['C'].width = 15
    
    def _create_sources_uses(self):
        """Create Sources & Uses of Funds"""
        ws = self._setup_sheet('Sources_Uses', 'Sources & Uses of Funds')
        
        # Get base EBITDA
        base_ebitda = self.financial_data.get('ebitda', 1000)
        entry_multiple = self.assumptions['entry_multiple']
        
        # Calculate transaction values
        enterprise_value = base_ebitda * entry_multiple
        
        row = 4
        # Uses of Funds
        ws[f'B{row}'] = 'Uses of Funds'
        ws[f'B{row}'].font = Font(bold=True, size=11, color='1F4E79')
        ws[f'C{row}'] = 'Amount'
        ws[f'D{row}'] = '% of Total'
        ws[f'C{row}'].font = Font(bold=True)
        ws[f'D{row}'].font = Font(bold=True)
        row += 1
        
        uses_start = row
        uses = [
            ('Purchase Enterprise Value', f'=IFERROR({base_ebitda}*entry_multiple,0)'),
            ('Transaction Fees', f'=IFERROR(B{row}*transaction_fees,0)'),
            ('Financing Fees', f'=IFERROR((senior_multiple+mezz_multiple+sub_multiple)*{base_ebitda}*financing_fees,0)'),
            ('Minimum Cash', self.assumptions['min_cash_balance']),
        ]
        
        for label, formula in uses:
            ws[f'B{row}'] = label
            ws[f'C{row}'] = formula
            ws[f'C{row}'].number_format = self.FORMATS['currency']
            row += 1
        
        uses_end = row - 1
        ws[f'B{row}'] = 'Total Uses'
        ws[f'B{row}'].font = Font(bold=True)
        ws[f'C{row}'] = f'=IFERROR(SUM(C{uses_start}:C{uses_end}),0)'
        ws[f'C{row}'].number_format = self.FORMATS['currency']
        ws[f'C{row}'].font = Font(bold=True)
        total_uses_row = row
        
        # Add % of Total formulas
        for r in range(uses_start, uses_end + 1):
            ws[f'D{r}'] = f'=IFERROR(C{r}/C{total_uses_row},0)'
            ws[f'D{r}'].number_format = self.FORMATS['percent']
        
        row += 3
        
        # Sources of Funds
        ws[f'B{row}'] = 'Sources of Funds'
        ws[f'B{row}'].font = Font(bold=True, size=11, color='1F4E79')
        ws[f'C{row}'] = 'Amount'
        ws[f'D{row}'] = '% of Total'
        ws[f'E{row}'] = 'x EBITDA'
        ws[f'C{row}'].font = Font(bold=True)
        ws[f'D{row}'].font = Font(bold=True)
        ws[f'E{row}'].font = Font(bold=True)
        row += 1
        
        sources_start = row
        sources = [
            ('Senior Debt', f'=IFERROR(senior_multiple*{base_ebitda},0)'),
            ('Mezzanine Debt', f'=IFERROR(mezz_multiple*{base_ebitda},0)'),
            ('Subordinated Debt', f'=IFERROR(sub_multiple*{base_ebitda},0)'),
            ('Management Rollover', f'=IFERROR((C{total_uses_row}-C{row}-C{row+1}-C{row+2})*mgmt_rollover,0)'),
        ]
        
        for label, formula in sources:
            ws[f'B{row}'] = label
            ws[f'C{row}'] = formula
            ws[f'C{row}'].number_format = self.FORMATS['currency']
            ws[f'E{row}'] = f'=IFERROR(C{row}/{base_ebitda},0)'
            ws[f'E{row}'].number_format = self.FORMATS['multiple']
            row += 1
        
        # Sponsor Equity (plug)
        ws[f'B{row}'] = 'Sponsor Equity (Plug)'
        ws[f'C{row}'] = f'=IFERROR(C{total_uses_row}-SUM(C{sources_start}:C{row-1}),0)'
        ws[f'C{row}'].number_format = self.FORMATS['currency']
        ws[f'C{row}'].style = 'lbo_output'
        ws[f'E{row}'] = f'=IFERROR(C{row}/{base_ebitda},0)'
        ws[f'E{row}'].number_format = self.FORMATS['multiple']
        sponsor_equity_row = row
        row += 1
        
        sources_end = row - 1
        ws[f'B{row}'] = 'Total Sources'
        ws[f'B{row}'].font = Font(bold=True)
        ws[f'C{row}'] = f'=IFERROR(SUM(C{sources_start}:C{sources_end}),0)'
        ws[f'C{row}'].number_format = self.FORMATS['currency']
        ws[f'C{row}'].font = Font(bold=True)
        total_sources_row = row
        
        # Add % of Total formulas
        for r in range(sources_start, sources_end + 1):
            ws[f'D{r}'] = f'=IFERROR(C{r}/C{total_sources_row},0)'
            ws[f'D{r}'].number_format = self.FORMATS['percent']
        
        # Balance check
        row += 2
        ws[f'B{row}'] = 'Sources - Uses Check'
        ws[f'B{row}'].font = Font(bold=True)
        ws[f'C{row}'] = f'=IFERROR(C{total_sources_row}-C{total_uses_row},0)'
        ws[f'C{row}'].number_format = self.FORMATS['currency']
        
        # Store key values as named ranges
        add_named_range(self.wb, 'total_uses', f"Sources_Uses!$C${total_uses_row}")
        add_named_range(self.wb, 'sponsor_equity', f"Sources_Uses!$C${sponsor_equity_row}")
        
        # Set column widths
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 12
    
    def _create_operating_model(self):
        """Create operating model projections"""
        ws = self._setup_sheet('Operating', 'Operating Model')
        
        holding_period = self.assumptions['holding_period']
        base_revenue = self.financial_data.get('revenue', 5000)
        
        # Year headers
        row = 4
        ws[f'B{row}'] = 'Fiscal Year'
        ws[f'B{row}'].font = Font(bold=True)
        for i in range(holding_period + 1):
            col = get_column_letter(3 + i)
            if i == 0:
                ws[f'{col}{row}'] = 'Entry'
            else:
                ws[f'{col}{row}'] = f'Year {i}'
            ws[f'{col}{row}'].style = 'lbo_header'
        
        row += 2
        
        # Revenue
        ws[f'B{row}'] = 'Revenue'
        ws[f'B{row}'].font = Font(bold=True)
        ws['C' + str(row)] = base_revenue
        ws['C' + str(row)].number_format = self.FORMATS['currency']
        for i in range(1, holding_period + 1):
            col = get_column_letter(3 + i)
            prev_col = get_column_letter(2 + i)
            ws[f'{col}{row}'] = f'=IFERROR({prev_col}{row}*(1+rev_growth),0)'
            ws[f'{col}{row}'].number_format = self.FORMATS['currency']
        revenue_row = row
        
        row += 1
        # Revenue Growth
        ws[f'B{row}'] = '  % Growth'
        ws[f'C{row}'] = '-'
        for i in range(1, holding_period + 1):
            col = get_column_letter(3 + i)
            prev_col = get_column_letter(2 + i)
            ws[f'{col}{row}'] = f'=IFERROR({col}{revenue_row}/{prev_col}{revenue_row}-1,0)'
            ws[f'{col}{row}'].number_format = self.FORMATS['percent']
        
        row += 2
        # EBITDA
        ws[f'B{row}'] = 'EBITDA'
        ws[f'B{row}'].font = Font(bold=True)
        for i in range(holding_period + 1):
            col = get_column_letter(3 + i)
            ws[f'{col}{row}'] = f'=IFERROR({col}{revenue_row}*ebitda_margin,0)'
            ws[f'{col}{row}'].number_format = self.FORMATS['currency']
        ebitda_row = row
        
        row += 1
        # EBITDA Margin
        ws[f'B{row}'] = '  % Margin'
        for i in range(holding_period + 1):
            col = get_column_letter(3 + i)
            ws[f'{col}{row}'] = f'=IFERROR({col}{ebitda_row}/{col}{revenue_row},0)'
            ws[f'{col}{row}'].number_format = self.FORMATS['percent']
        
        row += 2
        # D&A
        ws[f'B{row}'] = 'Depreciation & Amortization'
        for i in range(holding_period + 1):
            col = get_column_letter(3 + i)
            ws[f'{col}{row}'] = f'=IFERROR({col}{revenue_row}*da_pct,0)'
            ws[f'{col}{row}'].number_format = self.FORMATS['currency']
        da_row = row
        
        row += 1
        # EBIT
        ws[f'B{row}'] = 'EBIT'
        ws[f'B{row}'].font = Font(bold=True)
        for i in range(holding_period + 1):
            col = get_column_letter(3 + i)
            ws[f'{col}{row}'] = f'=IFERROR({col}{ebitda_row}-{col}{da_row},0)'
            ws[f'{col}{row}'].number_format = self.FORMATS['currency']
        ebit_row = row
        
        row += 2
        # Interest Expense (placeholder - linked from debt schedule)
        ws[f'B{row}'] = 'Interest Expense'
        for i in range(holding_period + 1):
            col = get_column_letter(3 + i)
            ws[f'{col}{row}'] = f"=IF(ISERROR(Debt_Schedule!{col}$50),0,Debt_Schedule!{col}$50)"
            ws[f'{col}{row}'].number_format = self.FORMATS['currency']
        interest_row = row
        
        row += 1
        # EBT
        ws[f'B{row}'] = 'EBT (Earnings Before Tax)'
        ws[f'B{row}'].font = Font(bold=True)
        for i in range(holding_period + 1):
            col = get_column_letter(3 + i)
            ws[f'{col}{row}'] = f'=IFERROR({col}{ebit_row}-{col}{interest_row},0)'
            ws[f'{col}{row}'].number_format = self.FORMATS['currency']
        ebt_row = row
        
        row += 1
        # Taxes
        ws[f'B{row}'] = 'Taxes'
        for i in range(holding_period + 1):
            col = get_column_letter(3 + i)
            ws[f'{col}{row}'] = f'=IFERROR(MAX(0,{col}{ebt_row}*tax_rate),0)'
            ws[f'{col}{row}'].number_format = self.FORMATS['currency']
        tax_row = row
        
        row += 1
        # Net Income
        ws[f'B{row}'] = 'Net Income'
        ws[f'B{row}'].font = Font(bold=True)
        for i in range(holding_period + 1):
            col = get_column_letter(3 + i)
            ws[f'{col}{row}'] = f'=IFERROR({col}{ebt_row}-{col}{tax_row},0)'
            ws[f'{col}{row}'].number_format = self.FORMATS['currency']
            ws[f'{col}{row}'].style = 'lbo_output'
        
        # Store references
        add_named_range(self.wb, 'op_ebitda_row', f"Operating!$B${ebitda_row}")
        add_named_range(self.wb, 'op_revenue_row', f"Operating!$B${revenue_row}")
        
        # Set column widths
        ws.column_dimensions['B'].width = 30
        for i in range(holding_period + 2):
            ws.column_dimensions[get_column_letter(3 + i)].width = 14
    
    def _create_debt_schedule(self):
        """Create detailed debt schedule with all tranches"""
        ws = self._setup_sheet('Debt_Schedule', 'Debt Schedule')
        
        holding_period = self.assumptions['holding_period']
        base_ebitda = self.financial_data.get('ebitda', 1000)
        
        # Year headers
        row = 4
        ws[f'B{row}'] = 'Fiscal Year'
        ws[f'B{row}'].font = Font(bold=True)
        for i in range(holding_period + 1):
            col = get_column_letter(3 + i)
            if i == 0:
                ws[f'{col}{row}'] = 'Entry'
            else:
                ws[f'{col}{row}'] = f'Year {i}'
            ws[f'{col}{row}'].style = 'lbo_header'
        
        row += 2
        
        # ============ SENIOR DEBT ============
        ws[f'B{row}'] = 'Senior Debt'
        ws[f'B{row}'].font = Font(bold=True, size=11, color=LBOStyler.COLORS['debt_senior'])
        row += 1
        
        # Beginning Balance
        ws[f'B{row}'] = 'Beginning Balance'
        ws['C' + str(row)] = f'=IFERROR(senior_multiple*{base_ebitda},0)'
        ws['C' + str(row)].number_format = self.FORMATS['currency']
        for i in range(1, holding_period + 1):
            col = get_column_letter(3 + i)
            prev_col = get_column_letter(2 + i)
            ws[f'{col}{row}'] = f'=IFERROR({prev_col}{row+3},0)'  # End of prev period
            ws[f'{col}{row}'].number_format = self.FORMATS['currency']
        senior_begin_row = row
        
        row += 1
        # Mandatory Amortization
        ws[f'B{row}'] = 'Mandatory Amortization'
        ws['C' + str(row)] = 0
        for i in range(1, holding_period + 1):
            col = get_column_letter(3 + i)
            ws[f'{col}{row}'] = f'=IFERROR(-C{senior_begin_row}/senior_amort,0)'
            ws[f'{col}{row}'].number_format = self.FORMATS['currency']
        senior_amort_row = row
        
        row += 1
        # Optional Prepayment (cash sweep)
        ws[f'B{row}'] = 'Optional Prepayment'
        for i in range(holding_period + 1):
            col = get_column_letter(3 + i)
            if i == 0:
                ws[f'{col}{row}'] = 0
            else:
                # Cash sweep - excess cash goes to debt paydown
                ws[f'{col}{row}'] = f'=IFERROR(-MIN({col}{senior_begin_row}+{col}{senior_amort_row},MAX(0,Cash_Flow!{col}$40)),0)'
            ws[f'{col}{row}'].number_format = self.FORMATS['currency']
        senior_prepay_row = row
        
        row += 1
        # Ending Balance
        ws[f'B{row}'] = 'Ending Balance'
        ws[f'B{row}'].font = Font(bold=True)
        for i in range(holding_period + 1):
            col = get_column_letter(3 + i)
            ws[f'{col}{row}'] = f'=IFERROR(MAX(0,{col}{senior_begin_row}+{col}{senior_amort_row}+{col}{senior_prepay_row}),0)'
            ws[f'{col}{row}'].number_format = self.FORMATS['currency']
            ws[f'{col}{row}'].style = 'lbo_output'
        senior_end_row = row
        
        row += 1
        # Interest Expense
        ws[f'B{row}'] = 'Interest Expense'
        for i in range(holding_period + 1):
            col = get_column_letter(3 + i)
            ws[f'{col}{row}'] = f'=IFERROR({col}{senior_begin_row}*senior_rate,0)'
            ws[f'{col}{row}'].number_format = self.FORMATS['currency']
        senior_interest_row = row
        
        row += 3
        
        # ============ MEZZANINE DEBT ============
        ws[f'B{row}'] = 'Mezzanine Debt'
        ws[f'B{row}'].font = Font(bold=True, size=11, color=LBOStyler.COLORS['debt_mezz'])
        row += 1
        
        # Beginning Balance
        ws[f'B{row}'] = 'Beginning Balance'
        ws['C' + str(row)] = f'=IFERROR(mezz_multiple*{base_ebitda},0)'
        ws['C' + str(row)].number_format = self.FORMATS['currency']
        for i in range(1, holding_period + 1):
            col = get_column_letter(3 + i)
            prev_col = get_column_letter(2 + i)
            ws[f'{col}{row}'] = f'=IFERROR({prev_col}{row+2},0)'  # End of prev period + PIK
            ws[f'{col}{row}'].number_format = self.FORMATS['currency']
        mezz_begin_row = row
        
        row += 1
        # PIK Interest (adds to principal)
        ws[f'B{row}'] = 'PIK Interest'
        for i in range(holding_period + 1):
            col = get_column_letter(3 + i)
            ws[f'{col}{row}'] = f'=IFERROR({col}{mezz_begin_row}*mezz_pik,0)'
            ws[f'{col}{row}'].number_format = self.FORMATS['currency']
        mezz_pik_row = row
        
        row += 1
        # Ending Balance
        ws[f'B{row}'] = 'Ending Balance'
        ws[f'B{row}'].font = Font(bold=True)
        for i in range(holding_period + 1):
            col = get_column_letter(3 + i)
            ws[f'{col}{row}'] = f'=IFERROR({col}{mezz_begin_row}+{col}{mezz_pik_row},0)'
            ws[f'{col}{row}'].number_format = self.FORMATS['currency']
            ws[f'{col}{row}'].style = 'lbo_output'
        mezz_end_row = row
        
        row += 1
        # Cash Interest
        ws[f'B{row}'] = 'Cash Interest'
        for i in range(holding_period + 1):
            col = get_column_letter(3 + i)
            ws[f'{col}{row}'] = f'=IFERROR({col}{mezz_begin_row}*mezz_rate,0)'
            ws[f'{col}{row}'].number_format = self.FORMATS['currency']
        mezz_interest_row = row
        
        row += 3
        
        # ============ SUBORDINATED DEBT ============
        ws[f'B{row}'] = 'Subordinated Debt'
        ws[f'B{row}'].font = Font(bold=True, size=11, color=LBOStyler.COLORS['debt_sub'])
        row += 1
        
        # Beginning Balance
        ws[f'B{row}'] = 'Beginning Balance'
        ws['C' + str(row)] = f'=IFERROR(sub_multiple*{base_ebitda},0)'
        ws['C' + str(row)].number_format = self.FORMATS['currency']
        for i in range(1, holding_period + 1):
            col = get_column_letter(3 + i)
            prev_col = get_column_letter(2 + i)
            ws[f'{col}{row}'] = f'=IFERROR({prev_col}{row+1},0)'
            ws[f'{col}{row}'].number_format = self.FORMATS['currency']
        sub_begin_row = row
        
        row += 1
        # Ending Balance (bullet at exit)
        ws[f'B{row}'] = 'Ending Balance'
        ws[f'B{row}'].font = Font(bold=True)
        for i in range(holding_period + 1):
            col = get_column_letter(3 + i)
            ws[f'{col}{row}'] = f'=IFERROR({col}{sub_begin_row},0)'
            ws[f'{col}{row}'].number_format = self.FORMATS['currency']
            ws[f'{col}{row}'].style = 'lbo_output'
        sub_end_row = row
        
        row += 1
        # Interest Expense
        ws[f'B{row}'] = 'Interest Expense'
        for i in range(holding_period + 1):
            col = get_column_letter(3 + i)
            ws[f'{col}{row}'] = f'=IFERROR({col}{sub_begin_row}*sub_rate,0)'
            ws[f'{col}{row}'].number_format = self.FORMATS['currency']
        sub_interest_row = row
        
        row += 3
        
        # ============ TOTAL DEBT SUMMARY ============
        ws[f'B{row}'] = 'Total Debt Summary'
        ws[f'B{row}'].font = Font(bold=True, size=11, color='1F4E79')
        row += 1
        
        # Total Debt
        ws[f'B{row}'] = 'Total Debt'
        ws[f'B{row}'].font = Font(bold=True)
        for i in range(holding_period + 1):
            col = get_column_letter(3 + i)
            ws[f'{col}{row}'] = f'=IFERROR({col}{senior_end_row}+{col}{mezz_end_row}+{col}{sub_end_row},0)'
            ws[f'{col}{row}'].number_format = self.FORMATS['currency']
            ws[f'{col}{row}'].style = 'lbo_output'
        total_debt_row = row
        row += 50 - row  # Move to row 50 for interest reference
        
        # Total Interest (at row 50 for reference from Operating Model)
        ws[f'B50'] = 'Total Interest Expense'
        ws[f'B50'].font = Font(bold=True)
        for i in range(holding_period + 1):
            col = get_column_letter(3 + i)
            ws[f'{col}50'] = f'=IFERROR({col}{senior_interest_row}+{col}{mezz_interest_row}+{col}{sub_interest_row},0)'
            ws[f'{col}50'].number_format = self.FORMATS['currency']
        
        # Set column widths
        ws.column_dimensions['B'].width = 25
        for i in range(holding_period + 2):
            ws.column_dimensions[get_column_letter(3 + i)].width = 14
    
    def _create_cash_flow(self):
        """Create cash flow statement for LBO"""
        ws = self._setup_sheet('Cash_Flow', 'Cash Flow Statement')
        
        holding_period = self.assumptions['holding_period']
        
        # Year headers
        row = 4
        ws[f'B{row}'] = 'Fiscal Year'
        ws[f'B{row}'].font = Font(bold=True)
        for i in range(holding_period + 1):
            col = get_column_letter(3 + i)
            if i == 0:
                ws[f'{col}{row}'] = 'Entry'
            else:
                ws[f'{col}{row}'] = f'Year {i}'
            ws[f'{col}{row}'].style = 'lbo_header'
        
        row += 2
        
        # EBITDA
        ws[f'B{row}'] = 'EBITDA'
        ws[f'B{row}'].font = Font(bold=True)
        for i in range(holding_period + 1):
            col = get_column_letter(3 + i)
            ws[f'{col}{row}'] = f'=IFERROR(Operating!{col}$10,0)'
            ws[f'{col}{row}'].number_format = self.FORMATS['currency']
        ebitda_row = row
        
        row += 1
        # Less: Interest
        ws[f'B{row}'] = 'Less: Cash Interest'
        for i in range(holding_period + 1):
            col = get_column_letter(3 + i)
            ws[f'{col}{row}'] = f'=IFERROR(-Debt_Schedule!{col}$50,0)'
            ws[f'{col}{row}'].number_format = self.FORMATS['currency']
        interest_row = row
        
        row += 1
        # Less: Taxes
        ws[f'B{row}'] = 'Less: Cash Taxes'
        for i in range(holding_period + 1):
            col = get_column_letter(3 + i)
            ws[f'{col}{row}'] = f'=IFERROR(-MAX(0,({col}{ebitda_row}+{col}{interest_row})*tax_rate),0)'
            ws[f'{col}{row}'].number_format = self.FORMATS['currency']
        tax_row = row
        
        row += 1
        # Less: CapEx
        ws[f'B{row}'] = 'Less: CapEx'
        for i in range(holding_period + 1):
            col = get_column_letter(3 + i)
            ws[f'{col}{row}'] = f'=IFERROR(-Operating!{col}$6*capex_pct,0)'
            ws[f'{col}{row}'].number_format = self.FORMATS['currency']
        capex_row = row
        
        row += 1
        # Less: Change in NWC
        ws[f'B{row}'] = 'Less: Change in NWC'
        for i in range(holding_period + 1):
            col = get_column_letter(3 + i)
            if i == 0:
                ws[f'{col}{row}'] = 0
            else:
                prev_col = get_column_letter(2 + i)
                ws[f'{col}{row}'] = f'=IFERROR(-(Operating!{col}$6-Operating!{prev_col}$6)*nwc_pct,0)'
            ws[f'{col}{row}'].number_format = self.FORMATS['currency']
        nwc_row = row
        
        row += 2
        # Free Cash Flow
        ws[f'B{row}'] = 'Free Cash Flow'
        ws[f'B{row}'].font = Font(bold=True)
        for i in range(holding_period + 1):
            col = get_column_letter(3 + i)
            ws[f'{col}{row}'] = f'=IFERROR(SUM({col}{ebitda_row}:{col}{nwc_row}),0)'
            ws[f'{col}{row}'].number_format = self.FORMATS['currency']
            ws[f'{col}{row}'].style = 'lbo_output'
        fcf_row = row
        
        row += 2
        # Mandatory Debt Repayment
        ws[f'B{row}'] = 'Less: Mandatory Debt Amortization'
        for i in range(holding_period + 1):
            col = get_column_letter(3 + i)
            ws[f'{col}{row}'] = f'=IFERROR(Debt_Schedule!{col}$8,0)'
            ws[f'{col}{row}'].number_format = self.FORMATS['currency']
        amort_row = row
        
        row += 2
        # Cash Available for Sweep (Row 40 for debt schedule reference)
        ws[f'B40'] = 'Cash Available for Sweep'
        ws[f'B40'].font = Font(bold=True)
        for i in range(holding_period + 1):
            col = get_column_letter(3 + i)
            ws[f'{col}40'] = f'=IFERROR({col}{fcf_row}+{col}{amort_row},0)'
            ws[f'{col}40'].number_format = self.FORMATS['currency']
            ws[f'{col}40'].style = 'lbo_output'
        
        # Set column widths
        ws.column_dimensions['B'].width = 30
        for i in range(holding_period + 2):
            ws.column_dimensions[get_column_letter(3 + i)].width = 14
    
    def _create_returns_analysis(self):
        """Create returns analysis with IRR, MoIC, payback"""
        ws = self._setup_sheet('Returns', 'Returns Analysis')
        
        holding_period = self.assumptions['holding_period']
        base_ebitda = self.financial_data.get('ebitda', 1000)
        
        row = 4
        # Exit Analysis
        ws[f'B{row}'] = 'Exit Analysis'
        ws[f'B{row}'].font = Font(bold=True, size=11, color='1F4E79')
        row += 1
        
        exit_col = get_column_letter(3 + holding_period)
        
        # Exit EBITDA
        ws[f'B{row}'] = 'Exit Year EBITDA'
        ws[f'C{row}'] = f'=IFERROR(Operating!{exit_col}$10,0)'
        ws[f'C{row}'].number_format = self.FORMATS['currency']
        exit_ebitda_row = row
        
        row += 1
        # Exit Multiple
        ws[f'B{row}'] = 'Exit EV/EBITDA Multiple'
        ws[f'C{row}'] = '=exit_multiple'
        ws[f'C{row}'].number_format = self.FORMATS['multiple']
        
        row += 1
        # Exit Enterprise Value
        ws[f'B{row}'] = 'Exit Enterprise Value'
        ws[f'C{row}'] = f'=IFERROR(C{exit_ebitda_row}*exit_multiple,0)'
        ws[f'C{row}'].number_format = self.FORMATS['currency']
        ws[f'C{row}'].font = Font(bold=True)
        exit_ev_row = row
        
        row += 1
        # Less: Exit Debt
        ws[f'B{row}'] = 'Less: Debt at Exit'
        ws[f'C{row}'] = f'=IFERROR(-Debt_Schedule!{exit_col}$45,0)'
        ws[f'C{row}'].number_format = self.FORMATS['currency']
        exit_debt_row = row
        
        row += 1
        # Exit Equity Value
        ws[f'B{row}'] = 'Exit Equity Value'
        ws[f'C{row}'] = f'=IFERROR(C{exit_ev_row}+C{exit_debt_row},0)'
        ws[f'C{row}'].number_format = self.FORMATS['currency']
        ws[f'C{row}'].font = Font(bold=True)
        ws[f'C{row}'].style = 'lbo_output'
        exit_equity_row = row
        
        row += 3
        # Returns Summary
        ws[f'B{row}'] = 'Returns Summary'
        ws[f'B{row}'].font = Font(bold=True, size=11, color=LBOStyler.COLORS['equity'])
        row += 1
        
        # Initial Equity Investment
        ws[f'B{row}'] = 'Initial Sponsor Equity'
        ws[f'C{row}'] = '=IFERROR(sponsor_equity,0)'
        ws[f'C{row}'].number_format = self.FORMATS['currency']
        initial_equity_row = row
        
        row += 1
        # Exit Equity Proceeds
        ws[f'B{row}'] = 'Exit Equity Proceeds'
        ws[f'C{row}'] = f'=IFERROR(C{exit_equity_row},0)'
        ws[f'C{row}'].number_format = self.FORMATS['currency']
        
        row += 2
        # MoIC
        ws[f'B{row}'] = 'Multiple on Invested Capital (MoIC)'
        ws[f'C{row}'] = f'=IFERROR(C{exit_equity_row}/C{initial_equity_row},0)'
        ws[f'C{row}'].number_format = self.FORMATS['multiple']
        ws[f'C{row}'].style = 'lbo_returns'
        moic_row = row
        
        row += 1
        # IRR (using Excel formula)
        ws[f'B{row}'] = 'Internal Rate of Return (IRR)'
        # Create cash flow array for IRR
        cf_range = f'-C{initial_equity_row}'
        for i in range(1, holding_period):
            cf_range += ',0'
        cf_range += f',C{exit_equity_row}'
        ws[f'C{row}'] = f'=IFERROR(IRR({{{cf_range}}}),0)'
        ws[f'C{row}'].number_format = self.FORMATS['irr']
        ws[f'C{row}'].style = 'lbo_returns'
        irr_row = row
        
        row += 1
        # Payback Period
        ws[f'B{row}'] = 'Payback Period (Years)'
        ws[f'C{row}'] = f'=IFERROR(holding_period/LN(C{moic_row})*LN(2),0)'
        ws[f'C{row}'].number_format = self.FORMATS['years']
        
        row += 3
        # Cash-on-Cash Return by Year
        ws[f'B{row}'] = 'Cash-on-Cash Return by Exit Year'
        ws[f'B{row}'].font = Font(bold=True, size=11, color='1F4E79')
        row += 1
        
        ws[f'B{row}'] = 'Exit Year'
        ws[f'C{row}'] = 'MoIC'
        ws[f'D{row}'] = 'IRR'
        ws[f'C{row}'].font = Font(bold=True)
        ws[f'D{row}'].font = Font(bold=True)
        row += 1
        
        for year in range(1, holding_period + 1):
            ws[f'B{row}'] = f'Year {year}'
            yr_col = get_column_letter(3 + year)
            ws[f'C{row}'] = f'=IFERROR((Operating!{yr_col}$10*exit_multiple-Debt_Schedule!{yr_col}$45)/sponsor_equity,0)'
            ws[f'C{row}'].number_format = self.FORMATS['multiple']
            ws[f'D{row}'] = f'=IFERROR(C{row}^(1/{year})-1,0)'
            ws[f'D{row}'].number_format = self.FORMATS['irr']
            row += 1
        
        # Set column widths
        ws.column_dimensions['B'].width = 35
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 12
    
    def _create_sensitivity(self):
        """Create sensitivity tables for returns"""
        ws = self._setup_sheet('Sensitivity', 'Sensitivity Analysis')
        
        row = 4
        # Exit Multiple vs Entry Multiple
        ws[f'B{row}'] = 'IRR Sensitivity: Exit Multiple vs Entry Multiple'
        ws[f'B{row}'].font = Font(bold=True, size=11, color='1F4E79')
        row += 2
        
        # Headers
        ws[f'B{row}'] = 'IRR'
        entry_multiples = [6.0, 7.0, 8.0, 9.0, 10.0]
        exit_multiples = [6.0, 7.0, 8.0, 9.0, 10.0, 11.0]
        
        # Column headers (Exit Multiples)
        for i, em in enumerate(exit_multiples):
            col = get_column_letter(3 + i)
            ws[f'{col}{row}'] = f'{em:.1f}x'
            ws[f'{col}{row}'].style = 'lbo_header'
        row += 1
        
        # Row headers (Entry multiples) and values
        base_ebitda = self.financial_data.get('ebitda', 1000)
        holding_period = self.assumptions['holding_period']
        
        for entry_m in entry_multiples:
            ws[f'B{row}'] = f'{entry_m:.1f}x Entry'
            ws[f'B{row}'].font = Font(bold=True)
            
            for i, exit_m in enumerate(exit_multiples):
                col = get_column_letter(3 + i)
                # Simplified IRR calculation for sensitivity
                # Entry EV = entry_m * EBITDA, Exit EV = exit_m * EBITDA * (1+growth)^years
                ws[f'{col}{row}'] = f'=IFERROR((({exit_m}*{base_ebitda}*(1+rev_growth)^{holding_period})/(({entry_m}*{base_ebitda})*0.4))^(1/{holding_period})-1,0)'
                ws[f'{col}{row}'].number_format = self.FORMATS['irr']
            row += 1
        
        row += 3
        
        # MoIC Sensitivity
        ws[f'B{row}'] = 'MoIC Sensitivity: Exit Multiple vs Leverage'
        ws[f'B{row}'].font = Font(bold=True, size=11, color='1F4E79')
        row += 2
        
        # Headers
        ws[f'B{row}'] = 'MoIC'
        leverage_levels = [3.0, 4.0, 5.0, 6.0]
        
        for i, em in enumerate(exit_multiples):
            col = get_column_letter(3 + i)
            ws[f'{col}{row}'] = f'{em:.1f}x Exit'
            ws[f'{col}{row}'].style = 'lbo_header'
        row += 1
        
        for lev in leverage_levels:
            ws[f'B{row}'] = f'{lev:.1f}x Leverage'
            ws[f'B{row}'].font = Font(bold=True)
            
            for i, exit_m in enumerate(exit_multiples):
                col = get_column_letter(3 + i)
                # MoIC = (Exit EV - Debt) / Equity = (Exit EV - Debt) / (Entry EV - Debt)
                entry_ev = self.assumptions['entry_multiple'] * base_ebitda
                debt = lev * base_ebitda
                equity = entry_ev - debt
                exit_ev = exit_m * base_ebitda * (1 + self.assumptions.get('revenue_growth', 0.08)) ** holding_period
                ws[f'{col}{row}'] = f'=IFERROR(({exit_ev}-{debt}*0.7)/{equity},0)'
                ws[f'{col}{row}'].number_format = self.FORMATS['multiple']
            row += 1
        
        # Set column widths
        ws.column_dimensions['B'].width = 20
        for i in range(8):
            ws.column_dimensions[get_column_letter(3 + i)].width = 12


def generate_lbo_model(
    company_name: str,
    financial_data: Dict[str, Any],
    lbo_assumptions: Dict[str, Any],
    industry_info: Dict[str, Any],
    output_path: str
) -> str:
    """Generate LBO model"""
    generator = LBOModelGenerator(
        company_name=company_name,
        financial_data=financial_data,
        lbo_assumptions=lbo_assumptions,
        industry_info=industry_info,
    )
    return generator.generate(output_path)
