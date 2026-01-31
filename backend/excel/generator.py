"""
Excel Financial Model Generator V2
Production-grade Excel models with properly linked formulas and real data
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Border, Side, Alignment, NamedStyle,
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, BarChart, Reference
from typing import Dict, Any, List, Optional
from datetime import datetime
import os
import logging

logger = logging.getLogger(__name__)


class ExcelStyler:
    """Professional Excel styles"""
    
    COLORS = {
        'primary': '1F4E79',
        'secondary': '2E75B6',
        'header': '1F4E79',
        'input': 'FFF2CC',
        'output': 'E2EFDA',
        'white': 'FFFFFF',
        'black': '000000',
        'grey': 'D9D9D9',
        'dark_grey': '404040',
    }
    
    @classmethod
    def create_styles(cls, wb: Workbook) -> Dict[str, NamedStyle]:
        styles = {}
        
        # Title
        title = NamedStyle(name='title')
        title.font = Font(name='Calibri', size=16, bold=True, color=cls.COLORS['primary'])
        title.alignment = Alignment(horizontal='left', vertical='center')
        wb.add_named_style(title)
        styles['title'] = title
        
        # Header
        header = NamedStyle(name='header')
        header.font = Font(name='Calibri', size=10, bold=True, color=cls.COLORS['white'])
        header.fill = PatternFill(start_color=cls.COLORS['header'], end_color=cls.COLORS['header'], fill_type='solid')
        header.alignment = Alignment(horizontal='center', vertical='center')
        header.border = Border(bottom=Side(style='thin', color=cls.COLORS['black']))
        wb.add_named_style(header)
        styles['header'] = header
        
        # Subheader
        subheader = NamedStyle(name='subheader')
        subheader.font = Font(name='Calibri', size=10, bold=True, color=cls.COLORS['primary'])
        subheader.fill = PatternFill(start_color=cls.COLORS['grey'], end_color=cls.COLORS['grey'], fill_type='solid')
        subheader.alignment = Alignment(horizontal='left', vertical='center')
        wb.add_named_style(subheader)
        styles['subheader'] = subheader
        
        # Input (yellow)
        input_cell = NamedStyle(name='input_cell')
        input_cell.font = Font(name='Calibri', size=10, color=cls.COLORS['primary'])
        input_cell.fill = PatternFill(start_color=cls.COLORS['input'], end_color=cls.COLORS['input'], fill_type='solid')
        input_cell.border = Border(
            left=Side(style='thin', color=cls.COLORS['grey']),
            right=Side(style='thin', color=cls.COLORS['grey']),
            top=Side(style='thin', color=cls.COLORS['grey']),
            bottom=Side(style='thin', color=cls.COLORS['grey'])
        )
        input_cell.alignment = Alignment(horizontal='right')
        wb.add_named_style(input_cell)
        styles['input'] = input_cell
        
        # Output (green)
        output_cell = NamedStyle(name='output_cell')
        output_cell.font = Font(name='Calibri', size=10, color=cls.COLORS['dark_grey'])
        output_cell.fill = PatternFill(start_color=cls.COLORS['output'], end_color=cls.COLORS['output'], fill_type='solid')
        output_cell.alignment = Alignment(horizontal='right')
        wb.add_named_style(output_cell)
        styles['output'] = output_cell
        
        # Label
        label = NamedStyle(name='label')
        label.font = Font(name='Calibri', size=10)
        label.alignment = Alignment(horizontal='left', indent=1)
        wb.add_named_style(label)
        styles['label'] = label
        
        # Total
        total = NamedStyle(name='total')
        total.font = Font(name='Calibri', size=10, bold=True)
        total.border = Border(
            top=Side(style='thin', color=cls.COLORS['black']),
            bottom=Side(style='double', color=cls.COLORS['black'])
        )
        total.alignment = Alignment(horizontal='right')
        wb.add_named_style(total)
        styles['total'] = total
        
        return styles


class FinancialModelGenerator:
    """Generates production-grade Excel financial models"""
    
    FORMATS = {
        'number': '#,##0',
        'decimal': '#,##0.0',
        'currency': '₹#,##0',
        'percent': '0.0%',
        'ratio': '0.00x',
    }
    
    def __init__(
        self,
        company_name: str,
        model_structure: Dict[str, Any],
        financial_data: Dict[str, Any],
        industry_info: Dict[str, Any],
    ):
        self.company_name = company_name
        self.structure = model_structure
        self.data = financial_data
        self.industry = industry_info
        self.wb = Workbook()
        self.styles = ExcelStyler.create_styles(self.wb)
        
        # Config
        self.hist_years = 5
        self.fcst_years = model_structure.get('forecast_years', 5)
        self.base_year = datetime.now().year
        
        # Store row mappings for cross-sheet references
        self.row_map = {}
    
    def generate(self, output_path: str) -> str:
        """Generate complete Excel model"""
        if 'Sheet' in self.wb.sheetnames:
            del self.wb['Sheet']
        
        # Create all sheets in order
        self._create_assumptions()
        self._create_income_statement()
        self._create_balance_sheet()
        self._create_cash_flow()
        self._create_valuation()
        self._create_comps()  # NEW: Comparable company analysis
        self._create_sensitivity()
        self._create_scenarios()
        self._create_dashboard()
        self._create_summary()
        
        # Save
        os.makedirs(os.path.dirname(output_path), exist_ok=True) if os.path.dirname(output_path) else None
        self.wb.save(output_path)
        logger.info(f"Generated: {output_path}")
        return output_path
    
    def _setup_sheet(self, name: str, title: str) -> tuple:
        """Setup a sheet with headers"""
        ws = self.wb.create_sheet(name)
        ws.column_dimensions['A'].width = 3
        ws.column_dimensions['B'].width = 32
        
        ws['B2'] = title
        ws['B2'].style = 'title'
        
        return ws
    
    def _add_year_headers(self, ws, row: int, start_col: int = 3) -> None:
        """Add year headers from historical to forecast"""
        # Historical years
        for i in range(self.hist_years):
            col = get_column_letter(start_col + i)
            year = self.base_year - self.hist_years + i
            ws[f'{col}{row}'] = f"FY{year}"
            ws[f'{col}{row}'].style = 'header'
            ws.column_dimensions[col].width = 13
        
        # Forecast years
        for i in range(self.fcst_years):
            col = get_column_letter(start_col + self.hist_years + i)
            year = self.base_year + i
            ws[f'{col}{row}'] = f"FY{year}E"
            ws[f'{col}{row}'].style = 'header'
            ws.column_dimensions[col].width = 13
    
    def _get_historical_value(self, statement: str, key: str, year_idx: int) -> Optional[float]:
        """Get historical value from data"""
        data = self.data.get(statement, {})
        historical = data.get('historical', [])
        
        if year_idx < len(historical):
            val = historical[year_idx].get(key)
            if val:
                return val / 10000000  # Convert to Crores
        return None
    
    def _create_assumptions(self) -> None:
        """Create assumptions sheet with real company and industry data"""
        ws = self._setup_sheet("Assumptions", "Model Assumptions")
        
        # Extract real data from sources
        company_info = self.data.get('company_info', {})
        model_assumptions = self.data.get('model_assumptions', {})
        damodaran = self.data.get('damodaran', {})
        real_financials = self.data.get('real_financials', {})
        income_stmt = self.data.get('income_statement', {})
        
        # Get real values with fallbacks
        # Shares outstanding from Yahoo Finance
        shares = company_info.get('shares_outstanding', 0) or 0
        if not shares and company_info.get('sharesOutstanding'):
            shares = (company_info.get('sharesOutstanding', 0) or 0) / 10000000  # Convert to Cr
        shares = shares if (shares and shares > 0) else model_assumptions.get('shares_outstanding', 10.0) or 10.0
        
        # Beta from Yahoo or Damodaran
        beta = company_info.get('beta', 0)
        if not beta or beta == 0:
            beta = self.industry.get('industry_beta', 1.0)
        if not beta or beta == 0:
            beta = damodaran.get('beta', {}).get('levered_beta', 1.0)
        
        # WACC components from Damodaran
        wacc_data = damodaran.get('wacc', {})
        cost_of_equity = wacc_data.get('cost_of_equity', 0.14)
        cost_of_debt = wacc_data.get('cost_of_debt', 0.09)
        debt_ratio = wacc_data.get('debt_ratio', 0.25)
        
        # Margins - Priority: Yahoo Finance > Screener > Damodaran industry avg
        margins = damodaran.get('margins', {})
        actual_financials = self.data.get('real_financials', {})
        
        # Use Yahoo Finance margins if available (most accurate)
        gross_margin = company_info.get('gross_margin', 0) or actual_financials.get('gross_margin', margins.get('gross_margin', 0.35))
        ebitda_margin = company_info.get('ebitda_margin', 0) or actual_financials.get('ebitda_margin', margins.get('ebitda_margin', 0.20))
        net_margin = company_info.get('profit_margin', 0) or actual_financials.get('net_margin', margins.get('net_margin', 0.10))
        operating_margin = company_info.get('operating_margin', 0) or actual_financials.get('operating_margin', margins.get('operating_margin', 0.15))
        
        # Growth from Yahoo Finance or Damodaran
        growth_data = damodaran.get('growth', {})
        revenue_growth = company_info.get('revenue_growth', 0) or growth_data.get('expected_growth', 0.10)
        
        # Capex from Damodaran
        capex_data = damodaran.get('capex', {})
        capex_pct = capex_data.get('capex_to_sales', 0.05)
        da_pct = capex_pct / capex_data.get('capex_to_depreciation', 1.5) if capex_data.get('capex_to_depreciation') else 0.04
        
        # ERP from Damodaran
        erp_data = damodaran.get('erp', {})
        risk_free_rate = erp_data.get('risk_free_rate', 0.07)
        equity_risk_premium = erp_data.get('total_erp', 0.055)
        
        # Tax rate from Damodaran
        tax_rate = damodaran.get('tax_rate', 0.25)
        if not tax_rate:
            tax_rate = model_assumptions.get('tax_rate', 0.25)
        
        # Data source note
        data_source = self.data.get('data_source', 'Yahoo Finance + Screener.in + Damodaran')
        
        ws['B4'] = f"Data Sources: {data_source}"
        ws['B4'].font = Font(italic=True, color='666666')
        ws['B5'] = "Yellow cells are inputs - modify to update projections"
        ws['B5'].font = Font(italic=True, color='1F4E79')
        
        row = 7
        
        # Define all assumptions with REAL values from data
        assumptions = [
            ("GROWTH ASSUMPTIONS", None, None, None),
            ("Revenue Growth Rate", revenue_growth, "percent", f"From Damodaran industry data"),
            ("Volume Growth", revenue_growth * 0.6, "percent", "Estimated 60% of revenue growth"),
            ("Price Inflation", revenue_growth * 0.4, "percent", "Estimated 40% of revenue growth"),
            ("", None, None, None),
            ("MARGIN ASSUMPTIONS (Real Data)", None, None, None),
            ("Gross Margin", gross_margin, "percent", "From company financials"),
            ("EBITDA Margin", ebitda_margin, "percent", "From company financials"),
            ("Operating Margin", operating_margin, "percent", "EBIT/Revenue"),
            ("Net Margin", net_margin, "percent", "From company financials"),
            ("SG&A as % of Revenue", max(gross_margin - ebitda_margin - 0.02, 0.05), "percent", "Calculated from margins"),
            ("", None, None, None),
            ("WORKING CAPITAL", None, None, None),
            ("Receivable Days", 45, "days", "DSO - industry standard"),
            ("Inventory Days", 30, "days", "DIO - industry standard"),
            ("Payable Days", 60, "days", "DPO - industry standard"),
            ("", None, None, None),
            ("CAPEX & D&A (Damodaran)", None, None, None),
            ("Capex % of Revenue", capex_pct, "percent", f"Damodaran industry data"),
            ("D&A % of Gross PPE", da_pct, "percent", f"Derived from capex/depreciation ratio"),
            ("", None, None, None),
            ("DEBT ASSUMPTIONS", None, None, None),
            ("Cost of Debt", cost_of_debt, "percent", f"Damodaran industry WACC"),
            ("Debt/Equity Ratio", debt_ratio / (1 - debt_ratio) if debt_ratio < 1 else 0.5, "ratio", f"From Damodaran"),
            ("", None, None, None),
            ("VALUATION INPUTS (Damodaran)", None, None, None),
            ("Risk-free Rate", risk_free_rate, "percent", "India 10Y G-Sec from Damodaran"),
            ("Equity Risk Premium", equity_risk_premium, "percent", f"India ERP from Damodaran"),
            ("Beta", beta, "ratio", f"From Yahoo Finance/Damodaran"),
            ("Cost of Equity", risk_free_rate + beta * equity_risk_premium, "percent", "CAPM: Rf + β × ERP"),
            ("Terminal Growth Rate", 0.04, "percent", "Conservative long-term GDP growth"),
            ("Tax Rate", tax_rate, "percent", "India corporate tax from Damodaran"),
            ("", None, None, None),
            ("COMPANY DATA (Screener.in + Yahoo)", None, None, None),
            ("Shares Outstanding (Cr)", shares, "number", "From Yahoo Finance"),
            ("Current Price (₹)", company_info.get('current_price', company_info.get('currentPrice', 0)), "currency", "Live market price"),
            ("Market Cap (₹ Cr)", company_info.get('market_cap', 0), "number", "Market capitalization"),
            ("", None, None, None),
            ("KEY RATIOS (Screener.in Real)", None, None, None),
            ("Stock P/E", company_info.get('stock_p/e', company_info.get('pe_ratio', 0)), "ratio", "From Screener.in"),
            ("Price to Book", company_info.get('pb_ratio', company_info.get('price_to_book', 0)), "ratio", "From Screener.in"),
            ("ROCE (%)", company_info.get('roce', 0), "percent", "Return on Capital Employed - Screener.in"),
            ("ROE (%)", company_info.get('roe', company_info.get('return_on_equity', 0)), "percent", "Return on Equity - Screener.in"),
            ("Book Value (₹)", company_info.get('book_value', 0), "currency", "From Screener.in"),
            ("Face Value (₹)", company_info.get('face_value', 10), "currency", "From Screener.in"),
            ("Dividend Yield (%)", company_info.get('dividend_yield', 0), "percent", "From Yahoo Finance"),
        ]
        
        for item in assumptions:
            name, value, unit, desc = item
            
            if not name:
                row += 1
                continue
            
            if value is None:  # Section header
                ws[f'B{row}'] = name
                ws[f'B{row}'].style = 'subheader'
                ws.merge_cells(f'B{row}:E{row}')
            else:
                ws[f'B{row}'] = name
                ws[f'B{row}'].style = 'label'
                
                ws[f'C{row}'] = value
                ws[f'C{row}'].style = 'input_cell'
                
                if unit == 'percent':
                    ws[f'C{row}'].number_format = self.FORMATS['percent']
                elif unit == 'ratio':
                    ws[f'C{row}'].number_format = self.FORMATS['ratio']
                elif unit == 'currency':
                    ws[f'C{row}'].number_format = self.FORMATS['currency']
                else:
                    ws[f'C{row}'].number_format = self.FORMATS['number']
                
                ws[f'D{row}'] = unit if unit else ""
                ws[f'E{row}'] = desc if desc else ""
                ws[f'E{row}'].font = Font(italic=True, size=9, color='666666')
                
                # Create named range (sanitize name)
                range_name = name.replace(' ', '_').replace('%', 'Pct').replace('/', '_').replace('&', 'And').replace('(', '').replace(')', '')
                try:
                    self.wb.create_named_range(range_name, ws, f'$C${row}')
                    self.row_map[name] = row
                except:
                    pass
            
            row += 1
        
        # Store assumption rows for formula references (updated row numbers)
        self.assum_rows = {
            'rev_growth': 8, 'gross_margin': 14, 'ebitda_margin': 15, 'operating_margin': 16,
            'net_margin': 17, 'sga_pct': 18,
            'recv_days': 21, 'inv_days': 22, 'pay_days': 23,
            'capex_pct': 26, 'da_pct': 27,
            'cost_of_debt': 30, 'debt_equity': 31,
            'risk_free': 34, 'erp': 35, 'beta': 36, 'cost_of_equity': 37,
            'terminal_growth': 38, 'tax_rate': 39,
            'shares': 42, 'current_price': 43, 'market_cap': 44
        }
        
        ws.column_dimensions['C'].width = 14
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 32

    
    def _create_income_statement(self) -> None:
        """Create income statement with linked formulas"""
        ws = self._setup_sheet("Income_Statement", "Income Statement (₹ Crores)")
        
        row = 4
        self._add_year_headers(ws, row)
        row += 2
        
        # Get base year data from multiple sources
        income = self.data.get('income_statement', {})
        real_financials = self.data.get('real_financials', {})
        company_info = self.data.get('company_info', {})
        
        # Priority: Screener real_financials > income_statement > calculated from market cap
        base_revenue = real_financials.get('revenue', 0) or 0
        if not base_revenue:
            base_revenue = income.get('revenue', 0) or income.get('totalRevenue', 0) or 0
        if base_revenue and base_revenue is not None:
            # Convert if not already in Crores (likely in raw value from API)
            if base_revenue > 1000000000:  # If > 100 Cr in absolute value
                base_revenue = base_revenue / 10000000  # Convert to Crores
        else:
            # Last resort: estimate from market cap and industry P/S ratio
            market_cap = company_info.get('market_cap', 0) or 0
            if market_cap and market_cap > 0:
                base_revenue = market_cap / 3  # Assume ~3x P/S ratio as estimate
            else:
                base_revenue = 10000  # Final fallback
        
        # Get real EBITDA if available
        base_ebitda = real_financials.get('ebitda', 0) or 0
        if not base_ebitda:
            ebitda_val = income.get('ebitda', 0) or 0
            if ebitda_val and ebitda_val > 1000000000:
                base_ebitda = ebitda_val / 10000000
            else:
                base_ebitda = ebitda_val or 0
        
        # Get real net income if available
        base_net_income = real_financials.get('net_income', 0) or 0
        if not base_net_income:
            ni_val = income.get('netIncome', 0) or 0
            if ni_val and ni_val > 1000000000:
                base_net_income = ni_val / 10000000
            else:
                base_net_income = ni_val or 0
        
        # Store for use in formulas
        self.base_financials = {
            'revenue': base_revenue,
            'ebitda': base_ebitda,
            'net_income': base_net_income,
        }
        
        # Row tracking for formula references
        rows = {}
        
        # Line items
        items = [
            ("Revenue", "revenue", True),
            ("Growth %", None, False),
            ("", None, False),
            ("Cost of Goods Sold", "cogs", False),
            ("Gross Profit", "gross", True),
            ("Gross Margin %", None, False),
            ("", None, False),
            ("SG&A Expenses", "sga", False),
            ("Other Operating Expenses", "other_opex", False),
            ("EBITDA", "ebitda", True),
            ("EBITDA Margin %", None, False),
            ("", None, False),
            ("Depreciation & Amortization", "da", False),
            ("EBIT", "ebit", True),
            ("EBIT Margin %", None, False),
            ("", None, False),
            ("Interest Expense", "interest", False),
            ("Pre-tax Income", "pbt", True),
            ("", None, False),
            ("Tax Expense", "tax", False),
            ("Net Income", "net_income", True),
            ("Net Margin %", None, False),
        ]
        
        for item_name, key, is_bold in items:
            if not item_name:
                row += 1
                continue
            
            ws[f'B{row}'] = item_name
            ws[f'B{row}'].style = 'label'
            if is_bold:
                ws[f'B{row}'].font = Font(bold=True)
            
            if key:
                rows[key] = row
            
            # Fill in values
            total_years = self.hist_years + self.fcst_years
            
            for i in range(total_years):
                col = get_column_letter(3 + i)
                prev_col = get_column_letter(2 + i) if i > 0 else None
                
                is_forecast = i >= self.hist_years
                
                if key == 'revenue':
                    if i == 0:
                        ws[f'{col}{row}'] = base_revenue
                    else:
                        a_row = self.assum_rows['rev_growth']
                        ws[f'{col}{row}'] = f"=IFERROR({prev_col}{row}*(1+Assumptions!$C${a_row}),0)"
                    ws[f'{col}{row}'].number_format = self.FORMATS['number']
                    
                elif item_name == 'Growth %':
                    if i > 0:
                        ws[f'{col}{row}'] = f"=IFERROR({col}{row-1}/{prev_col}{row-1}-1,0)"
                    ws[f'{col}{row}'].number_format = self.FORMATS['percent']
                    
                elif key == 'cogs':
                    a_row = self.assum_rows['gross_margin']
                    ws[f'{col}{row}'] = f"=IFERROR(-{col}{rows['revenue']}*(1-Assumptions!$C${a_row}),0)"
                    ws[f'{col}{row}'].number_format = self.FORMATS['number']
                    
                elif key == 'gross':
                    ws[f'{col}{row}'] = f"={col}{rows['revenue']}+{col}{rows['cogs']}"
                    ws[f'{col}{row}'].number_format = self.FORMATS['number']
                    
                elif item_name == 'Gross Margin %':
                    ws[f'{col}{row}'] = f"=IFERROR({col}{row-1}/{col}{rows['revenue']},0)"
                    ws[f'{col}{row}'].number_format = self.FORMATS['percent']
                    
                elif key == 'sga':
                    a_row = self.assum_rows['sga_pct']
                    ws[f'{col}{row}'] = f"=IFERROR(-{col}{rows['revenue']}*Assumptions!$C${a_row},0)"
                    ws[f'{col}{row}'].number_format = self.FORMATS['number']
                    
                elif key == 'other_opex':
                    ws[f'{col}{row}'] = f"=-{col}{rows['revenue']}*0.02"
                    ws[f'{col}{row}'].number_format = self.FORMATS['number']
                    
                elif key == 'ebitda':
                    ws[f'{col}{row}'] = f"={col}{rows['gross']}+{col}{rows['sga']}+{col}{rows['other_opex']}"
                    ws[f'{col}{row}'].number_format = self.FORMATS['number']
                    ws[f'{col}{row}'].style = 'output_cell'
                    
                elif item_name == 'EBITDA Margin %':
                    ws[f'{col}{row}'] = f"=IFERROR({col}{row-1}/{col}{rows['revenue']},0)"
                    ws[f'{col}{row}'].number_format = self.FORMATS['percent']
                    
                elif key == 'da':
                    a_row = self.assum_rows['da_pct']
                    ws[f'{col}{row}'] = f"=IFERROR(-{col}{rows['revenue']}*Assumptions!$C${a_row},0)"
                    ws[f'{col}{row}'].number_format = self.FORMATS['number']
                    
                elif key == 'ebit':
                    ws[f'{col}{row}'] = f"={col}{rows['ebitda']}+{col}{rows['da']}"
                    ws[f'{col}{row}'].number_format = self.FORMATS['number']
                    
                elif item_name == 'EBIT Margin %':
                    ws[f'{col}{row}'] = f"=IFERROR({col}{row-1}/{col}{rows['revenue']},0)"
                    ws[f'{col}{row}'].number_format = self.FORMATS['percent']
                    
                elif key == 'interest':
                    # Link to balance sheet for debt balance
                    ws[f'{col}{row}'] = f"=-{col}{rows['revenue']}*0.02"  # Placeholder - will link to BS
                    ws[f'{col}{row}'].number_format = self.FORMATS['number']
                    
                elif key == 'pbt':
                    ws[f'{col}{row}'] = f"={col}{rows['ebit']}+{col}{rows['interest']}"
                    ws[f'{col}{row}'].number_format = self.FORMATS['number']
                    
                elif key == 'tax':
                    a_row = self.assum_rows['tax_rate']
                    ws[f'{col}{row}'] = f"=IFERROR(-MAX({col}{rows['pbt']},0)*Assumptions!$C${a_row},0)"
                    ws[f'{col}{row}'].number_format = self.FORMATS['number']
                    
                elif key == 'net_income':
                    ws[f'{col}{row}'] = f"={col}{rows['pbt']}+{col}{rows['tax']}"
                    ws[f'{col}{row}'].number_format = self.FORMATS['number']
                    ws[f'{col}{row}'].style = 'output_cell'
                    
                elif item_name == 'Net Margin %':
                    ws[f'{col}{row}'] = f"=IFERROR({col}{row-1}/{col}{rows['revenue']},0)"
                    ws[f'{col}{row}'].number_format = self.FORMATS['percent']
            
            row += 1
        
        self.row_map['is'] = rows
    
    def _create_balance_sheet(self) -> None:
        """Create balance sheet with linked formulas"""
        ws = self._setup_sheet("Balance_Sheet", "Balance Sheet (₹ Crores)")
        
        row = 4
        self._add_year_headers(ws, row)
        row += 2
        
        rows = {}
        is_rows = self.row_map.get('is', {})
        
        items = [
            ("ASSETS", None, "header"),
            ("Cash & Equivalents", "cash", "asset"),
            ("Trade Receivables", "ar", "asset"),
            ("Inventory", "inv", "asset"),
            ("Other Current Assets", "other_ca", "asset"),
            ("Total Current Assets", "tca", "total"),
            ("", None, None),
            ("Gross PP&E", "ppe_gross", "asset"),
            ("Accumulated Depreciation", "accum_dep", "asset"),
            ("Net PP&E", "ppe_net", "total"),
            ("Other Non-Current Assets", "other_nca", "asset"),
            ("Total Assets", "ta", "total"),
            ("", None, None),
            ("LIABILITIES", None, "header"),
            ("Trade Payables", "ap", "liability"),
            ("Accrued Expenses", "accrued", "liability"),
            ("Short-term Debt", "st_debt", "liability"),
            ("Total Current Liabilities", "tcl", "total"),
            ("", None, None),
            ("Long-term Debt", "lt_debt", "liability"),
            ("Other Non-Current Liabilities", "other_ncl", "liability"),
            ("Total Liabilities", "tl", "total"),
            ("", None, None),
            ("EQUITY", None, "header"),
            ("Share Capital", "share_cap", "equity"),
            ("Retained Earnings", "retained", "equity"),
            ("Total Equity", "te", "total"),
            ("", None, None),
            ("Total Liabilities & Equity", "tle", "total"),
            ("Balance Check (should be 0)", "check", "check"),
        ]
        
        # Starting values from real data
        company_info = self.data.get('company_info', {})
        balance_sheet = self.data.get('balance_sheet', {})
        real_financials = getattr(self, 'base_financials', {})
        
        # Get real total assets if available
        base_ta = balance_sheet.get('totalAssets', 0) or 0
        if base_ta and base_ta > 1000000000:
            base_ta = base_ta / 10000000  # Convert to Crores
        
        if not base_ta or base_ta == 0:
            # Estimate from market cap (assets ~ 1.5x market cap for typical company)
            market_cap = company_info.get('market_cap', 0) or 0
            if market_cap and market_cap > 0:
                base_ta = market_cap * 1.5
            else:
                # Or from revenue (asset turnover ratio ~0.8)
                base_revenue = real_financials.get('revenue', 10000) or 10000
                base_ta = base_revenue * 1.25  # Typical asset turnover
        
        for item_name, key, item_type in items:
            if not item_name:
                row += 1
                continue
            
            ws[f'B{row}'] = item_name
            
            if item_type == "header":
                ws[f'B{row}'].style = 'subheader'
            else:
                ws[f'B{row}'].style = 'label'
                if item_type == "total":
                    ws[f'B{row}'].font = Font(bold=True)
            
            if key:
                rows[key] = row
            
            total_years = self.hist_years + self.fcst_years
            
            for i in range(total_years):
                col = get_column_letter(3 + i)
                prev_col = get_column_letter(2 + i) if i > 0 else None
                is_col = col  # Same column in IS
                
                if key == 'cash':
                    if i == 0:
                        ws[f'{col}{row}'] = base_ta * 0.1
                    else:
                        # Cash = prior cash + net income - capex + depreciation - working capital change
                        ws[f'{col}{row}'] = f"={prev_col}{row}+Cash_Flow!{col}28"
                    ws[f'{col}{row}'].number_format = self.FORMATS['number']
                    
                elif key == 'ar':
                    a_row = self.assum_rows['recv_days']
                    ws[f'{col}{row}'] = f"=IFERROR(Income_Statement!{col}6*Assumptions!$C${a_row}/365,0)"
                    ws[f'{col}{row}'].number_format = self.FORMATS['number']
                    
                elif key == 'inv':
                    a_row = self.assum_rows['inv_days']
                    ws[f'{col}{row}'] = f"=IFERROR(ABS(Income_Statement!{col}9)*Assumptions!$C${a_row}/365,0)"
                    ws[f'{col}{row}'].number_format = self.FORMATS['number']
                    
                elif key == 'other_ca':
                    if i == 0:
                        ws[f'{col}{row}'] = base_ta * 0.05
                    else:
                        ws[f'{col}{row}'] = f"={prev_col}{row}*1.02"
                    ws[f'{col}{row}'].number_format = self.FORMATS['number']
                    
                elif key == 'tca':
                    ws[f'{col}{row}'] = f"=SUM({col}{rows['cash']}:{col}{rows['other_ca']})"
                    ws[f'{col}{row}'].number_format = self.FORMATS['number']
                    
                elif key == 'ppe_gross':
                    if i == 0:
                        ws[f'{col}{row}'] = base_ta * 0.6
                    else:
                        a_row = self.assum_rows['capex_pct']
                        ws[f'{col}{row}'] = f"=IFERROR({prev_col}{row}+Income_Statement!{col}6*Assumptions!$C${a_row},{prev_col}{row})"
                    ws[f'{col}{row}'].number_format = self.FORMATS['number']
                    
                elif key == 'accum_dep':
                    if i == 0:
                        ws[f'{col}{row}'] = -base_ta * 0.2
                    else:
                        ws[f'{col}{row}'] = f"={prev_col}{row}+Income_Statement!{col}18"
                    ws[f'{col}{row}'].number_format = self.FORMATS['number']
                    
                elif key == 'ppe_net':
                    ws[f'{col}{row}'] = f"={col}{rows['ppe_gross']}+{col}{rows['accum_dep']}"
                    ws[f'{col}{row}'].number_format = self.FORMATS['number']
                    
                elif key == 'other_nca':
                    if i == 0:
                        ws[f'{col}{row}'] = base_ta * 0.1
                    else:
                        ws[f'{col}{row}'] = f"={prev_col}{row}*1.01"
                    ws[f'{col}{row}'].number_format = self.FORMATS['number']
                    
                elif key == 'ta':
                    ws[f'{col}{row}'] = f"={col}{rows['tca']}+{col}{rows['ppe_net']}+{col}{rows['other_nca']}"
                    ws[f'{col}{row}'].number_format = self.FORMATS['number']
                    ws[f'{col}{row}'].style = 'output_cell'
                    
                elif key == 'ap':
                    a_row = self.assum_rows['pay_days']
                    ws[f'{col}{row}'] = f"=IFERROR(ABS(Income_Statement!{col}9)*Assumptions!$C${a_row}/365,0)"
                    ws[f'{col}{row}'].number_format = self.FORMATS['number']
                    
                elif key == 'accrued':
                    if i == 0:
                        ws[f'{col}{row}'] = base_ta * 0.03
                    else:
                        ws[f'{col}{row}'] = f"={prev_col}{row}*1.02"
                    ws[f'{col}{row}'].number_format = self.FORMATS['number']
                    
                elif key == 'st_debt':
                    if i == 0:
                        ws[f'{col}{row}'] = base_ta * 0.05
                    else:
                        ws[f'{col}{row}'] = f"={prev_col}{row}"
                    ws[f'{col}{row}'].number_format = self.FORMATS['number']
                    
                elif key == 'tcl':
                    ws[f'{col}{row}'] = f"={col}{rows['ap']}+{col}{rows['accrued']}+{col}{rows['st_debt']}"
                    ws[f'{col}{row}'].number_format = self.FORMATS['number']
                    
                elif key == 'lt_debt':
                    if i == 0:
                        ws[f'{col}{row}'] = base_ta * 0.25
                    else:
                        ws[f'{col}{row}'] = f"={prev_col}{row}"
                    ws[f'{col}{row}'].number_format = self.FORMATS['number']
                    
                elif key == 'other_ncl':
                    if i == 0:
                        ws[f'{col}{row}'] = base_ta * 0.02
                    else:
                        ws[f'{col}{row}'] = f"={prev_col}{row}*1.01"
                    ws[f'{col}{row}'].number_format = self.FORMATS['number']
                    
                elif key == 'tl':
                    ws[f'{col}{row}'] = f"={col}{rows['tcl']}+{col}{rows['lt_debt']}+{col}{rows['other_ncl']}"
                    ws[f'{col}{row}'].number_format = self.FORMATS['number']
                    
                elif key == 'share_cap':
                    if i == 0:
                        ws[f'{col}{row}'] = base_ta * 0.2
                    else:
                        ws[f'{col}{row}'] = f"={prev_col}{row}"
                    ws[f'{col}{row}'].number_format = self.FORMATS['number']
                    
                elif key == 'retained':
                    if i == 0:
                        ws[f'{col}{row}'] = base_ta * 0.45
                    else:
                        # Retained = prior + net income - dividends
                        ws[f'{col}{row}'] = f"={prev_col}{row}+Income_Statement!{col}26-Income_Statement!{col}26*0.2"
                    ws[f'{col}{row}'].number_format = self.FORMATS['number']
                    
                elif key == 'te':
                    ws[f'{col}{row}'] = f"={col}{rows['share_cap']}+{col}{rows['retained']}"
                    ws[f'{col}{row}'].number_format = self.FORMATS['number']
                    
                elif key == 'tle':
                    ws[f'{col}{row}'] = f"={col}{rows['tl']}+{col}{rows['te']}"
                    ws[f'{col}{row}'].number_format = self.FORMATS['number']
                    ws[f'{col}{row}'].style = 'output_cell'
                    
                elif key == 'check':
                    ws[f'{col}{row}'] = f"=ROUND({col}{rows['ta']}-{col}{rows['tle']},0)"
                    ws[f'{col}{row}'].number_format = self.FORMATS['number']
                    ws[f'{col}{row}'].style = 'output_cell'
            
            row += 1
        
        self.row_map['bs'] = rows
    
    def _create_cash_flow(self) -> None:
        """Create cash flow statement"""
        ws = self._setup_sheet("Cash_Flow", "Cash Flow Statement (₹ Crores)")
        
        row = 4
        self._add_year_headers(ws, row)
        row += 2
        
        rows = {}
        is_rows = self.row_map.get('is', {})
        bs_rows = self.row_map.get('bs', {})
        
        items = [
            ("OPERATING ACTIVITIES", None, "header"),
            ("Net Income", "ni", "ops"),
            ("Add: Depreciation", "dep", "ops"),
            ("Change in Receivables", "chg_ar", "ops"),
            ("Change in Inventory", "chg_inv", "ops"),
            ("Change in Payables", "chg_ap", "ops"),
            ("Change in Other Working Capital", "chg_other", "ops"),
            ("Operating Cash Flow", "ocf", "total"),
            ("", None, None),
            ("INVESTING ACTIVITIES", None, "header"),
            ("Capital Expenditure", "capex", "inv"),
            ("Other Investing", "other_inv", "inv"),
            ("Investing Cash Flow", "icf", "total"),
            ("", None, None),
            ("FINANCING ACTIVITIES", None, "header"),
            ("Dividends Paid", "div", "fin"),
            ("Change in Debt", "chg_debt", "fin"),
            ("Financing Cash Flow", "fcf", "total"),
            ("", None, None),
            ("Net Change in Cash", "net_cash", "total"),
            ("Opening Cash", "open_cash", None),
            ("Closing Cash", "close_cash", "total"),
        ]
        
        for item_name, key, item_type in items:
            if not item_name:
                row += 1
                continue
            
            ws[f'B{row}'] = item_name
            
            if item_type == "header":
                ws[f'B{row}'].style = 'subheader'
            else:
                ws[f'B{row}'].style = 'label'
                if item_type == "total":
                    ws[f'B{row}'].font = Font(bold=True)
            
            if key:
                rows[key] = row
            
            total_years = self.hist_years + self.fcst_years
            
            for i in range(total_years):
                col = get_column_letter(3 + i)
                prev_col = get_column_letter(2 + i) if i > 0 else None
                
                if key == 'ni':
                    ws[f'{col}{row}'] = f"=IFERROR(Income_Statement!{col}26,0)"
                    
                elif key == 'dep':
                    ws[f'{col}{row}'] = f"=IFERROR(-Income_Statement!{col}18,0)"
                    
                elif key == 'chg_ar':
                    if i > 0:
                        ws[f'{col}{row}'] = f"=IFERROR(Balance_Sheet!{prev_col}8-Balance_Sheet!{col}8,0)"
                    else:
                        ws[f'{col}{row}'] = 0
                        
                elif key == 'chg_inv':
                    if i > 0:
                        ws[f'{col}{row}'] = f"=IFERROR(Balance_Sheet!{prev_col}9-Balance_Sheet!{col}9,0)"
                    else:
                        ws[f'{col}{row}'] = 0
                        
                elif key == 'chg_ap':
                    if i > 0:
                        ws[f'{col}{row}'] = f"=IFERROR(Balance_Sheet!{col}20-Balance_Sheet!{prev_col}20,0)"
                    else:
                        ws[f'{col}{row}'] = 0
                        
                elif key == 'chg_other':
                    ws[f'{col}{row}'] = 0
                    
                elif key == 'ocf':
                    ws[f'{col}{row}'] = f"=SUM({col}{rows['ni']}:{col}{rows['chg_other']})"
                    ws[f'{col}{row}'].style = 'output_cell'
                    
                elif key == 'capex':
                    ws[f'{col}{row}'] = f"=IFERROR(-Income_Statement!{col}6*Assumptions!$C$21,0)"
                    
                elif key == 'other_inv':
                    ws[f'{col}{row}'] = 0
                    
                elif key == 'icf':
                    ws[f'{col}{row}'] = f"={col}{rows['capex']}+{col}{rows['other_inv']}"
                    
                elif key == 'div':
                    ws[f'{col}{row}'] = f"=IFERROR(-Income_Statement!{col}26*0.2,0)"
                    
                elif key == 'chg_debt':
                    ws[f'{col}{row}'] = 0
                    
                elif key == 'fcf':
                    ws[f'{col}{row}'] = f"={col}{rows['div']}+{col}{rows['chg_debt']}"
                    
                elif key == 'net_cash':
                    ws[f'{col}{row}'] = f"={col}{rows['ocf']}+{col}{rows['icf']}+{col}{rows['fcf']}"
                    ws[f'{col}{row}'].style = 'output_cell'
                    
                elif key == 'open_cash':
                    if i > 0:
                        ws[f'{col}{row}'] = f"=IFERROR(Balance_Sheet!{prev_col}7,0)"
                    else:
                        ws[f'{col}{row}'] = 2000  # Starting cash
                        
                elif key == 'close_cash':
                    ws[f'{col}{row}'] = f"={col}{rows['open_cash']}+{col}{rows['net_cash']}"
                    ws[f'{col}{row}'].style = 'output_cell'
                
                if key:
                    ws[f'{col}{row}'].number_format = self.FORMATS['number']
            
            row += 1
        
        self.row_map['cf'] = rows
    
    def _create_valuation(self) -> None:
        """Create DCF valuation"""
        ws = self._setup_sheet("Valuation", "DCF Valuation (₹ Crores)")
        
        row = 5
        
        # WACC Section
        ws[f'B{row}'] = "WACC CALCULATION"
        ws[f'B{row}'].style = 'subheader'
        row += 1
        
        wacc_items = [
            ("Risk-free Rate", "=Assumptions!$C$30", "percent"),
            ("Equity Risk Premium", "=Assumptions!$C$31", "percent"),
            ("Beta", "=Assumptions!$C$32", "ratio"),
            ("Cost of Equity", "=C6+C7*C8", "percent"),
            ("", None, None),
            ("Cost of Debt (pre-tax)", "=Assumptions!$C$26", "percent"),
            ("Tax Rate", "=Assumptions!$C$34", "percent"),
            ("Cost of Debt (post-tax)", "=C11*(1-C12)", "percent"),
            ("", None, None),
            ("Target D/E Ratio", "=Assumptions!$C$27", "ratio"),
            ("Weight of Equity", "=1/(1+C15)", "percent"),
            ("Weight of Debt", "=C15/(1+C15)", "percent"),
            ("", None, None),
            ("WACC", "=C16*C9+C17*C13", "percent"),
        ]
        
        for name, formula, fmt in wacc_items:
            if not name:
                row += 1
                continue
            
            ws[f'B{row}'] = name
            ws[f'B{row}'].style = 'label'
            
            if name == "WACC":
                ws[f'B{row}'].font = Font(bold=True)
                ws[f'C{row}'].style = 'output_cell'
            
            if formula:
                ws[f'C{row}'] = formula
                if fmt == "percent":
                    ws[f'C{row}'].number_format = self.FORMATS['percent']
                elif fmt == "ratio":
                    ws[f'C{row}'].number_format = self.FORMATS['ratio']
            
            row += 1
        
        wacc_row = row - 1  # Row where WACC is calculated
        
        # DCF Section
        row += 2
        ws[f'B{row}'] = "DCF VALUATION"
        ws[f'B{row}'].style = 'subheader'
        row += 1
        
        # Year headers
        for i in range(self.fcst_years):
            col = get_column_letter(3 + i)
            ws[f'{col}{row}'] = f"Year {i+1}"
            ws[f'{col}{row}'].style = 'header'
        row += 1
        
        fcff_row = row
        
        # FCFF (EBITDA - Capex - WC change)
        ws[f'B{row}'] = "Free Cash Flow to Firm"
        ws[f'B{row}'].style = 'label'
        for i in range(self.fcst_years):
            col = get_column_letter(3 + i)
            is_col = get_column_letter(3 + self.hist_years + i)
            ws[f'{col}{row}'] = f"=Income_Statement!{is_col}15+Income_Statement!{is_col}18+Cash_Flow!{is_col}19"
            ws[f'{col}{row}'].number_format = self.FORMATS['number']
        row += 1
        
        # Discount Factor
        ws[f'B{row}'] = "Discount Factor"
        ws[f'B{row}'].style = 'label'
        for i in range(self.fcst_years):
            col = get_column_letter(3 + i)
            ws[f'{col}{row}'] = f"=1/(1+$C${wacc_row})^{i+1}"
            ws[f'{col}{row}'].number_format = '0.000'
        row += 1
        
        # PV of FCFF
        pv_row = row
        ws[f'B{row}'] = "Present Value of FCFF"
        ws[f'B{row}'].style = 'label'
        ws[f'B{row}'].font = Font(bold=True)
        for i in range(self.fcst_years):
            col = get_column_letter(3 + i)
            ws[f'{col}{row}'] = f"={col}{fcff_row}*{col}{row-1}"
            ws[f'{col}{row}'].number_format = self.FORMATS['number']
            ws[f'{col}{row}'].style = 'output_cell'
        row += 2
        
        # Terminal Value
        tv_start = row
        tg_row = row  # Terminal growth rate row
        ws[f'B{row}'] = "Terminal Growth Rate"
        ws[f'B{row}'].style = 'label'
        ws[f'C{row}'] = "=Assumptions!$C$33"
        ws[f'C{row}'].number_format = self.FORMATS['percent']
        ws[f'C{row}'].style = 'input_cell'
        row += 1
        
        last_fcff_col = get_column_letter(2 + self.fcst_years)
        tv_row = row  # Terminal value row
        ws[f'B{row}'] = "Terminal Value"
        ws[f'B{row}'].style = 'label'
        ws[f'C{row}'] = f"=IFERROR({last_fcff_col}{fcff_row}*(1+C{tg_row})/($C${wacc_row}-C{tg_row}),0)"
        ws[f'C{row}'].number_format = self.FORMATS['number']
        row += 1
        
        pv_tv_row = row  # PV of terminal value row
        ws[f'B{row}'] = "PV of Terminal Value"
        ws[f'B{row}'].style = 'label'
        ws[f'C{row}'] = f"=C{tv_row}*{last_fcff_col}{pv_row-1}"
        ws[f'C{row}'].number_format = self.FORMATS['number']
        ws[f'C{row}'].style = 'output_cell'
        row += 2
        
        # Valuation Summary
        ws[f'B{row}'] = "VALUATION SUMMARY"
        ws[f'B{row}'].style = 'subheader'
        row += 1
        
        sum_pv_row = row  # Sum of PV of FCFF row
        ev_row = row + 2  # Enterprise Value row
        net_debt_row = row + 3  # Net Debt row
        equity_value_row = row + 4  # Equity Value row
        shares_row = row + 6  # Shares Outstanding row
        share_price_row = row + 7  # Share Price row
        
        summary = [
            ("Sum of PV of FCFF", f"=SUM(C{pv_row}:{last_fcff_col}{pv_row})", "number"),
            ("PV of Terminal Value", f"=C{pv_tv_row}", "number"),
            ("Enterprise Value", f"=C{sum_pv_row}+C{sum_pv_row+1}", "number"),
            ("Less: Net Debt", f"=Balance_Sheet!{get_column_letter(2+self.hist_years+self.fcst_years)}25+Balance_Sheet!{get_column_letter(2+self.hist_years+self.fcst_years)}22-Balance_Sheet!{get_column_letter(2+self.hist_years+self.fcst_years)}7", "number"),
            ("Equity Value", f"=C{ev_row}-C{net_debt_row}", "number"),
            ("", None, None),
            ("Shares Outstanding (Cr)", "=Assumptions!$C$37", "decimal"),
            ("Implied Share Price (₹)", f"=IFERROR(C{equity_value_row}/C{shares_row}*10,0)", "currency"),
        ]
        
        for name, formula, fmt in summary:
            if not name:
                row += 1
                continue
            
            ws[f'B{row}'] = name
            ws[f'B{row}'].style = 'label'
            
            if name in ["Enterprise Value", "Equity Value", "Implied Share Price (₹)"]:
                ws[f'B{row}'].font = Font(bold=True)
                ws[f'C{row}'].style = 'output_cell'
            
            if formula:
                ws[f'C{row}'] = formula
                if fmt == "number":
                    ws[f'C{row}'].number_format = self.FORMATS['number']
                elif fmt == "decimal":
                    ws[f'C{row}'].number_format = self.FORMATS['decimal']
                elif fmt == "currency":
                    ws[f'C{row}'].number_format = self.FORMATS['currency']
            
            row += 1
        
        # Store valuation row references for Summary sheet and Dashboard
        self.row_map['valuation'] = {
            'wacc': wacc_row,
            'pv_fcff': pv_row,
            'pv_tv': pv_tv_row,
            'sum_pv': sum_pv_row,
            'ev': ev_row,
            'net_debt': net_debt_row,
            'equity_value': equity_value_row,
            'shares': shares_row,
            'share_price': share_price_row,
        }
        
        ws.column_dimensions['C'].width = 15
    
    def _create_comps(self) -> None:
        """Create Comparable Company Analysis sheet"""
        from openpyxl.formatting.rule import ColorScaleRule
        
        ws = self._setup_sheet("Comps", f"{self.company_name} - Comparable Company Analysis")
        
        row = 5
        ws[f'B{row}'] = "COMPARABLE COMPANY ANALYSIS"
        ws[f'B{row}'].style = 'subheader'
        ws.merge_cells(f'B{row}:J{row}')
        row += 2
        
        # Headers
        headers = [
            ("Company", 25),
            ("Market Cap", 15),
            ("Revenue", 15),
            ("EBITDA", 15),
            ("EBITDA Margin", 15),
            ("P/E", 12),
            ("EV/EBITDA", 12),
            ("EV/Revenue", 12),
            ("ROE", 12),
        ]
        
        for i, (header, width) in enumerate(headers):
            col = get_column_letter(2 + i)
            ws[f'{col}{row}'] = header
            ws[f'{col}{row}'].style = 'header'
            ws.column_dimensions[col].width = width
        row += 1
        
        # Target company row (from model data)
        ws[f'B{row}'] = f"{self.company_name} (Target)"
        ws[f'B{row}'].font = Font(bold=True)
        ws[f'C{row}'] = "=Valuation!C46/100"  # Market cap
        ws[f'C{row}'].number_format = self.FORMATS['number']
        ws[f'D{row}'] = "=Income_Statement!C6"  # Revenue
        ws[f'D{row}'].number_format = self.FORMATS['number']
        ws[f'E{row}'] = "=Income_Statement!C15"  # EBITDA
        ws[f'E{row}'].number_format = self.FORMATS['number']
        ws[f'F{row}'] = "=IFERROR(E{0}/D{0},0)".format(row)
        ws[f'F{row}'].number_format = self.FORMATS['percent']
        ws[f'G{row}'] = "=IFERROR(Valuation!C46/Income_Statement!C26,0)"  # P/E
        ws[f'G{row}'].number_format = self.FORMATS['decimal']
        ws[f'H{row}'] = "=IFERROR(Valuation!C44/E{0},0)".format(row)  # EV/EBITDA
        ws[f'H{row}'].number_format = self.FORMATS['decimal']
        ws[f'I{row}'] = "=IFERROR(Valuation!C44/D{0},0)".format(row)  # EV/Revenue
        ws[f'I{row}'].number_format = self.FORMATS['decimal']
        ws[f'J{row}'] = "=Assumptions!C14"  # ROE placeholder
        ws[f'J{row}'].number_format = self.FORMATS['percent']
        target_row = row
        row += 1
        
        # Sample peer companies (placeholders for user input)
        peer_data = [
            ("Peer Company 1", 50000, 25000, 5000, 0.20, 15.0, 8.5, 1.8, 0.18),
            ("Peer Company 2", 35000, 18000, 3500, 0.19, 18.0, 9.0, 1.7, 0.15),
            ("Peer Company 3", 75000, 40000, 8500, 0.21, 12.0, 7.5, 1.6, 0.22),
            ("Peer Company 4", 28000, 15000, 2800, 0.19, 20.0, 9.5, 1.8, 0.14),
            ("Peer Company 5", 60000, 32000, 6800, 0.21, 14.0, 8.0, 1.7, 0.19),
        ]
        
        peer_start = row
        for name, mcap, rev, ebitda, margin, pe, ev_ebitda, ev_rev, roe in peer_data:
            ws[f'B{row}'] = name
            ws[f'B{row}'].style = 'input_cell'
            ws[f'C{row}'] = mcap
            ws[f'C{row}'].style = 'input_cell'
            ws[f'C{row}'].number_format = self.FORMATS['number']
            ws[f'D{row}'] = rev
            ws[f'D{row}'].style = 'input_cell'
            ws[f'D{row}'].number_format = self.FORMATS['number']
            ws[f'E{row}'] = ebitda
            ws[f'E{row}'].style = 'input_cell'
            ws[f'E{row}'].number_format = self.FORMATS['number']
            ws[f'F{row}'] = margin
            ws[f'F{row}'].style = 'input_cell'
            ws[f'F{row}'].number_format = self.FORMATS['percent']
            ws[f'G{row}'] = pe
            ws[f'G{row}'].style = 'input_cell'
            ws[f'G{row}'].number_format = self.FORMATS['decimal']
            ws[f'H{row}'] = ev_ebitda
            ws[f'H{row}'].style = 'input_cell'
            ws[f'H{row}'].number_format = self.FORMATS['decimal']
            ws[f'I{row}'] = ev_rev
            ws[f'I{row}'].style = 'input_cell'
            ws[f'I{row}'].number_format = self.FORMATS['decimal']
            ws[f'J{row}'] = roe
            ws[f'J{row}'].style = 'input_cell'
            ws[f'J{row}'].number_format = self.FORMATS['percent']
            row += 1
        peer_end = row - 1
        
        # Summary Statistics
        row += 2
        ws[f'B{row}'] = "PEER STATISTICS"
        ws[f'B{row}'].style = 'subheader'
        ws.merge_cells(f'B{row}:J{row}')
        row += 1
        
        stats = [
            ("Mean", "AVERAGE"),
            ("Median", "MEDIAN"),
            ("25th Percentile", "PERCENTILE"),
            ("75th Percentile", "PERCENTILE"),
        ]
        
        for stat_name, func in stats:
            ws[f'B{row}'] = stat_name
            ws[f'B{row}'].style = 'label'
            
            for i, col in enumerate(['C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']):
                if "PERCENTILE" in func:
                    pct = 0.25 if "25th" in stat_name else 0.75
                    ws[f'{col}{row}'] = f"=IFERROR({func}({col}{peer_start}:{col}{peer_end},{pct}),0)"
                else:
                    ws[f'{col}{row}'] = f"=IFERROR({func}({col}{peer_start}:{col}{peer_end}),0)"
                
                if col == 'F' or col == 'J':
                    ws[f'{col}{row}'].number_format = self.FORMATS['percent']
                elif col in ['G', 'H', 'I']:
                    ws[f'{col}{row}'].number_format = self.FORMATS['decimal']
                else:
                    ws[f'{col}{row}'].number_format = self.FORMATS['number']
            row += 1
        
        # Implied Valuation
        row += 2
        ws[f'B{row}'] = "IMPLIED VALUATION FROM COMPS"
        ws[f'B{row}'].style = 'subheader'
        ws.merge_cells(f'B{row}:E{row}')
        row += 1
        
        ws[f'B{row}'] = "Metric"
        ws[f'C{row}'] = "Low"
        ws[f'D{row}'] = "Mid"
        ws[f'E{row}'] = "High"
        for col in ['B', 'C', 'D', 'E']:
            ws[f'{col}{row}'].style = 'header'
        row += 1
        
        mean_row = peer_end + 4  # where mean was calculated
        p25_row = mean_row + 2
        p75_row = mean_row + 3
        
        # Implied EV from EV/EBITDA
        ws[f'B{row}'] = "EV (from EV/EBITDA)"
        ws[f'C{row}'] = f"=IFERROR(H{p25_row}*E{target_row},0)"
        ws[f'D{row}'] = f"=IFERROR(H{mean_row}*E{target_row},0)"
        ws[f'E{row}'] = f"=IFERROR(H{p75_row}*E{target_row},0)"
        for col in ['C', 'D', 'E']:
            ws[f'{col}{row}'].number_format = self.FORMATS['number']
        row += 1
        
        # Implied EV from EV/Revenue
        ws[f'B{row}'] = "EV (from EV/Revenue)"
        ws[f'C{row}'] = f"=IFERROR(I{p25_row}*D{target_row},0)"
        ws[f'D{row}'] = f"=IFERROR(I{mean_row}*D{target_row},0)"
        ws[f'E{row}'] = f"=IFERROR(I{p75_row}*D{target_row},0)"
        for col in ['C', 'D', 'E']:
            ws[f'{col}{row}'].number_format = self.FORMATS['number']
        row += 1
        
        # Implied Share Price
        row += 1
        ws[f'B{row}'] = "Implied Share Price Range"
        ws[f'B{row}'].font = Font(bold=True)
        shares = "Valuation!C49"  # shares outstanding
        ws[f'C{row}'] = f"=IFERROR(C{row-2}/{shares},0)"
        ws[f'D{row}'] = f"=IFERROR(D{row-2}/{shares},0)"  
        ws[f'E{row}'] = f"=IFERROR(E{row-2}/{shares},0)"
        for col in ['C', 'D', 'E']:
            ws[f'{col}{row}'].number_format = self.FORMATS['currency']
            ws[f'{col}{row}'].style = 'output_cell'
    
    def _create_summary(self) -> None:

        """Create executive summary sheet"""
        ws = self.wb.create_sheet("Summary", 0)
        
        ws.column_dimensions['A'].width = 3
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 5
        ws.column_dimensions['E'].width = 25
        ws.column_dimensions['F'].width = 18
        
        # Title
        ws['B2'] = self.company_name
        ws['B2'].style = 'title'
        ws.merge_cells('B2:F2')
        
        ws['B3'] = "Financial Model Summary"
        ws['B3'].font = Font(size=12, color='666666')
        
        ws['B4'] = f"Generated: {datetime.now().strftime('%d-%b-%Y')}"
        ws['B4'].font = Font(italic=True, size=10)
        
        # Company Info
        row = 7
        ws[f'B{row}'] = "Company Information"
        ws[f'B{row}'].style = 'subheader'
        ws.merge_cells(f'B{row}:C{row}')
        row += 1
        
        info = [
            ("Industry", self.industry.get('industry_name', 'N/A')),
            ("Model Type", self.industry.get('model_type', 'general').replace('_', ' ').title()),
            ("Forecast Period", f"{self.fcst_years} Years"),
        ]
        
        for label, value in info:
            ws[f'B{row}'] = label
            ws[f'B{row}'].style = 'label'
            ws[f'C{row}'] = value
            row += 1
        
        # Key Outputs - Use dynamic row references from valuation sheet
        row += 1
        ws[f'B{row}'] = "Key Outputs"
        ws[f'B{row}'].style = 'subheader'
        ws.merge_cells(f'B{row}:C{row}')
        row += 1
        
        # Get dynamic row references from valuation sheet
        val_rows = self.row_map.get('valuation', {})
        wacc_row = val_rows.get('wacc', 19)
        ev_row = val_rows.get('ev', 44)
        equity_value_row = val_rows.get('equity_value', 46)
        share_price_row = val_rows.get('share_price', 50)
        
        outputs = [
            ("Implied Share Price", f"=Valuation!C{share_price_row}", "currency"),
            ("Enterprise Value", f"=Valuation!C{ev_row}", "number"),
            ("Equity Value", f"=Valuation!C{equity_value_row}", "number"),
            ("WACC", f"=Valuation!C{wacc_row}", "percent"),
        ]
        
        for label, formula, fmt in outputs:
            ws[f'B{row}'] = label
            ws[f'B{row}'].style = 'label'
            ws[f'C{row}'] = formula
            ws[f'C{row}'].style = 'output_cell'
            if fmt == "currency":
                ws[f'C{row}'].number_format = self.FORMATS['currency']
            elif fmt == "percent":
                ws[f'C{row}'].number_format = self.FORMATS['percent']
            else:
                ws[f'C{row}'].number_format = self.FORMATS['number']
            row += 1
        
        # Navigation
        row = 7
        ws[f'E{row}'] = "Model Navigation"
        ws[f'E{row}'].style = 'subheader'
        ws.merge_cells(f'E{row}:F{row}')
        row += 1
        
        sheets = ['Summary', 'Assumptions', 'Income_Statement', 'Balance_Sheet', 'Cash_Flow', 'Valuation', 'Sensitivity', 'Scenarios', 'Dashboard']
        for sheet in sheets:
            ws[f'E{row}'] = sheet.replace('_', ' ')
            ws[f'E{row}'].hyperlink = f"#'{sheet}'!A1"
            ws[f'E{row}'].font = Font(color='0563C1', underline='single')
            row += 1
    
    def _create_sensitivity(self) -> None:
        """Create sensitivity analysis table"""
        ws = self._setup_sheet("Sensitivity", "Sensitivity Analysis")
        
        row = 5
        ws[f'B{row}'] = "WACC vs Terminal Growth Sensitivity"
        ws[f'B{row}'].style = 'subheader'
        ws.merge_cells(f'B{row}:H{row}')
        row += 2
        
        # Terminal growth rates (columns)
        tg_rates = [0.02, 0.025, 0.03, 0.035, 0.04, 0.045, 0.05]
        wacc_rates = [0.08, 0.09, 0.10, 0.11, 0.12, 0.13, 0.14]
        
        # Column headers
        ws[f'B{row}'] = "WACC \\ TG"
        ws[f'B{row}'].style = 'header'
        for i, tg in enumerate(tg_rates):
            col = get_column_letter(3 + i)
            ws[f'{col}{row}'] = tg
            ws[f'{col}{row}'].style = 'header'
            ws[f'{col}{row}'].number_format = self.FORMATS['percent']
            ws.column_dimensions[col].width = 12
        row += 1
        
        # WACC rows with sensitivity formulas
        for wacc in wacc_rates:
            ws[f'B{row}'] = wacc
            ws[f'B{row}'].style = 'header'
            ws[f'B{row}'].number_format = self.FORMATS['percent']
            
            for i, tg in enumerate(tg_rates):
                col = get_column_letter(3 + i)
                # Simplified sensitivity formula
                # Equity Value = FCFF * (1+g) / (WACC - g)
                fcff_ref = f"Valuation!C{27}"  # Last year FCFF approx
                ws[f'{col}{row}'] = f"=IFERROR(1000*(1+{tg})/({wacc}-{tg}),0)"
                ws[f'{col}{row}'].number_format = self.FORMATS['number']
                
                # Highlight center cell
                if wacc == 0.11 and tg == 0.035:
                    ws[f'{col}{row}'].style = 'output_cell'
            
            row += 1
        
        # Add color scale formatting to first table
        from openpyxl.formatting.rule import ColorScaleRule
        first_table_end = row - 1
        first_table_start = first_table_end - len(wacc_rates) + 1
        color_range = f"C{first_table_start}:I{first_table_end}"
        ws.conditional_formatting.add(color_range,
            ColorScaleRule(
                start_type='min', start_color='F8696B',  # Red
                mid_type='percentile', mid_value=50, mid_color='FFEB84',  # Yellow
                end_type='max', end_color='63BE7B'  # Green
            )
        )
        
        # Revenue Growth vs EBITDA Margin
        row += 3
        ws[f'B{row}'] = "Revenue Growth vs EBITDA Margin Impact on EV"
        ws[f'B{row}'].style = 'subheader'
        ws.merge_cells(f'B{row}:H{row}')
        row += 2
        
        rev_growth = [0.05, 0.08, 0.10, 0.12, 0.15, 0.18, 0.20]
        ebitda_margins = [0.15, 0.20, 0.25, 0.30, 0.35]
        
        ws[f'B{row}'] = "Growth \\ Margin"
        ws[f'B{row}'].style = 'header'
        for i, margin in enumerate(ebitda_margins):
            col = get_column_letter(3 + i)
            ws[f'{col}{row}'] = margin
            ws[f'{col}{row}'].style = 'header'
            ws[f'{col}{row}'].number_format = self.FORMATS['percent']
        second_header_row = row
        row += 1
        
        second_start_row = row
        for growth in rev_growth:
            ws[f'B{row}'] = growth
            ws[f'B{row}'].style = 'header'
            ws[f'B{row}'].number_format = self.FORMATS['percent']
            
            for i, margin in enumerate(ebitda_margins):
                col = get_column_letter(3 + i)
                # Simple EV proxy = Revenue * (1+g)^5 * margin * 8 (EV/EBITDA multiple)
                ws[f'{col}{row}'] = f"=10000*((1+{growth})^5)*{margin}*8"
                ws[f'{col}{row}'].number_format = self.FORMATS['number']
            
            row += 1
        
        # Add color scale to second table
        second_table_end = row - 1
        color_range2 = f"C{second_start_row}:G{second_table_end}"
        ws.conditional_formatting.add(color_range2,
            ColorScaleRule(
                start_type='min', start_color='F8696B',
                mid_type='percentile', mid_value=50, mid_color='FFEB84',
                end_type='max', end_color='63BE7B'
            )
        )
    

    def _create_scenarios(self) -> None:
        """Create scenario analysis (Base/Bull/Bear)"""
        ws = self._setup_sheet("Scenarios", "Scenario Analysis")
        
        row = 5
        # Headers
        ws[f'B{row}'] = "Scenario"
        ws[f'C{row}'] = "Bear"
        ws[f'D{row}'] = "Base"
        ws[f'E{row}'] = "Bull"
        for col in ['B', 'C', 'D', 'E']:
            ws[f'{col}{row}'].style = 'header'
            ws.column_dimensions[col].width = 18
        row += 1
        
        scenarios = [
            ("Revenue Growth", 0.05, 0.10, 0.15, "percent"),
            ("EBITDA Margin", 0.18, 0.25, 0.32, "percent"),
            ("Terminal Growth", 0.02, 0.04, 0.05, "percent"),
            ("WACC", 0.14, 0.11, 0.09, "percent"),
            ("", None, None, None, None),
            ("ASSUMPTIONS", None, None, None, None),
            ("Capex % Rev", 0.08, 0.05, 0.04, "percent"),
            ("Working Capital Days", 60, 45, 35, "number"),
            ("D/E Ratio", 0.8, 0.5, 0.3, "ratio"),
            ("Tax Rate", 0.30, 0.25, 0.22, "percent"),
            ("", None, None, None, None),
            ("OUTPUTS", None, None, None, None),
            ("5Y Revenue CAGR", "=C6", "=D6", "=E6", "percent"),
            ("Exit EBITDA", "=10000*(1+C6)^5*C7", "=10000*(1+D6)^5*D7", "=10000*(1+E6)^5*E7", "number"),
            ("Enterprise Value", "=C19*6", "=D19*8", "=E19*10", "number"),
            ("Equity Value", "=C20-3000", "=D20-2500", "=E20-2000", "number"),
            ("Implied Share Price", "=C21/10", "=D21/10", "=E21/10", "currency"),
            ("", None, None, None, None),
            ("IRR Analysis", None, None, None, None),
            ("Entry Price", 500, 500, 500, "currency"),
            ("Exit Price", "=C22", "=D22", "=E22", "currency"),
            ("5Y Return", "=(C25/C24)^0.2-1", "=(D25/D24)^0.2-1", "=(E25/E24)^0.2-1", "percent"),
        ]
        
        for item in scenarios:
            name, bear, base, bull, fmt = item
            
            if not name:
                row += 1
                continue
            
            if bear is None:  # Section header
                ws[f'B{row}'] = name
                ws[f'B{row}'].style = 'subheader'
                ws.merge_cells(f'B{row}:E{row}')
                row += 1
                continue
            
            ws[f'B{row}'] = name
            ws[f'B{row}'].style = 'label'
            
            for col, val in [('C', bear), ('D', base), ('E', bull)]:
                ws[f'{col}{row}'] = val
                if col == 'D':
                    ws[f'{col}{row}'].style = 'input_cell'
                
                if fmt == "percent":
                    ws[f'{col}{row}'].number_format = self.FORMATS['percent']
                elif fmt == "ratio":
                    ws[f'{col}{row}'].number_format = self.FORMATS['ratio']
                elif fmt == "currency":
                    ws[f'{col}{row}'].number_format = self.FORMATS['currency']
                else:
                    ws[f'{col}{row}'].number_format = self.FORMATS['number']
            
            row += 1
    
    def _create_dashboard(self) -> None:
        """Create visual dashboard with charts"""
        from openpyxl.chart import LineChart, BarChart, Reference
        from openpyxl.chart.series import SeriesLabel
        
        ws = self._setup_sheet("Dashboard", f"{self.company_name} - Financial Dashboard")
        
        # First create data for charts
        row = 40  # Put data below where charts will be
        data_start = row
        
        # Revenue data
        ws[f'B{row}'] = "Chart Data"
        ws[f'B{row}'].style = 'subheader'
        row += 1
        
        ws[f'B{row}'] = "Year"
        total_years = self.hist_years + self.fcst_years
        for i in range(total_years):
            col = get_column_letter(3 + i)
            year = self.base_year - self.hist_years + i
            ws[f'{col}{row}'] = f"FY{year}"
        row += 1
        
        # Revenue
        ws[f'B{row}'] = "Revenue"
        for i in range(total_years):
            col = get_column_letter(3 + i)
            is_col = col
            ws[f'{col}{row}'] = f"=Income_Statement!{is_col}6"
            ws[f'{col}{row}'].number_format = self.FORMATS['number']
        rev_row = row
        row += 1
        
        # EBITDA
        ws[f'B{row}'] = "EBITDA"
        for i in range(total_years):
            col = get_column_letter(3 + i)
            ws[f'{col}{row}'] = f"=Income_Statement!{col}15"
            ws[f'{col}{row}'].number_format = self.FORMATS['number']
        ebitda_row = row
        row += 1
        
        # Net Income
        ws[f'B{row}'] = "Net Income"
        for i in range(total_years):
            col = get_column_letter(3 + i)
            ws[f'{col}{row}'] = f"=Income_Statement!{col}26"
            ws[f'{col}{row}'].number_format = self.FORMATS['number']
        ni_row = row
        row += 1
        
        # EBITDA Margin
        ws[f'B{row}'] = "EBITDA Margin %"
        for i in range(total_years):
            col = get_column_letter(3 + i)
            ws[f'{col}{row}'] = f"=IFERROR({col}{ebitda_row}/{col}{rev_row},0)"
            ws[f'{col}{row}'].number_format = self.FORMATS['percent']
        margin_row = row
        row += 1
        
        # Create Revenue & EBITDA Bar Chart
        chart1 = BarChart()
        chart1.type = "col"
        chart1.grouping = "clustered"
        chart1.title = "Revenue & EBITDA (₹ Crores)"
        chart1.style = 10
        
        data = Reference(ws, min_col=3, min_row=rev_row, max_col=2+total_years, max_row=ebitda_row)
        cats = Reference(ws, min_col=3, min_row=data_start+1, max_col=2+total_years)
        chart1.add_data(data, from_rows=True, titles_from_data=False)
        chart1.set_categories(cats)
        # Set series titles properly using SeriesLabel
        if len(chart1.series) > 0:
            chart1.series[0].tx = SeriesLabel(v="Revenue")
        if len(chart1.series) > 1:
            chart1.series[1].tx = SeriesLabel(v="EBITDA")
        chart1.shape = 4
        chart1.width = 18
        chart1.height = 10
        
        ws.add_chart(chart1, "B5")
        
        # Create Margin Line Chart
        chart2 = LineChart()
        chart2.title = "EBITDA Margin Trend"
        chart2.style = 10
        chart2.y_axis.title = "Margin %"
        
        data2 = Reference(ws, min_col=3, min_row=margin_row, max_col=2+total_years)
        chart2.add_data(data2, from_rows=True)
        chart2.set_categories(cats)
        if len(chart2.series) > 0:
            chart2.series[0].tx = SeriesLabel(v="EBITDA Margin")
        chart2.width = 12
        chart2.height = 8
        
        ws.add_chart(chart2, "L5")
        
        # Net Income Line Chart
        chart3 = LineChart()
        chart3.title = "Net Income Trend (₹ Crores)"
        chart3.style = 12
        
        data3 = Reference(ws, min_col=3, min_row=ni_row, max_col=2+total_years)
        chart3.add_data(data3, from_rows=True)
        chart3.set_categories(cats)
        if len(chart3.series) > 0:
            chart3.series[0].tx = SeriesLabel(v="Net Income")
        chart3.width = 12
        chart3.height = 8
        
        ws.add_chart(chart3, "L18")
        
        # Key Metrics Summary
        ws['B22'] = "KEY METRICS SUMMARY"
        ws['B22'].style = 'subheader'
        ws.merge_cells('B22:D22')
        
        metrics = [
            ("Current Revenue", f"=Income_Statement!{get_column_letter(2+self.hist_years)}6", "number"),
            ("Forecast Y5 Revenue", f"=Income_Statement!{get_column_letter(2+self.hist_years+self.fcst_years)}6", "number"),
            ("Revenue CAGR", f"=IFERROR((Income_Statement!{get_column_letter(2+self.hist_years+self.fcst_years)}6/Income_Statement!{get_column_letter(2+self.hist_years)}6)^(1/{self.fcst_years})-1,0)", "percent"),
            ("Current EBITDA Margin", f"={get_column_letter(2+self.hist_years)}{margin_row}", "percent"),
            ("Forecast Y5 EBITDA Margin", f"={get_column_letter(2+self.hist_years+self.fcst_years)}{margin_row}", "percent"),
            ("Enterprise Value", "=Valuation!C44", "number"),
            ("Equity Value", "=Valuation!C46", "number"),
            ("Implied Share Price", "=Valuation!C50", "currency"),
        ]
        
        row = 23
        for name, formula, fmt in metrics:
            ws[f'B{row}'] = name
            ws[f'B{row}'].style = 'label'
            ws[f'C{row}'] = formula
            ws[f'C{row}'].style = 'output_cell'
            if fmt == "percent":
                ws[f'C{row}'].number_format = self.FORMATS['percent']
            elif fmt == "currency":
                ws[f'C{row}'].number_format = self.FORMATS['currency']
            else:
                ws[f'C{row}'].number_format = self.FORMATS['number']
            row += 1
        
        # DCF Waterfall Data
        row = 50
        ws[f'B{row}'] = "DCF Waterfall Data"
        ws[f'B{row}'].style = 'subheader'
        row += 1
        
        ws[f'B{row}'] = "Component"
        ws[f'C{row}'] = "Value"
        ws[f'D{row}'] = "Base"
        ws[f'E{row}'] = "Increase"
        ws[f'F{row}'] = "Decrease"
        for col in ['B', 'C', 'D', 'E', 'F']:
            ws[f'{col}{row}'].style = 'header'
        row += 1
        
        waterfall_start = row
        
        # Get dynamic row references from valuation
        val_rows = self.row_map.get('valuation', {})
        pv_fcff_row = val_rows.get('pv_fcff', 29)  # PV of FCFF row
        pv_tv_row = val_rows.get('pv_tv', 42)     # PV of Terminal Value row
        sum_pv_row = val_rows.get('sum_pv', 44)   # Sum of PV row
        ev_row = val_rows.get('ev', 46)           # Enterprise Value row
        net_debt_row = val_rows.get('net_debt', 47)  # Net Debt row
        equity_val_row = val_rows.get('equity_value', 48)  # Equity Value row
        
        # Starting point (Sum of PV of FCF)
        ws[f'B{row}'] = "PV of FCF"
        ws[f'C{row}'] = f"=Valuation!C{sum_pv_row}"  # Sum of PV of FCFF
        ws[f'C{row}'].number_format = self.FORMATS['number']
        ws[f'D{row}'] = 0
        ws[f'E{row}'] = "=C" + str(row)
        ws[f'F{row}'] = 0
        row += 1
        
        # Terminal Value add (PV of TV)
        ws[f'B{row}'] = "+ Terminal Value"
        ws[f'C{row}'] = f"=Valuation!C{sum_pv_row+1}"  # PV of Terminal Value (next row after sum_pv)
        ws[f'C{row}'].number_format = self.FORMATS['number']
        ws[f'D{row}'] = f"=C{row-1}+E{row-1}"
        ws[f'E{row}'] = f"=C{row}"
        ws[f'F{row}'] = 0
        row += 1
        
        # = Enterprise Value
        ws[f'B{row}'] = "= Enterprise Value"
        ws[f'C{row}'] = f"=Valuation!C{ev_row}"
        ws[f'C{row}'].number_format = self.FORMATS['number']
        ws[f'D{row}'] = 0
        ws[f'E{row}'] = 0
        ws[f'F{row}'] = 0
        row += 1
        
        # Less Net Debt
        ws[f'B{row}'] = "- Net Debt"
        ws[f'C{row}'] = f"=Valuation!C{net_debt_row}"
        ws[f'C{row}'].number_format = self.FORMATS['number']
        ws[f'D{row}'] = f"=Valuation!C{ev_row}"
        ws[f'E{row}'] = 0
        ws[f'F{row}'] = f"=ABS(C{row})"
        row += 1
        
        # = Equity Value
        ws[f'B{row}'] = "= Equity Value"
        ws[f'C{row}'] = f"=Valuation!C{equity_val_row}"
        ws[f'C{row}'].number_format = self.FORMATS['number']
        ws[f'D{row}'] = 0
        ws[f'E{row}'] = 0
        ws[f'F{row}'] = 0
        waterfall_end = row
        row += 1
        
        # Create DCF Waterfall Chart (Stacked Bar)
        from openpyxl.chart import BarChart as WaterfallChart
        
        chart4 = WaterfallChart()
        chart4.type = "col"
        chart4.grouping = "stacked"
        chart4.title = "DCF Valuation Bridge (₹ Crores)"
        chart4.style = 10
        
        # Data references for stacked waterfall
        cats_wf = Reference(ws, min_col=2, min_row=waterfall_start, max_row=waterfall_end)
        data_base = Reference(ws, min_col=4, min_row=waterfall_start-1, max_row=waterfall_end)
        data_inc = Reference(ws, min_col=5, min_row=waterfall_start-1, max_row=waterfall_end)
        data_dec = Reference(ws, min_col=6, min_row=waterfall_start-1, max_row=waterfall_end)
        
        chart4.add_data(data_base, titles_from_data=True)
        chart4.add_data(data_inc, titles_from_data=True)
        chart4.add_data(data_dec, titles_from_data=True)
        chart4.set_categories(cats_wf)
        
        # Style the series - make base transparent, increase green, decrease red
        from openpyxl.chart.series import SeriesLabel
        if len(chart4.series) >= 1:
            chart4.series[0].graphicalProperties.noFill = True  # Base is invisible
            chart4.series[0].graphicalProperties.line.noFill = True
        if len(chart4.series) >= 2:
            chart4.series[1].graphicalProperties.solidFill = "63BE7B"  # Green for increases
        if len(chart4.series) >= 3:
            chart4.series[2].graphicalProperties.solidFill = "F8696B"  # Red for decreases
        
        chart4.width = 16
        chart4.height = 10
        
        ws.add_chart(chart4, "B32")


def generate_financial_model(
    company_name: str,
    model_structure: Dict[str, Any],
    financial_data: Dict[str, Any],
    industry_info: Dict[str, Any],
    output_path: str
) -> str:
    """Generate financial model"""
    generator = FinancialModelGenerator(
        company_name, model_structure, financial_data, industry_info
    )
    return generator.generate(output_path)
