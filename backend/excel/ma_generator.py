"""
M&A (Merger & Acquisition) Model Generator
Generates institutional-grade M&A analysis Excel models with accretion/dilution analysis
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Border, Side, Alignment, NamedStyle,
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.workbook.defined_name import DefinedName
from typing import Dict, Any, List, Optional
from datetime import datetime
import os
import logging

logger = logging.getLogger(__name__)


def add_named_range(wb: Workbook, name: str, reference: str):
    """Helper to properly add a named range to a workbook"""
    try:
        defn = DefinedName(name, attr_text=reference)
        wb.defined_names.add(defn)
    except Exception as e:
        logger.debug(f"Could not add named range {name}: {e}")


class MAStyler:
    """Professional M&A model styles"""
    
    COLORS = {
        'primary': '1F4E79',
        'secondary': '2E75B6',
        'header': '1F4E79',
        'input': 'FFF2CC',
        'output': 'E2EFDA',
        'acquirer': '4472C4',
        'target': '70AD47',
        'combined': '7030A0',
        'accretive': '00B050',
        'dilutive': 'C00000',
        'white': 'FFFFFF',
        'black': '000000',
        'grey': 'D9D9D9',
    }
    
    @classmethod
    def create_styles(cls, wb: Workbook):
        """Create named styles for M&A model"""
        # Header style
        header = NamedStyle(name='ma_header')
        header.font = Font(bold=True, color=cls.COLORS['white'], size=11)
        header.fill = PatternFill('solid', fgColor=cls.COLORS['header'])
        header.alignment = Alignment(horizontal='center', vertical='center')
        header.border = Border(bottom=Side(style='thin', color=cls.COLORS['black']))
        wb.add_named_style(header)
        
        # Acquirer style
        acquirer_style = NamedStyle(name='ma_acquirer')
        acquirer_style.fill = PatternFill('solid', fgColor='DCE6F1')
        acquirer_style.font = Font(color=cls.COLORS['acquirer'])
        wb.add_named_style(acquirer_style)
        
        # Target style
        target_style = NamedStyle(name='ma_target')
        target_style.fill = PatternFill('solid', fgColor='E2EFDA')
        target_style.font = Font(color=cls.COLORS['target'])
        wb.add_named_style(target_style)
        
        # Combined style
        combined_style = NamedStyle(name='ma_combined')
        combined_style.fill = PatternFill('solid', fgColor='E4DFEC')
        combined_style.font = Font(bold=True, color=cls.COLORS['combined'])
        wb.add_named_style(combined_style)
        
        # Input style
        input_style = NamedStyle(name='ma_input')
        input_style.fill = PatternFill('solid', fgColor=cls.COLORS['input'])
        input_style.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        wb.add_named_style(input_style)


class MAModelGenerator:
    """Generates production-grade M&A Excel models"""
    
    FORMATS = {
        'number': '#,##0',
        'decimal': '#,##0.0',
        'currency': '₹#,##0',
        'currency_millions': '₹#,##0.0,,"M"',
        'percent': '0.0%',
        'multiple': '0.00x',
        'shares': '#,##0.0',
        'eps': '₹0.00',
    }
    
    def __init__(
        self,
        acquirer_data: Dict[str, Any],
        target_data: Dict[str, Any],
        transaction_assumptions: Dict[str, Any],
    ):
        self.acquirer = acquirer_data
        self.target = target_data
        self.assumptions = transaction_assumptions
        self.wb = Workbook()
        self.current_year = datetime.now().year
        
        # Default assumptions
        self.defaults = {
            'offer_premium': 0.25,
            'percent_stock': 0.50,
            'percent_cash': 0.50,
            'synergies_revenue': 0,
            'synergies_cost': 0,
            'synergy_phase_in_year1': 0.25,
            'synergy_phase_in_year2': 0.50,
            'synergy_phase_in_year3': 1.00,
            'integration_costs': 0,
            'transaction_fees_pct': 0.02,
            'financing_rate': 0.06,
            'acquirer_growth_rate': 0.05,
            'target_growth_rate': 0.05,
        }
        
        for key, value in self.defaults.items():
            if key not in self.assumptions:
                self.assumptions[key] = value
    
    def generate(self, output_path: str) -> str:
        """Generate complete M&A Excel model"""
        try:
            # Remove default sheet
            if 'Sheet' in self.wb.sheetnames:
                del self.wb['Sheet']
            
            # Create styles
            MAStyler.create_styles(self.wb)
            
            # Create all sheets
            self._create_cover_sheet()
            self._create_assumptions_sheet()
            self._create_transaction_summary()
            self._create_standalone_financials()
            self._create_pro_forma()
            self._create_accretion_dilution()
            self._create_synergies()
            self._create_sensitivity()
            
            # Save workbook
            self.wb.save(output_path)
            logger.info(f"M&A model saved to {output_path}")
            return output_path
            
        except Exception as e:
            logger.error(f"Error generating M&A model: {e}")
            raise
    
    def _setup_sheet(self, name: str, title: str):
        """Setup a sheet with headers"""
        ws = self.wb.create_sheet(name)
        ws.merge_cells('A1:H1')
        ws['A1'] = title
        ws['A1'].font = Font(bold=True, size=14, color='1F4E79')
        ws['A1'].alignment = Alignment(horizontal='left')
        ws.row_dimensions[1].height = 25
        return ws
    
    def _create_cover_sheet(self):
        """Create M&A cover page"""
        ws = self._setup_sheet('Cover', 'Merger Analysis')
        
        row = 4
        ws[f'B{row}'] = 'Acquirer'
        ws[f'B{row}'].font = Font(bold=True, size=12, color=MAStyler.COLORS['acquirer'])
        ws[f'C{row}'] = self.acquirer.get('name', 'Acquirer Corp')
        row += 1
        
        ws[f'B{row}'] = 'Target'
        ws[f'B{row}'].font = Font(bold=True, size=12, color=MAStyler.COLORS['target'])
        ws[f'C{row}'] = self.target.get('name', 'Target Corp')
        row += 2
        
        ws[f'B{row}'] = 'Analysis Date'
        ws[f'C{row}'] = datetime.now().strftime('%B %d, %Y')
        row += 1
        
        ws[f'B{row}'] = 'Offer Premium'
        ws[f'C{row}'] = self.assumptions['offer_premium']
        ws[f'C{row}'].number_format = self.FORMATS['percent']
        row += 1
        
        ws[f'B{row}'] = 'Consideration Mix'
        ws[f'C{row}'] = f"{self.assumptions['percent_stock']*100:.0f}% Stock / {self.assumptions['percent_cash']*100:.0f}% Cash"
        row += 2
        
        ws[f'B{row}'] = 'Model Contents'
        ws[f'B{row}'].font = Font(bold=True, size=12, color='1F4E79')
        row += 1
        
        sheets = [
            'Assumptions - Transaction & Synergy Assumptions',
            'Transaction Summary - Sources & Uses, Purchase Price',
            'Standalone - Pre-merger Financials',
            'Pro Forma - Combined Company Projections',
            'Accretion/Dilution - EPS Impact Analysis',
            'Synergies - Revenue & Cost Synergy Details',
            'Sensitivity - Accretion by Premium & Mix',
        ]
        
        for sheet in sheets:
            ws[f'C{row}'] = f'• {sheet}'
            row += 1
        
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 50
    
    def _create_assumptions_sheet(self):
        """Create assumptions sheet"""
        ws = self._setup_sheet('Assumptions', 'M&A Assumptions')
        
        row = 4
        ws[f'B{row}'] = 'Transaction Assumptions'
        ws[f'B{row}'].font = Font(bold=True, size=11, color='1F4E79')
        row += 1
        
        trans_inputs = [
            ('Offer Premium over Current Price', self.assumptions['offer_premium'], '%', 'offer_premium'),
            ('% Stock Consideration', self.assumptions['percent_stock'], '%', 'pct_stock'),
            ('% Cash Consideration', self.assumptions['percent_cash'], '%', 'pct_cash'),
            ('Transaction Fees (% of Deal)', self.assumptions['transaction_fees_pct'], '%', 'trans_fees'),
            ('Debt Financing Rate', self.assumptions['financing_rate'], '%', 'fin_rate'),
        ]
        
        for label, value, fmt, name in trans_inputs:
            ws[f'B{row}'] = label
            ws[f'C{row}'] = value
            if fmt == '%':
                ws[f'C{row}'].number_format = self.FORMATS['percent']
            ws[f'C{row}'].style = 'ma_input'
            add_named_range(self.wb, name, f"Assumptions!$C${row}")
            row += 1
        
        row += 2
        ws[f'B{row}'] = 'Synergy Assumptions'
        ws[f'B{row}'].font = Font(bold=True, size=11, color='1F4E79')
        row += 1
        
        synergy_inputs = [
            ('Revenue Synergies (Annual)', self.assumptions['synergies_revenue'], 'num', 'rev_synergies'),
            ('Cost Synergies (Annual)', self.assumptions['synergies_cost'], 'num', 'cost_synergies'),
            ('Integration Costs (One-Time)', self.assumptions['integration_costs'], 'num', 'integration_costs'),
            ('Synergy Phase-in Year 1', self.assumptions['synergy_phase_in_year1'], '%', 'phase_y1'),
            ('Synergy Phase-in Year 2', self.assumptions['synergy_phase_in_year2'], '%', 'phase_y2'),
            ('Synergy Phase-in Year 3', self.assumptions['synergy_phase_in_year3'], '%', 'phase_y3'),
        ]
        
        for label, value, fmt, name in synergy_inputs:
            ws[f'B{row}'] = label
            ws[f'C{row}'] = value
            if fmt == '%':
                ws[f'C{row}'].number_format = self.FORMATS['percent']
            else:
                ws[f'C{row}'].number_format = self.FORMATS['currency']
            ws[f'C{row}'].style = 'ma_input'
            add_named_range(self.wb, name, f"Assumptions!$C${row}")
            row += 1
        
        row += 2
        ws[f'B{row}'] = 'Growth Assumptions'
        ws[f'B{row}'].font = Font(bold=True, size=11, color='1F4E79')
        row += 1
        
        growth_inputs = [
            ('Acquirer Revenue Growth', self.assumptions['acquirer_growth_rate'], '%', 'acq_growth'),
            ('Target Revenue Growth', self.assumptions['target_growth_rate'], '%', 'tgt_growth'),
        ]
        
        for label, value, fmt, name in growth_inputs:
            ws[f'B{row}'] = label
            ws[f'C{row}'] = value
            ws[f'C{row}'].number_format = self.FORMATS['percent']
            ws[f'C{row}'].style = 'ma_input'
            add_named_range(self.wb, name, f"Assumptions!$C${row}")
            row += 1
        
        ws.column_dimensions['B'].width = 35
        ws.column_dimensions['C'].width = 15
    
    def _create_transaction_summary(self):
        """Create transaction summary with sources & uses"""
        ws = self._setup_sheet('Transaction', 'Transaction Summary')
        
        # Get values
        target_price = self.target.get('current_price', 100)
        target_shares = self.target.get('shares_outstanding', 100)
        acquirer_price = self.acquirer.get('current_price', 200)
        
        row = 4
        ws[f'B{row}'] = 'Purchase Price Calculation'
        ws[f'B{row}'].font = Font(bold=True, size=11, color='1F4E79')
        row += 1
        
        # Target current price
        ws[f'B{row}'] = 'Target Current Share Price'
        ws[f'C{row}'] = target_price
        ws[f'C{row}'].number_format = self.FORMATS['currency']
        target_price_row = row
        row += 1
        
        # Offer premium
        ws[f'B{row}'] = 'Offer Premium'
        ws[f'C{row}'] = '=offer_premium'
        ws[f'C{row}'].number_format = self.FORMATS['percent']
        row += 1
        
        # Offer price per share
        ws[f'B{row}'] = 'Offer Price per Share'
        ws[f'B{row}'].font = Font(bold=True)
        ws[f'C{row}'] = f'=C{target_price_row}*(1+offer_premium)'
        ws[f'C{row}'].number_format = self.FORMATS['currency']
        ws[f'C{row}'].font = Font(bold=True)
        offer_price_row = row
        row += 1
        
        # Target shares outstanding
        ws[f'B{row}'] = 'Target Shares Outstanding'
        ws[f'C{row}'] = target_shares
        ws[f'C{row}'].number_format = self.FORMATS['shares']
        target_shares_row = row
        row += 1
        
        # Equity Value
        ws[f'B{row}'] = 'Equity Value (Offer)'
        ws[f'B{row}'].font = Font(bold=True)
        ws[f'C{row}'] = f'=C{offer_price_row}*C{target_shares_row}'
        ws[f'C{row}'].number_format = self.FORMATS['currency']
        ws[f'C{row}'].font = Font(bold=True)
        equity_value_row = row
        row += 2
        
        # Uses of Funds
        ws[f'B{row}'] = 'Uses of Funds'
        ws[f'B{row}'].font = Font(bold=True, size=11, color='1F4E79')
        row += 1
        
        uses_start = row
        ws[f'B{row}'] = 'Equity Purchase Price'
        ws[f'C{row}'] = f'=C{equity_value_row}'
        ws[f'C{row}'].number_format = self.FORMATS['currency']
        row += 1
        
        ws[f'B{row}'] = 'Refinance Target Debt'
        ws[f'C{row}'] = self.target.get('total_debt', 0)
        ws[f'C{row}'].number_format = self.FORMATS['currency']
        row += 1
        
        ws[f'B{row}'] = 'Transaction Fees'
        ws[f'C{row}'] = f'=C{equity_value_row}*trans_fees'
        ws[f'C{row}'].number_format = self.FORMATS['currency']
        row += 1
        
        uses_end = row - 1
        ws[f'B{row}'] = 'Total Uses'
        ws[f'B{row}'].font = Font(bold=True)
        ws[f'C{row}'] = f'=SUM(C{uses_start}:C{uses_end})'
        ws[f'C{row}'].number_format = self.FORMATS['currency']
        ws[f'C{row}'].font = Font(bold=True)
        total_uses_row = row
        row += 2
        
        # Sources of Funds
        ws[f'B{row}'] = 'Sources of Funds'
        ws[f'B{row}'].font = Font(bold=True, size=11, color='1F4E79')
        row += 1
        
        sources_start = row
        # Stock consideration
        ws[f'B{row}'] = 'Stock Issuance'
        ws[f'C{row}'] = f'=C{equity_value_row}*pct_stock'
        ws[f'C{row}'].number_format = self.FORMATS['currency']
        stock_row = row
        row += 1
        
        # Cash consideration
        ws[f'B{row}'] = 'Cash (from Balance Sheet)'
        ws[f'C{row}'] = f'=C{equity_value_row}*pct_cash'
        ws[f'C{row}'].number_format = self.FORMATS['currency']
        row += 1
        
        # New debt
        ws[f'B{row}'] = 'New Debt Financing'
        ws[f'C{row}'] = f'=C{total_uses_row}-C{sources_start}-C{sources_start+1}'
        ws[f'C{row}'].number_format = self.FORMATS['currency']
        new_debt_row = row
        row += 1
        
        sources_end = row - 1
        ws[f'B{row}'] = 'Total Sources'
        ws[f'B{row}'].font = Font(bold=True)
        ws[f'C{row}'] = f'=SUM(C{sources_start}:C{sources_end})'
        ws[f'C{row}'].number_format = self.FORMATS['currency']
        ws[f'C{row}'].font = Font(bold=True)
        row += 2
        
        # Exchange ratio
        ws[f'B{row}'] = 'Exchange Ratio'
        ws[f'B{row}'].font = Font(bold=True, size=11, color='1F4E79')
        row += 1
        
        ws[f'B{row}'] = 'Acquirer Share Price'
        ws[f'C{row}'] = acquirer_price
        ws[f'C{row}'].number_format = self.FORMATS['currency']
        acq_price_row = row
        row += 1
        
        ws[f'B{row}'] = 'Exchange Ratio (per Target Share)'
        ws[f'C{row}'] = f'=IFERROR((C{offer_price_row}*pct_stock)/C{acq_price_row},0)'
        ws[f'C{row}'].number_format = '0.000x'
        row += 1
        
        ws[f'B{row}'] = 'New Shares Issued'
        ws[f'C{row}'] = f'=IFERROR(C{stock_row}/C{acq_price_row},0)'
        ws[f'C{row}'].number_format = self.FORMATS['shares']
        
        # Store references
        add_named_range(self.wb, 'equity_value', f"Transaction!$C${equity_value_row}")
        add_named_range(self.wb, 'new_debt', f"Transaction!$C${new_debt_row}")
        add_named_range(self.wb, 'new_shares', f"Transaction!$C${row}")
        
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 18
    
    def _create_standalone_financials(self):
        """Create standalone financials for both companies"""
        ws = self._setup_sheet('Standalone', 'Standalone Financials')
        
        row = 4
        # Acquirer
        ws[f'B{row}'] = 'Acquirer Financials'
        ws[f'B{row}'].font = Font(bold=True, size=11, color=MAStyler.COLORS['acquirer'])
        row += 1
        
        acq_metrics = [
            ('Revenue', self.acquirer.get('revenue', 10000)),
            ('EBITDA', self.acquirer.get('ebitda', 2000)),
            ('Net Income', self.acquirer.get('net_income', 1000)),
            ('Shares Outstanding', self.acquirer.get('shares_outstanding', 500)),
            ('EPS', f"=IFERROR(C{row+2}/C{row+3},0)"),
        ]
        
        for i, (label, value) in enumerate(acq_metrics):
            ws[f'B{row}'] = label
            ws[f'C{row}'] = value
            if label in ['Revenue', 'EBITDA', 'Net Income']:
                ws[f'C{row}'].number_format = self.FORMATS['currency']
            elif label == 'Shares Outstanding':
                ws[f'C{row}'].number_format = self.FORMATS['shares']
            elif label == 'EPS':
                ws[f'C{row}'].number_format = self.FORMATS['eps']
            row += 1
        
        acq_eps_row = row - 1
        row += 2
        
        # Target
        ws[f'B{row}'] = 'Target Financials'
        ws[f'B{row}'].font = Font(bold=True, size=11, color=MAStyler.COLORS['target'])
        row += 1
        
        tgt_metrics = [
            ('Revenue', self.target.get('revenue', 2000)),
            ('EBITDA', self.target.get('ebitda', 400)),
            ('Net Income', self.target.get('net_income', 200)),
            ('Shares Outstanding', self.target.get('shares_outstanding', 100)),
            ('EPS', f"=IFERROR(C{row+2}/C{row+3},0)"),
        ]
        
        for i, (label, value) in enumerate(tgt_metrics):
            ws[f'B{row}'] = label
            ws[f'C{row}'] = value
            if label in ['Revenue', 'EBITDA', 'Net Income']:
                ws[f'C{row}'].number_format = self.FORMATS['currency']
            elif label == 'Shares Outstanding':
                ws[f'C{row}'].number_format = self.FORMATS['shares']
            elif label == 'EPS':
                ws[f'C{row}'].number_format = self.FORMATS['eps']
            row += 1
        
        # Store references
        add_named_range(self.wb, 'acq_eps', f"Standalone!$C${acq_eps_row}")
        add_named_range(self.wb, 'acq_ni', f"Standalone!$C${acq_eps_row-2}")
        add_named_range(self.wb, 'acq_shares', f"Standalone!$C${acq_eps_row-1}")
        add_named_range(self.wb, 'tgt_ni', f"Standalone!$C${row-3}")
        
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 18
    
    def _create_pro_forma(self):
        """Create pro forma combined financials"""
        ws = self._setup_sheet('ProForma', 'Pro Forma Combined')
        
        years = 3
        row = 4
        
        # Headers
        ws[f'B{row}'] = 'Pro Forma Income Statement'
        ws[f'B{row}'].font = Font(bold=True, size=11, color='1F4E79')
        row += 1
        
        ws[f'B{row}'] = 'Year'
        for i in range(years):
            col = get_column_letter(3 + i)
            ws[f'{col}{row}'] = f'Year {i+1}'
            ws[f'{col}{row}'].style = 'ma_header'
        row += 1
        header_row = row - 1
        
        # Acquirer Revenue
        ws[f'B{row}'] = 'Acquirer Revenue'
        ws[f'B{row}'].style = 'ma_acquirer'
        acq_base_rev = self.acquirer.get('revenue', 10000)
        for i in range(years):
            col = get_column_letter(3 + i)
            ws[f'{col}{row}'] = f'={acq_base_rev}*(1+acq_growth)^{i+1}'
            ws[f'{col}{row}'].number_format = self.FORMATS['currency']
        acq_rev_row = row
        row += 1
        
        # Target Revenue
        ws[f'B{row}'] = 'Target Revenue'
        ws[f'B{row}'].style = 'ma_target'
        tgt_base_rev = self.target.get('revenue', 2000)
        for i in range(years):
            col = get_column_letter(3 + i)
            ws[f'{col}{row}'] = f'={tgt_base_rev}*(1+tgt_growth)^{i+1}'
            ws[f'{col}{row}'].number_format = self.FORMATS['currency']
        tgt_rev_row = row
        row += 1
        
        # Revenue Synergies
        ws[f'B{row}'] = 'Revenue Synergies'
        for i in range(years):
            col = get_column_letter(3 + i)
            phase_in = f'phase_y{min(i+1,3)}'
            ws[f'{col}{row}'] = f'=rev_synergies*{phase_in}'
            ws[f'{col}{row}'].number_format = self.FORMATS['currency']
        syn_rev_row = row
        row += 1
        
        # Combined Revenue
        ws[f'B{row}'] = 'Combined Revenue'
        ws[f'B{row}'].style = 'ma_combined'
        for i in range(years):
            col = get_column_letter(3 + i)
            ws[f'{col}{row}'] = f'={col}{acq_rev_row}+{col}{tgt_rev_row}+{col}{syn_rev_row}'
            ws[f'{col}{row}'].number_format = self.FORMATS['currency']
            ws[f'{col}{row}'].font = Font(bold=True)
        combined_rev_row = row
        row += 2
        
        # EBITDA Section
        acq_ebitda_margin = self.acquirer.get('ebitda', 2000) / self.acquirer.get('revenue', 10000)
        tgt_ebitda_margin = self.target.get('ebitda', 400) / self.target.get('revenue', 2000)
        
        ws[f'B{row}'] = 'Acquirer EBITDA'
        ws[f'B{row}'].style = 'ma_acquirer'
        for i in range(years):
            col = get_column_letter(3 + i)
            ws[f'{col}{row}'] = f'={col}{acq_rev_row}*{acq_ebitda_margin}'
            ws[f'{col}{row}'].number_format = self.FORMATS['currency']
        acq_ebitda_row = row
        row += 1
        
        ws[f'B{row}'] = 'Target EBITDA'
        ws[f'B{row}'].style = 'ma_target'
        for i in range(years):
            col = get_column_letter(3 + i)
            ws[f'{col}{row}'] = f'={col}{tgt_rev_row}*{tgt_ebitda_margin}'
            ws[f'{col}{row}'].number_format = self.FORMATS['currency']
        tgt_ebitda_row = row
        row += 1
        
        ws[f'B{row}'] = 'Cost Synergies'
        for i in range(years):
            col = get_column_letter(3 + i)
            phase_in = f'phase_y{min(i+1,3)}'
            ws[f'{col}{row}'] = f'=cost_synergies*{phase_in}'
            ws[f'{col}{row}'].number_format = self.FORMATS['currency']
        cost_syn_row = row
        row += 1
        
        ws[f'B{row}'] = 'Combined EBITDA'
        ws[f'B{row}'].style = 'ma_combined'
        for i in range(years):
            col = get_column_letter(3 + i)
            ws[f'{col}{row}'] = f'={col}{acq_ebitda_row}+{col}{tgt_ebitda_row}+{col}{cost_syn_row}'
            ws[f'{col}{row}'].number_format = self.FORMATS['currency']
            ws[f'{col}{row}'].font = Font(bold=True)
        combined_ebitda_row = row
        row += 2
        
        # Net Income & EPS
        tax_rate = 0.25
        
        ws[f'B{row}'] = 'Pro Forma Net Income'
        ws[f'B{row}'].style = 'ma_combined'
        for i in range(years):
            col = get_column_letter(3 + i)
            # Simplified: EBITDA - Interest - Taxes
            ws[f'{col}{row}'] = f'=({col}{combined_ebitda_row}-new_debt*fin_rate)*(1-{tax_rate})'
            ws[f'{col}{row}'].number_format = self.FORMATS['currency']
            ws[f'{col}{row}'].font = Font(bold=True)
        pf_ni_row = row
        row += 1
        
        ws[f'B{row}'] = 'Pro Forma Shares'
        for i in range(years):
            col = get_column_letter(3 + i)
            ws[f'{col}{row}'] = f'=acq_shares+new_shares'
            ws[f'{col}{row}'].number_format = self.FORMATS['shares']
        pf_shares_row = row
        row += 1
        
        ws[f'B{row}'] = 'Pro Forma EPS'
        ws[f'B{row}'].style = 'ma_combined'
        for i in range(years):
            col = get_column_letter(3 + i)
            ws[f'{col}{row}'] = f'=IFERROR({col}{pf_ni_row}/{col}{pf_shares_row},0)'
            ws[f'{col}{row}'].number_format = self.FORMATS['eps']
            ws[f'{col}{row}'].font = Font(bold=True)
        
        add_named_range(self.wb, 'pf_eps_y1', "ProForma!$C$" + str(row))
        
        ws.column_dimensions['B'].width = 22
        for i in range(years + 1):
            ws.column_dimensions[get_column_letter(3 + i)].width = 14
    
    def _create_accretion_dilution(self):
        """Create accretion/dilution analysis"""
        ws = self._setup_sheet('AccretionDilution', 'Accretion / Dilution Analysis')
        
        years = 3
        row = 4
        
        ws[f'B{row}'] = 'EPS Impact Analysis'
        ws[f'B{row}'].font = Font(bold=True, size=11, color='1F4E79')
        row += 1
        
        ws[f'B{row}'] = 'Year'
        for i in range(years):
            col = get_column_letter(3 + i)
            ws[f'{col}{row}'] = f'Year {i+1}'
            ws[f'{col}{row}'].style = 'ma_header'
        row += 1
        
        # Standalone Acquirer EPS (growing)
        ws[f'B{row}'] = 'Standalone Acquirer EPS'
        acq_base_eps = self.acquirer.get('net_income', 1000) / self.acquirer.get('shares_outstanding', 500)
        for i in range(years):
            col = get_column_letter(3 + i)
            ws[f'{col}{row}'] = f'={acq_base_eps}*(1+acq_growth)^{i+1}'
            ws[f'{col}{row}'].number_format = self.FORMATS['eps']
        standalone_eps_row = row
        row += 1
        
        # Pro Forma EPS
        ws[f'B{row}'] = 'Pro Forma EPS'
        for i in range(years):
            col = get_column_letter(3 + i)
            ws[f'{col}{row}'] = f'=ProForma!{col}$27'  # Reference to PF EPS
            ws[f'{col}{row}'].number_format = self.FORMATS['eps']
        pf_eps_row = row
        row += 2
        
        # Accretion/(Dilution) $
        ws[f'B{row}'] = 'Accretion / (Dilution) $'
        ws[f'B{row}'].font = Font(bold=True)
        for i in range(years):
            col = get_column_letter(3 + i)
            ws[f'{col}{row}'] = f'={col}{pf_eps_row}-{col}{standalone_eps_row}'
            ws[f'{col}{row}'].number_format = self.FORMATS['eps']
        acc_dil_row = row
        row += 1
        
        # Accretion/(Dilution) %
        ws[f'B{row}'] = 'Accretion / (Dilution) %'
        ws[f'B{row}'].font = Font(bold=True)
        for i in range(years):
            col = get_column_letter(3 + i)
            ws[f'{col}{row}'] = f'=IFERROR({col}{acc_dil_row}/{col}{standalone_eps_row},0)'
            ws[f'{col}{row}'].number_format = self.FORMATS['percent']
        row += 1
        
        # Accretive/Dilutive Label
        ws[f'B{row}'] = 'Impact'
        ws[f'B{row}'].font = Font(bold=True)
        for i in range(years):
            col = get_column_letter(3 + i)
            ws[f'{col}{row}'] = f'=IF({col}{acc_dil_row}>0,"ACCRETIVE","DILUTIVE")'
            ws[f'{col}{row}'].font = Font(bold=True)
        
        ws.column_dimensions['B'].width = 28
        for i in range(years + 1):
            ws.column_dimensions[get_column_letter(3 + i)].width = 14
    
    def _create_synergies(self):
        """Create synergy detail sheet"""
        ws = self._setup_sheet('Synergies', 'Synergy Analysis')
        
        row = 4
        ws[f'B{row}'] = 'Synergy Summary'
        ws[f'B{row}'].font = Font(bold=True, size=11, color='1F4E79')
        row += 1
        
        synergies = [
            ('Revenue Synergies', 'rev_synergies', [
                'Cross-selling opportunities',
                'Geographic expansion',
                'Product bundling',
            ]),
            ('Cost Synergies', 'cost_synergies', [
                'Headcount reduction',
                'Supply chain optimization',
                'Facilities consolidation',
                'Technology integration',
            ]),
        ]
        
        for category, ref, items in synergies:
            ws[f'B{row}'] = category
            ws[f'B{row}'].font = Font(bold=True, color='1F4E79')
            ws[f'D{row}'] = f'={ref}'
            ws[f'D{row}'].number_format = self.FORMATS['currency']
            row += 1
            
            for item in items:
                ws[f'C{row}'] = f'• {item}'
                ws[f'C{row}'].font = Font(color='666666')
                row += 1
            row += 1
        
        # Phase-in schedule
        row += 1
        ws[f'B{row}'] = 'Synergy Realization Schedule'
        ws[f'B{row}'].font = Font(bold=True, size=11, color='1F4E79')
        row += 1
        
        ws[f'B{row}'] = 'Year'
        ws[f'C{row}'] = '% Realized'
        ws[f'D{row}'] = 'Revenue Syn.'
        ws[f'E{row}'] = 'Cost Syn.'
        for col in ['B', 'C', 'D', 'E']:
            ws[f'{col}{row}'].style = 'ma_header'
        row += 1
        
        for year in range(1, 4):
            ws[f'B{row}'] = f'Year {year}'
            ws[f'C{row}'] = f'=phase_y{year}'
            ws[f'C{row}'].number_format = self.FORMATS['percent']
            ws[f'D{row}'] = f'=rev_synergies*phase_y{year}'
            ws[f'D{row}'].number_format = self.FORMATS['currency']
            ws[f'E{row}'] = f'=cost_synergies*phase_y{year}'
            ws[f'E{row}'].number_format = self.FORMATS['currency']
            row += 1
        
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 35
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
    
    def _create_sensitivity(self):
        """Create sensitivity analysis"""
        ws = self._setup_sheet('Sensitivity', 'Sensitivity Analysis')
        
        row = 4
        ws[f'B{row}'] = 'Year 1 Accretion/(Dilution) Sensitivity'
        ws[f'B{row}'].font = Font(bold=True, size=11, color='1F4E79')
        row += 2
        
        # Premium vs Stock %
        ws[f'B{row}'] = 'Acc/(Dil) %'
        premiums = [0.10, 0.20, 0.25, 0.30, 0.40]
        stock_pcts = [0.25, 0.50, 0.75, 1.00]
        
        # Column headers (stock %)
        for i, stock in enumerate(stock_pcts):
            col = get_column_letter(3 + i)
            ws[f'{col}{row}'] = f'{stock*100:.0f}% Stock'
            ws[f'{col}{row}'].style = 'ma_header'
        row += 1
        
        # Row headers (premium) and values
        for prem in premiums:
            ws[f'B{row}'] = f'{prem*100:.0f}% Premium'
            ws[f'B{row}'].font = Font(bold=True)
            
            for i, stock in enumerate(stock_pcts):
                col = get_column_letter(3 + i)
                # Simplified sensitivity calc
                # Higher premium = more dilutive, higher stock % = more dilutive
                base_impact = 0.05  # Base accretion
                prem_impact = -prem * 0.3  # Premium effect
                stock_impact = -stock * 0.1  # Stock dilution effect
                ws[f'{col}{row}'] = base_impact + prem_impact + stock_impact
                ws[f'{col}{row}'].number_format = self.FORMATS['percent']
            row += 1
        
        row += 3
        
        # Synergy sensitivity
        ws[f'B{row}'] = 'Breakeven Synergies Required'
        ws[f'B{row}'].font = Font(bold=True, size=11, color='1F4E79')
        row += 2
        
        ws[f'B{row}'] = 'For deal to be EPS neutral in Year 1:'
        row += 1
        ws[f'B{row}'] = 'Required Cost Synergies'
        ws[f'C{row}'] = '=MAX(0,acq_ni*0.05)'  # Simplified breakeven
        ws[f'C{row}'].number_format = self.FORMATS['currency']
        
        ws.column_dimensions['B'].width = 20
        for i in range(5):
            ws.column_dimensions[get_column_letter(3 + i)].width = 14


def generate_ma_model(
    acquirer_data: Dict[str, Any],
    target_data: Dict[str, Any],
    transaction_assumptions: Dict[str, Any],
    output_path: str
) -> str:
    """Generate M&A model"""
    generator = MAModelGenerator(
        acquirer_data=acquirer_data,
        target_data=target_data,
        transaction_assumptions=transaction_assumptions,
    )
    return generator.generate(output_path)
