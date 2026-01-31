"""
PDF Exporter Module - Corrected Version
Generates PDF summary reports from financial models
"""

import os
import logging
from typing import Dict, Any, Optional
from datetime import datetime

logger = logging.getLogger(__name__)

# Try to import reportlab
try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, letter
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch, cm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image, PageBreak
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False
    logger.warning("reportlab not installed. Run: pip install reportlab")


def generate_pdf_report(
    output_path: str,
    company_name: str,
    industry: str,
    valuation_data: Dict[str, Any],
    assumptions: Dict[str, Any],
    commentary: Optional[Dict[str, str]] = None
) -> bool:
    """
    Generate a PDF summary report
    
    Args:
        output_path: Path to save the PDF
        company_name: Name of the company
        industry: Industry classification
        valuation_data: Valuation metrics
        assumptions: Model assumptions
        commentary: AI-generated commentary (optional)
    
    Returns:
        True if successful, False otherwise
    """
    if not REPORTLAB_AVAILABLE:
        logger.error("reportlab not available for PDF generation")
        return False
    
    try:
        doc = SimpleDocTemplate(
            output_path,
            pagesize=A4,
            rightMargin=72,
            leftMargin=72,
            topMargin=72,
            bottomMargin=72
        )
        
        styles = getSampleStyleSheet()
        
        # Custom styles
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            spaceAfter=30,
            textColor=colors.HexColor('#1a365d'),
            alignment=TA_CENTER
        )
        
        subtitle_style = ParagraphStyle(
            'CustomSubtitle',
            parent=styles['Normal'],
            fontSize=12,
            spaceAfter=20,
            textColor=colors.HexColor('#4a5568'),
            alignment=TA_CENTER
        )
        
        section_style = ParagraphStyle(
            'SectionHeader',
            parent=styles['Heading2'],
            fontSize=14,
            spaceBefore=20,
            spaceAfter=10,
            textColor=colors.HexColor('#2d3748'),
            borderWidth=1,
            borderColor=colors.HexColor('#e2e8f0'),
            borderPadding=5
        )
        
        body_style = ParagraphStyle(
            'BodyText',
            parent=styles['Normal'],
            fontSize=10,
            spaceAfter=8,
            textColor=colors.HexColor('#1a202c')
        )
        
        story = []
        
        # Title
        story.append(Paragraph(company_name, title_style))
        story.append(Paragraph(f"Financial Model Summary | {industry.title()}", subtitle_style))
        story.append(Paragraph(f"Generated: {datetime.now().strftime('%d %B %Y')}", subtitle_style))
        story.append(Spacer(1, 30))
        
        # Key Metrics Table
        story.append(Paragraph("Valuation Summary", section_style))
        
        metrics_data = [
            ["Metric", "Value"],
            ["Enterprise Value", f"₹{valuation_data.get('enterprise_value', 0):,.0f} Cr"],
            ["Equity Value", f"₹{valuation_data.get('equity_value', 0):,.0f} Cr"],
            ["Implied Share Price", f"₹{valuation_data.get('share_price', 0):,.2f}"],
            ["WACC", f"{valuation_data.get('wacc', 0) * 100:.1f}%"],
        ]
        
        metrics_table = Table(metrics_data, colWidths=[3*inch, 2*inch])
        metrics_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2d3748')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f7fafc')),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#e2e8f0')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f7fafc')]),
        ]))
        story.append(metrics_table)
        story.append(Spacer(1, 20))
        
        # Assumptions Table
        story.append(Paragraph("Key Assumptions", section_style))
        
        assumptions_data = [
            ["Parameter", "Value"],
            ["Revenue Growth", f"{assumptions.get('revenue_growth', 0) * 100:.1f}%"],
            ["EBITDA Margin", f"{assumptions.get('ebitda_margin', 0) * 100:.1f}%"],
            ["Tax Rate", f"{assumptions.get('tax_rate', 0) * 100:.1f}%"],
            ["Terminal Growth", f"{assumptions.get('terminal_growth', 0) * 100:.1f}%"],
            ["Risk-Free Rate", f"{assumptions.get('risk_free_rate', 0.07) * 100:.1f}%"],
            ["Beta", f"{assumptions.get('beta', 1.0):.2f}"],
        ]
        
        assumptions_table = Table(assumptions_data, colWidths=[3*inch, 2*inch])
        assumptions_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#38a169')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#e2e8f0')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f0fff4')]),
        ]))
        story.append(assumptions_table)
        story.append(Spacer(1, 20))
        
        # AI Commentary (if available)
        if commentary:
            story.append(Paragraph("Investment Analysis", section_style))
            
            if commentary.get('investment_thesis'):
                story.append(Paragraph("<b>Investment Thesis:</b>", body_style))
                story.append(Paragraph(commentary['investment_thesis'], body_style))
                story.append(Spacer(1, 10))
            
            if commentary.get('key_risks'):
                story.append(Paragraph("<b>Key Risks:</b>", body_style))
                story.append(Paragraph(commentary['key_risks'], body_style))
                story.append(Spacer(1, 10))
            
            if commentary.get('recommendation'):
                story.append(Paragraph("<b>Recommendation:</b>", body_style))
                story.append(Paragraph(commentary['recommendation'], body_style))
                story.append(Spacer(1, 10))
            
            if commentary.get('sensitivity_note'):
                story.append(Paragraph("<b>Sensitivity:</b>", body_style))
                story.append(Paragraph(commentary['sensitivity_note'], body_style))
        
        # Footer
        story.append(Spacer(1, 40))
        footer_style = ParagraphStyle(
            'Footer',
            parent=styles['Normal'],
            fontSize=8,
            textColor=colors.HexColor('#718096'),
            alignment=TA_CENTER
        )
        story.append(Paragraph(
            "This report is generated by AI Financial Modeler. "
            "The projections are based on assumptions and historical data. "
            "Please conduct your own due diligence before making investment decisions.",
            footer_style
        ))
        
        # Build PDF
        doc.build(story)
        
        logger.info(f"PDF report generated: {output_path}")
        return True
        
    except Exception as e:
        logger.error(f"Error generating PDF report: {e}")
        return False


def create_pdf_report(company_name: str, excel_path: str, output_path: str) -> bool:
    """
    Wrapper function to create PDF report from Excel model
    
    Args:
        company_name: Name of the company
        excel_path: Path to the generated Excel model
        output_path: Path to save the PDF report
    
    Returns:
        True if successful, False otherwise
    """
    if not REPORTLAB_AVAILABLE:
        logger.error("reportlab not available for PDF generation")
        return False
    
    try:
        import openpyxl
        
        # Load the Excel workbook (without data_only to access formulas if needed)
        wb = openpyxl.load_workbook(excel_path)
        
        # Extract industry from Summary sheet (Row 8, Column 3)
        industry = "General"
        if "Summary" in wb.sheetnames:
            summary = wb["Summary"]
            industry_cell = summary.cell(8, 3).value
            if industry_cell:
                industry = str(industry_cell)
        
        # Extract valuation data from Valuation sheet
        valuation_data = {
            'enterprise_value': 0,
            'equity_value': 0,
            'share_price': 0,
            'wacc': 0.10,
        }
        
        if "Valuation" in wb.sheetnames:
            val_sheet = wb["Valuation"]
            try:
                # Read specific cells: C19=WACC, C35=EV, C37=Equity, C40=Share Price
                wacc_val = val_sheet.cell(19, 3).value
                if wacc_val and isinstance(wacc_val, (int, float)):
                    valuation_data['wacc'] = float(wacc_val) if wacc_val <= 1 else wacc_val / 100
                
                ev_val = val_sheet.cell(35, 3).value
                if ev_val and isinstance(ev_val, (int, float)):
                    valuation_data['enterprise_value'] = float(ev_val)
                
                equity_val = val_sheet.cell(37, 3).value
                if equity_val and isinstance(equity_val, (int, float)):
                    valuation_data['equity_value'] = float(equity_val)
                
                price_val = val_sheet.cell(40, 3).value
                if price_val and isinstance(price_val, (int, float)):
                    valuation_data['share_price'] = float(price_val)
            except Exception as e:
                logger.warning(f"Error reading valuation data: {e}")
        
        # Extract assumptions from Assumptions sheet (Column 2 = labels, Column 3 = values)
        assumptions = {
            'revenue_growth': 0.10,
            'ebitda_margin': 0.20,
            'tax_rate': 0.25,
            'terminal_growth': 0.03,
            'risk_free_rate': 0.07,
            'beta': 1.0,
        }
        
        if "Assumptions" in wb.sheetnames:
            assump_sheet = wb["Assumptions"]
            try:
                keyword_map = {
                    'ebitda margin': 'ebitda_margin',
                    'terminal growth': 'terminal_growth',
                    'risk-free rate': 'risk_free_rate',
                    'beta': 'beta',
                    'tax rate': 'tax_rate',
                    'revenue growth': 'revenue_growth',
                }
                
                for row in range(1, 41):
                    label = assump_sheet.cell(row, 2).value
                    if label and isinstance(label, str):
                        label_lower = label.lower()
                        value_cell = assump_sheet.cell(row, 3).value
                        
                        if value_cell and isinstance(value_cell, (int, float)):
                            val = float(value_cell)
                            # Convert percentages to decimals
                            if val > 1 and any(x in label_lower for x in ["rate", "growth", "margin"]):
                                val = val / 100
                            
                            for keyword, key in keyword_map.items():
                                if keyword in label_lower:
                                    assumptions[key] = val
                                    break
            except Exception as e:
                logger.warning(f"Error reading assumptions: {e}")
        
        wb.close()
        
        # Generate the PDF
        return generate_pdf_report(
            output_path=output_path,
            company_name=company_name,
            industry=industry,
            valuation_data=valuation_data,
            assumptions=assumptions,
            commentary=None
        )
        
    except Exception as e:
        logger.error(f"Error creating PDF report from Excel: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return False


def is_available() -> bool:
    """Check if PDF export is available"""
    return REPORTLAB_AVAILABLE
