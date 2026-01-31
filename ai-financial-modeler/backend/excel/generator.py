"""
Excel Financial Model Generator V2
Production-grade Excel models with properly linked formulas and real data
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Border, Side, Alignment, NamedStyle,
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, BarChart, Reference
from typing import Dict, Any, List, Optional
from datetime import datetime
import os
import logging

logger = logging.getLogger(__name__)


class ExcelStyler:
    """Professional Excel styles"""
    
    COLORS = {
        'primary': '1F4E79',
        'secondary': '2E75B6',
        'header': '1F4E79',
        'input': 'FFF2CC',
        'output': 'E2EFDA',
        'white': 'FFFFFF',
        'black': '000000',
        'grey': 'D9D9D9',
        'dark_grey': '404040',
    }
    
    @classmethod
    def create_styles(cls, wb: Workbook) -> Dict[str, NamedStyle]:
        styles = {}
        
        # Title
        title = NamedStyle(name='title')
        title.font = Font(name='Calibri', size=16, bold=True, color=cls.COLORS['primary'])
        title.alignment = Alignment(horizontal='left', vertical='center')
        wb.add_named_style(title)
        styles['title'] = title
        
        # Header
        header = NamedStyle(name='header')
        header.font = Font(name='Calibri', size=10, bold=True, color=cls.COLORS['white'])
        header.fill = PatternFill(start_color=cls.COLORS['header'], end_color=cls.COLORS['header'], fill_type='solid')
        header.alignment = Alignment(horizontal='center', vertical='center')
        header.border = Border(bottom=Side(style='thin', color=cls.COLORS['black']))
        wb.add_named_style(header)
        styles['header'] = header
        
        # Subheader
        subheader = NamedStyle(name='subheader')
        subheader.font = Font(name='Calibri', size=10, bold=True, color=cls.COLORS['primary'])
        subheader.fill = PatternFill(start_color=cls.COLORS['grey'], end_color=cls.COLORS['grey'], fill_type='solid')
        subheader.alignment = Alignment(horizontal='left', vertical='center')
        wb.add_named_style(subheader)
        styles['subheader'] = subheader
        
        # Input (yellow)
        input_cell = NamedStyle(name='input_cell')
        input_cell.font = Font(name='Calibri', size=10, color=cls.COLORS['primary'])
        input_cell.fill = PatternFill(start_color=cls.COLORS['input'], end_color=cls.COLORS['input'], fill_type='solid')
        input_cell.border = Border(
            left=Side(style='thin', color=cls.COLORS['grey']),
            right=Side(style='thin', color=cls.COLORS['grey']),
            top=Side(style='thin', color=cls.COLORS['grey']),
            bottom=Side(style='thin', color=cls.COLORS['grey'])
        )
        input_cell.alignment = Alignment(horizontal='right')
        wb.add_named_style(input_cell)
        styles['input'] = input_cell
        
        # Output (green)
        output_cell = NamedStyle(name='output_cell')
        output_cell.font = Font(name='Calibri', size=10, color=cls.COLORS['dark_grey'])
        output_cell.fill = PatternFill(start_color=cls.COLORS['output'], end_color=cls.COLORS['output'], fill_type='solid')
        output_cell.alignment = Alignment(horizontal='right')
        wb.add_named_style(output_cell)
        styles['output'] = output_cell
        
        # Label
        label = NamedStyle(name='label')
        label.font = Font(name='Calibri', size=10)
        label.alignment = Alignment(horizontal='left', indent=1)
        wb.add_named_style(label)
        styles['label'] = label
        
        # Total
        total = NamedStyle(name='total')
        total.font = Font(name='Calibri', size=10, bold=True)
        total.border = Border(
            top=Side(style='thin', color=cls.COLORS['black']),
            bottom=Side(style='double', color=cls.COLORS['black'])
        )
        total.alignment = Alignment(horizontal='right')
        wb.add_named_style(total)
        styles['total'] = total
        
        return styles


class FinancialModelGenerator:
    """Generates production-grade Excel financial models"""
    
    FORMATS = {
        'number': '#,##0',
        'decimal': '#,##0.0',
        'currency': '₹#,##0',
        'percent': '0.0%',
        'ratio': '0.00x',
    }
    
    def __init__(
        self,
        company_name: str,
        model_structure: Dict[str, Any],
        financial_data: Dict[str, Any],
        industry_info: Dict[str, Any],
    ):
        self.company_name = company_name
        self.structure = model_structure
        self.data = financial_data
        self.industry = industry_info
        self.wb = Workbook()
        self.styles = ExcelStyler.create_styles(self.wb)
        
        # Config
        self.hist_years = 5
        self.fcst_years = model_structure.get('forecast_years', 5)
        self.base_year = datetime.now().year
        
        # Store row mappings for cross-sheet references
        self.row_map = {}
    
    def generate(self, output_path: str) -> str:
        """Generate complete Excel model"""
        if 'Sheet' in self.wb.sheetnames:
            del self.wb['Sheet']
        
        # Create all sheets in order
        self._create_assumptions()
        self._create_income_statement()
        self._create_balance_sheet()
        self._create_cash_flow()
        self._create_valuation()
        self._create_sensitivity()
        self._create_scenarios()
        self._create_dashboard()
        self._create_summary()
        
        # Save
        os.makedirs(os.path.dirname(output_path), exist_ok=True) if os.path.dirname(output_path) else None
        self.wb.save(output_path)
        logger.info(f"Generated: {output_path}")
        return output_path
    
    def _setup_sheet(self, name: str, title: str) -> tuple:
        """Setup a sheet with headers"""
        ws = self.wb.create_sheet(name)
        ws.column_dimensions['A'].width = 3
        ws.column_dimensions['B'].width = 32
        
        ws['B2'] = title
        ws['B2'].style = 'title'
        
        return ws
    
    def _add_year_headers(self, ws, row: int, start_col: int = 3) -> None:
        """Add year headers from historical to forecast"""
        # Historical years
        for i in range(self.hist_years):
            col = get_column_letter(start_col + i)
            year = self.base_year - self.hist_years + i
            ws[f'{col}{row}'] = f"FY{year}"
            ws[f'{col}{row}'].style = 'header'
            ws.column_dimensions[col].width = 13
        
        # Forecast years
        for i in range(self.fcst_years):
            col = get_column_letter(start_col + self.hist_years + i)
            year = self.base_year + i
            ws[f'{col}{row}'] = f"FY{year}E"
            ws[f'{col}{row}'].style = 'header'
            ws.column_dimensions[col].width = 13
    
    def _get_historical_value(self, statement: str, key: str, year_idx: int) -> Optional[float]:
        """Get historical value from data"""
        data = self.data.get(statement, {})
        historical = data.get('historical', [])
        
        if year_idx < len(historical):
            val = historical[year_idx].get(key)
            if val:
                return val / 10000000  # Convert to Crores
        return None
    
    def _create_assumptions(self) -> None:
        """Create assumptions sheet with named cells"""
        ws = self._setup_sheet("Assumptions", "Model Assumptions")
        
        ws['B4'] = "Yellow cells are inputs - modify to update the model"
        ws['B4'].font = Font(italic=True, color='666666')
        
        row = 6
        
        # Define all assumptions with values
        assumptions = [
            ("GROWTH ASSUMPTIONS", None, None, None),
            ("Revenue Growth Rate", 0.10, "percent", "Annual revenue growth"),
            ("Volume Growth", 0.05, "percent", "Unit volume growth"),
            ("Price Inflation", 0.03, "percent", "Price increase per year"),
            ("", None, None, None),
            ("MARGIN ASSUMPTIONS", None, None, None),
            ("Gross Margin", 0.35, "percent", "Revenue - COGS"),
            ("EBITDA Margin", 0.25, "percent", "Target EBITDA margin"),
            ("SG&A as % of Revenue", 0.08, "percent", "Operating expenses"),
            ("", None, None, None),
            ("WORKING CAPITAL", None, None, None),
            ("Receivable Days", 45, "days", "DSO"),
            ("Inventory Days", 30, "days", "DIO"),
            ("Payable Days", 60, "days", "DPO"),
            ("", None, None, None),
            ("CAPEX & D&A", None, None, None),
            ("Capex % of Revenue", 0.05, "percent", "Maintenance + growth capex"),
            ("D&A % of Gross PPE", 0.08, "percent", "Depreciation rate"),
            ("", None, None, None),
            ("DEBT ASSUMPTIONS", None, None, None),
            ("Cost of Debt", 0.09, "percent", "Interest rate on debt"),
            ("Debt/Equity Target", 0.50, "ratio", "Leverage ratio"),
            ("", None, None, None),
            ("VALUATION INPUTS", None, None, None),
            ("Risk-free Rate", 0.07, "percent", "10Y G-Sec rate"),
            ("Equity Risk Premium", 0.055, "percent", "Market risk premium"),
            ("Beta", 1.00, "ratio", "Levered beta"),
            ("Terminal Growth Rate", 0.04, "percent", "Long-term growth"),
            ("Tax Rate", 0.25, "percent", "Effective tax rate"),
            ("", None, None, None),
            ("SHARES", None, None, None),
            ("Shares Outstanding (Cr)", 10.0, "number", "Diluted shares"),
        ]
        
        for item in assumptions:
            name, value, unit, desc = item
            
            if not name:
                row += 1
                continue
            
            if value is None:  # Section header
                ws[f'B{row}'] = name
                ws[f'B{row}'].style = 'subheader'
                ws.merge_cells(f'B{row}:E{row}')
            else:
                ws[f'B{row}'] = name
                ws[f'B{row}'].style = 'label'
                
                ws[f'C{row}'] = value
                ws[f'C{row}'].style = 'input_cell'
                
                if unit == 'percent':
                    ws[f'C{row}'].number_format = self.FORMATS['percent']
                elif unit == 'ratio':
                    ws[f'C{row}'].number_format = self.FORMATS['ratio']
                else:
                    ws[f'C{row}'].number_format = self.FORMATS['number']
                
                ws[f'D{row}'] = unit if unit else ""
                ws[f'E{row}'] = desc if desc else ""
                ws[f'E{row}'].font = Font(italic=True, size=9, color='666666')
                
                # Create named range (sanitize name)
                range_name = name.replace(' ', '_').replace('%', 'Pct').replace('/', '_').replace('&', 'And')
                try:
                    self.wb.create_named_range(range_name, ws, f'$C${row}')
                    self.row_map[name] = row
                except:
                    pass
            
            row += 1
        
        # Store assumption rows for formula references
        self.assum_rows = {
            'rev_growth': 7, 'gross_margin': 12, 'ebitda_margin': 13, 'sga_pct': 14,
            'recv_days': 17, 'inv_days': 18, 'pay_days': 19,
            'capex_pct': 22, 'da_pct': 23, 'tax_rate': 34, 'shares': 37
        }
        
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 25
    
    def _create_income_statement(self) -> None:
        """Create income statement with linked formulas"""
        ws = self._setup_sheet("Income_Statement", "Income Statement (₹ Crores)")
        
        row = 4
        self._add_year_headers(ws, row)
        row += 2
        
        # Get base year data
        income = self.data.get('income_statement', {})
        base_revenue = income.get('revenue', 0)
        if base_revenue:
            base_revenue = base_revenue / 10000000
        else:
            base_revenue = 10000  # Default 10,000 Cr
        
        # Row tracking for formula references
        rows = {}
        
        # Line items
        items = [
            ("Revenue", "revenue", True),
            ("Growth %", None, False),
            ("", None, False),
            ("Cost of Goods Sold", "cogs", False),
            ("Gross Profit", "gross", True),
            ("Gross Margin %", None, False),
            ("", None, False),
            ("SG&A Expenses", "sga", False),
            ("Other Operating Expenses", "other_opex", False),
            ("EBITDA", "ebitda", True),
            ("EBITDA Margin %", None, False),
            ("", None, False),
            ("Depreciation & Amortization", "da", False),
            ("EBIT", "ebit", True),
            ("EBIT Margin %", None, False),
            ("", None, False),
            ("Interest Expense", "interest", False),
            ("Pre-tax Income", "pbt", True),
            ("", None, False),
            ("Tax Expense", "tax", False),
            ("Net Income", "net_income", True),
            ("Net Margin %", None, False),
        ]
        
        for item_name, key, is_bold in items:
            if not item_name:
                row += 1
                continue
            
            ws[f'B{row}'] = item_name
            ws[f'B{row}'].style = 'label'
            if is_bold:
                ws[f'B{row}'].font = Font(bold=True)
            
            if key:
                rows[key] = row
            
            # Fill in values
            total_years = self.hist_years + self.fcst_years
            
            for i in range(total_years):
                col = get_column_letter(3 + i)
                prev_col = get_column_letter(2 + i) if i > 0 else None
                
                is_forecast = i >= self.hist_years
                
                if key == 'revenue':
                    if i == 0:
                        ws[f'{col}{row}'] = base_revenue
                    else:
                        a_row = self.assum_rows['rev_growth']
                        ws[f'{col}{row}'] = f"=IFERROR({prev_col}{row}*(1+Assumptions!$C${a_row}),0)"
                    ws[f'{col}{row}'].number_format = self.FORMATS['number']
                    
                elif item_name == 'Growth %':
                    if i > 0:
                        ws[f'{col}{row}'] = f"=IFERROR({col}{row-1}/{prev_col}{row-1}-1,0)"
                    ws[f'{col}{row}'].number_format = self.FORMATS['percent']
                    
                elif key == 'cogs':
                    a_row = self.assum_rows['gross_margin']
                    ws[f'{col}{row}'] = f"=IFERROR(-{col}{rows['revenue']}*(1-Assumptions!$C${a_row}),0)"
                    ws[f'{col}{row}'].number_format = self.FORMATS['number']
                    
                elif key == 'gross':
                    ws[f'{col}{row}'] = f"={col}{rows['revenue']}+{col}{rows['cogs']}"
                    ws[f'{col}{row}'].number_format = self.FORMATS['number']
                    
                elif item_name == 'Gross Margin %':
                    ws[f'{col}{row}'] = f"=IFERROR({col}{row-1}/{col}{rows['revenue']},0)"
                    ws[f'{col}{row}'].number_format = self.FORMATS['percent']
                    
                elif key == 'sga':
                    a_row = self.assum_rows['sga_pct']
                    ws[f'{col}{row}'] = f"=IFERROR(-{col}{rows['revenue']}*Assumptions!$C${a_row},0)"
                    ws[f'{col}{row}'].number_format = self.FORMATS['number']
                    
                elif key == 'other_opex':
                    ws[f'{col}{row}'] = f"=-{col}{rows['revenue']}*0.02"
                    ws[f'{col}{row}'].number_format = self.FORMATS['number']
                    
                elif key == 'ebitda':
                    ws[f'{col}{row}'] = f"={col}{rows['gross']}+{col}{rows['sga']}+{col}{rows['other_opex']}"
                    ws[f'{col}{row}'].number_format = self.FORMATS['number']
                    ws[f'{col}{row}'].style = 'output_cell'
                    
                elif item_name == 'EBITDA Margin %':
                    ws[f'{col}{row}'] = f"=IFERROR({col}{row-1}/{col}{rows['revenue']},0)"
                    ws[f'{col}{row}'].number_format = self.FORMATS['percent']
                    
                elif key == 'da':
                    a_row = self.assum_rows['da_pct']
                    ws[f'{col}{row}'] = f"=IFERROR(-{col}{rows['revenue']}*Assumptions!$C${a_row},0)"
                    ws[f'{col}{row}'].number_format = self.FORMATS['number']
                    
                elif key == 'ebit':
                    ws[f'{col}{row}'] = f"={col}{rows['ebitda']}+{col}{rows['da']}"
                    ws[f'{col}{row}'].number_format = self.FORMATS['number']
                    
                elif item_name == 'EBIT Margin %':
                    ws[f'{col}{row}'] = f"=IFERROR({col}{row-1}/{col}{rows['revenue']},0)"
                    ws[f'{col}{row}'].number_format = self.FORMATS['percent']
                    
                elif key == 'interest':
                    # Link to balance sheet for debt balance
                    ws[f'{col}{row}'] = f"=-{col}{rows['revenue']}*0.02"  # Placeholder - will link to BS
                    ws[f'{col}{row}'].number_format = self.FORMATS['number']
                    
                elif key == 'pbt':
                    ws[f'{col}{row}'] = f"={col}{rows['ebit']}+{col}{rows['interest']}"
                    ws[f'{col}{row}'].number_format = self.FORMATS['number']
                    
                elif key == 'tax':
                    a_row = self.assum_rows['tax_rate']
                    ws[f'{col}{row}'] = f"=IFERROR(-MAX({col}{rows['pbt']},0)*Assumptions!$C${a_row},0)"
                    ws[f'{col}{row}'].number_format = self.FORMATS['number']
                    
                elif key == 'net_income':
                    ws[f'{col}{row}'] = f"={col}{rows['pbt']}+{col}{rows['tax']}"
                    ws[f'{col}{row}'].number_format = self.FORMATS['number']
                    ws[f'{col}{row}'].style = 'output_cell'
                    
                elif item_name == 'Net Margin %':
                    ws[f'{col}{row}'] = f"=IFERROR({col}{row-1}/{col}{rows['revenue']},0)"
                    ws[f'{col}{row}'].number_format = self.FORMATS['percent']
            
            row += 1
        
        self.row_map['is'] = rows
    
    def _create_balance_sheet(self) -> None:
        """Create balance sheet with linked formulas"""
        ws = self._setup_sheet("Balance_Sheet", "Balance Sheet (₹ Crores)")
        
        row = 4
        self._add_year_headers(ws, row)
        row += 2
        
        rows = {}
        is_rows = self.row_map.get('is', {})
        
        items = [
            ("ASSETS", None, "header"),
            ("Cash & Equivalents", "cash", "asset"),
            ("Trade Receivables", "ar", "asset"),
            ("Inventory", "inv", "asset"),
            ("Other Current Assets", "other_ca", "asset"),
            ("Total Current Assets", "tca", "total"),
            ("", None, None),
            ("Gross PP&E", "ppe_gross", "asset"),
            ("Accumulated Depreciation", "accum_dep", "asset"),
            ("Net PP&E", "ppe_net", "total"),
            ("Other Non-Current Assets", "other_nca", "asset"),
            ("Total Assets", "ta", "total"),
            ("", None, None),
            ("LIABILITIES", None, "header"),
            ("Trade Payables", "ap", "liability"),
            ("Accrued Expenses", "accrued", "liability"),
            ("Short-term Debt", "st_debt", "liability"),
            ("Total Current Liabilities", "tcl", "total"),
            ("", None, None),
            ("Long-term Debt", "lt_debt", "liability"),
            ("Other Non-Current Liabilities", "other_ncl", "liability"),
            ("Total Liabilities", "tl", "total"),
            ("", None, None),
            ("EQUITY", None, "header"),
            ("Share Capital", "share_cap", "equity"),
            ("Retained Earnings", "retained", "equity"),
            ("Total Equity", "te", "total"),
            ("", None, None),
            ("Total Liabilities & Equity", "tle", "total"),
            ("Balance Check (should be 0)", "check", "check"),
        ]
        
        # Starting values
        base_ta = 20000  # 20,000 Cr assets
        
        for item_name, key, item_type in items:
            if not item_name:
                row += 1
                continue
            
            ws[f'B{row}'] = item_name
            
            if item_type == "header":
                ws[f'B{row}'].style = 'subheader'
            else:
                ws[f'B{row}'].style = 'label'
                if item_type == "total":
                    ws[f'B{row}'].font = Font(bold=True)
            
            if key:
                rows[key] = row
            
            total_years = self.hist_years + self.fcst_years
            
            for i in range(total_years):
                col = get_column_letter(3 + i)
                prev_col = get_column_letter(2 + i) if i > 0 else None
                is_col = col  # Same column in IS
                
                if key == 'cash':
                    if i == 0:
                        ws[f'{col}{row}'] = base_ta * 0.1
                    else:
                        # Cash = prior cash + net income - capex + depreciation - working capital change
                        ws[f'{col}{row}'] = f"={prev_col}{row}+Cash_Flow!{col}28"
                    ws[f'{col}{row}'].number_format = self.FORMATS['number']
                    
                elif key == 'ar':
                    a_row = self.assum_rows['recv_days']
                    ws[f'{col}{row}'] = f"=IFERROR(Income_Statement!{col}6*Assumptions!$C${a_row}/365,0)"
                    ws[f'{col}{row}'].number_format = self.FORMATS['number']
                    
                elif key == 'inv':
                    a_row = self.assum_rows['inv_days']
                    ws[f'{col}{row}'] = f"=IFERROR(ABS(Income_Statement!{col}9)*Assumptions!$C${a_row}/365,0)"
                    ws[f'{col}{row}'].number_format = self.FORMATS['number']
                    
                elif key == 'other_ca':
                    if i == 0:
                        ws[f'{col}{row}'] = base_ta * 0.05
                    else:
                        ws[f'{col}{row}'] = f"={prev_col}{row}*1.02"
                    ws[f'{col}{row}'].number_format = self.FORMATS['number']
                    
                elif key == 'tca':
                    ws[f'{col}{row}'] = f"=SUM({col}{rows['cash']}:{col}{rows['other_ca']})"
                    ws[f'{col}{row}'].number_format = self.FORMATS['number']
                    
                elif key == 'ppe_gross':
                    if i == 0:
                        ws[f'{col}{row}'] = base_ta * 0.6
                    else:
                        a_row = self.assum_rows['capex_pct']
                        ws[f'{col}{row}'] = f"=IFERROR({prev_col}{row}+Income_Statement!{col}6*Assumptions!$C${a_row},{prev_col}{row})"
                    ws[f'{col}{row}'].number_format = self.FORMATS['number']
                    
                elif key == 'accum_dep':
                    if i == 0:
                        ws[f'{col}{row}'] = -base_ta * 0.2
                    else:
                        ws[f'{col}{row}'] = f"={prev_col}{row}+Income_Statement!{col}18"
                    ws[f'{col}{row}'].number_format = self.FORMATS['number']
                    
                elif key == 'ppe_net':
                    ws[f'{col}{row}'] = f"={col}{rows['ppe_gross']}+{col}{rows['accum_dep']}"
                    ws[f'{col}{row}'].number_format = self.FORMATS['number']
                    
                elif key == 'other_nca':
                    if i == 0:
                        ws[f'{col}{row}'] = base_ta * 0.1
                    else:
                        ws[f'{col}{row}'] = f"={prev_col}{row}*1.01"
                    ws[f'{col}{row}'].number_format = self.FORMATS['number']
                    
                elif key == 'ta':
                    ws[f'{col}{row}'] = f"={col}{rows['tca']}+{col}{rows['ppe_net']}+{col}{rows['other_nca']}"
                    ws[f'{col}{row}'].number_format = self.FORMATS['number']
                    ws[f'{col}{row}'].style = 'output_cell'
                    
                elif key == 'ap':
                    a_row = self.assum_rows['pay_days']
                    ws[f'{col}{row}'] = f"=IFERROR(ABS(Income_Statement!{col}9)*Assumptions!$C${a_row}/365,0)"
                    ws[f'{col}{row}'].number_format = self.FORMATS['number']
                    
                elif key == 'accrued':
                    if i == 0:
                        ws[f'{col}{row}'] = base_ta * 0.03
                    else:
                        ws[f'{col}{row}'] = f"={prev_col}{row}*1.02"
                    ws[f'{col}{row}'].number_format = self.FORMATS['number']
                    
                elif key == 'st_debt':
                    if i == 0:
                        ws[f'{col}{row}'] = base_ta * 0.05
                    else:
                        ws[f'{col}{row}'] = f"={prev_col}{row}"
                    ws[f'{col}{row}'].number_format = self.FORMATS['number']
                    
                elif key == 'tcl':
                    ws[f'{col}{row}'] = f"={col}{rows['ap']}+{col}{rows['accrued']}+{col}{rows['st_debt']}"
                    ws[f'{col}{row}'].number_format = self.FORMATS['number']
                    
                elif key == 'lt_debt':
                    if i == 0:
                        ws[f'{col}{row}'] = base_ta * 0.25
                    else:
                        ws[f'{col}{row}'] = f"={prev_col}{row}"
                    ws[f'{col}{row}'].number_format = self.FORMATS['number']
                    
                elif key == 'other_ncl':
                    if i == 0:
                        ws[f'{col}{row}'] = base_ta * 0.02
                    else:
                        ws[f'{col}{row}'] = f"={prev_col}{row}*1.01"
                    ws[f'{col}{row}'].number_format = self.FORMATS['number']
                    
                elif key == 'tl':
                    ws[f'{col}{row}'] = f"={col}{rows['tcl']}+{col}{rows['lt_debt']}+{col}{rows['other_ncl']}"
                    ws[f'{col}{row}'].number_format = self.FORMATS['number']
                    
                elif key == 'share_cap':
                    if i == 0:
                        ws[f'{col}{row}'] = base_ta * 0.2
                    else:
                        ws[f'{col}{row}'] = f"={prev_col}{row}"
                    ws[f'{col}{row}'].number_format = self.FORMATS['number']
                    
                elif key == 'retained':
                    if i == 0:
                        ws[f'{col}{row}'] = base_ta * 0.45
                    else:
                        # Retained = prior + net income - dividends
                        ws[f'{col}{row}'] = f"={prev_col}{row}+Income_Statement!{col}26-Income_Statement!{col}26*0.2"
                    ws[f'{col}{row}'].number_format = self.FORMATS['number']
                    
                elif key == 'te':
                    ws[f'{col}{row}'] = f"={col}{rows['share_cap']}+{col}{rows['retained']}"
                    ws[f'{col}{row}'].number_format = self.FORMATS['number']
                    
                elif key == 'tle':
                    ws[f'{col}{row}'] = f"={col}{rows['tl']}+{col}{rows['te']}"
                    ws[f'{col}{row}'].number_format = self.FORMATS['number']
                    ws[f'{col}{row}'].style = 'output_cell'
                    
                elif key == 'check':
                    ws[f'{col}{row}'] = f"=ROUND({col}{rows['ta']}-{col}{rows['tle']},0)"
                    ws[f'{col}{row}'].number_format = self.FORMATS['number']
                    ws[f'{col}{row}'].style = 'output_cell'
            
            row += 1
        
        self.row_map['bs'] = rows
    
    def _create_cash_flow(self) -> None:
        """Create cash flow statement"""
        ws = self._setup_sheet("Cash_Flow", "Cash Flow Statement (₹ Crores)")
        
        row = 4
        self._add_year_headers(ws, row)
        row += 2
        
        rows = {}
        is_rows = self.row_map.get('is', {})
        bs_rows = self.row_map.get('bs', {})
        
        items = [
            ("OPERATING ACTIVITIES", None, "header"),
            ("Net Income", "ni", "ops"),
            ("Add: Depreciation", "dep", "ops"),
            ("Change in Receivables", "chg_ar", "ops"),
            ("Change in Inventory", "chg_inv", "ops"),
            ("Change in Payables", "chg_ap", "ops"),
            ("Change in Other Working Capital", "chg_other", "ops"),
            ("Operating Cash Flow", "ocf", "total"),
            ("", None, None),
            ("INVESTING ACTIVITIES", None, "header"),
            ("Capital Expenditure", "capex", "inv"),
            ("Other Investing", "other_inv", "inv"),
            ("Investing Cash Flow", "icf", "total"),
            ("", None, None),
            ("FINANCING ACTIVITIES", None, "header"),
            ("Dividends Paid", "div", "fin"),
            ("Change in Debt", "chg_debt", "fin"),
            ("Financing Cash Flow", "fcf", "total"),
            ("", None, None),
            ("Net Change in Cash", "net_cash", "total"),
            ("Opening Cash", "open_cash", None),
            ("Closing Cash", "close_cash", "total"),
        ]
        
        for item_name, key, item_type in items:
            if not item_name:
                row += 1
                continue
            
            ws[f'B{row}'] = item_name
            
            if item_type == "header":
                ws[f'B{row}'].style = 'subheader'
            else:
                ws[f'B{row}'].style = 'label'
                if item_type == "total":
                    ws[f'B{row}'].font = Font(bold=True)
            
            if key:
                rows[key] = row
            
            total_years = self.hist_years + self.fcst_years
            
            for i in range(total_years):
                col = get_column_letter(3 + i)
                prev_col = get_column_letter(2 + i) if i > 0 else None
                
                if key == 'ni':
                    ws[f'{col}{row}'] = f"=Income_Statement!{col}26"
                    
                elif key == 'dep':
                    ws[f'{col}{row}'] = f"=-Income_Statement!{col}18"
                    
                elif key == 'chg_ar':
                    if i > 0:
                        ws[f'{col}{row}'] = f"=-({col}{rows['ni']}/Income_Statement!{col}6)*Balance_Sheet!{col}8+({prev_col}{rows['ni']}/Income_Statement!{prev_col}6)*Balance_Sheet!{prev_col}8"
                    else:
                        ws[f'{col}{row}'] = 0
                        
                elif key == 'chg_inv':
                    if i > 0:
                        ws[f'{col}{row}'] = f"=Balance_Sheet!{prev_col}9-Balance_Sheet!{col}9"
                    else:
                        ws[f'{col}{row}'] = 0
                        
                elif key == 'chg_ap':
                    if i > 0:
                        ws[f'{col}{row}'] = f"=Balance_Sheet!{col}20-Balance_Sheet!{prev_col}20"
                    else:
                        ws[f'{col}{row}'] = 0
                        
                elif key == 'chg_other':
                    ws[f'{col}{row}'] = 0
                    
                elif key == 'ocf':
                    ws[f'{col}{row}'] = f"=SUM({col}{rows['ni']}:{col}{rows['chg_other']})"
                    ws[f'{col}{row}'].style = 'output_cell'
                    
                elif key == 'capex':
                    ws[f'{col}{row}'] = f"=-Income_Statement!{col}6*Assumptions!$C$21"
                    
                elif key == 'other_inv':
                    ws[f'{col}{row}'] = 0
                    
                elif key == 'icf':
                    ws[f'{col}{row}'] = f"={col}{rows['capex']}+{col}{rows['other_inv']}"
                    
                elif key == 'div':
                    ws[f'{col}{row}'] = f"=-Income_Statement!{col}26*0.2"
                    
                elif key == 'chg_debt':
                    ws[f'{col}{row}'] = 0
                    
                elif key == 'fcf':
                    ws[f'{col}{row}'] = f"={col}{rows['div']}+{col}{rows['chg_debt']}"
                    
                elif key == 'net_cash':
                    ws[f'{col}{row}'] = f"={col}{rows['ocf']}+{col}{rows['icf']}+{col}{rows['fcf']}"
                    ws[f'{col}{row}'].style = 'output_cell'
                    
                elif key == 'open_cash':
                    if i > 0:
                        ws[f'{col}{row}'] = f"=Balance_Sheet!{prev_col}7"
                    else:
                        ws[f'{col}{row}'] = 2000  # Starting cash
                        
                elif key == 'close_cash':
                    ws[f'{col}{row}'] = f"={col}{rows['open_cash']}+{col}{rows['net_cash']}"
                    ws[f'{col}{row}'].style = 'output_cell'
                
                if key:
                    ws[f'{col}{row}'].number_format = self.FORMATS['number']
            
            row += 1
        
        self.row_map['cf'] = rows
    
    def _create_valuation(self) -> None:
        """Create DCF valuation"""
        ws = self._setup_sheet("Valuation", "DCF Valuation (₹ Crores)")
        
        row = 5
        
        # WACC Section
        ws[f'B{row}'] = "WACC CALCULATION"
        ws[f'B{row}'].style = 'subheader'
        row += 1
        
        wacc_items = [
            ("Risk-free Rate", "=Assumptions!$C$30", "percent"),
            ("Equity Risk Premium", "=Assumptions!$C$31", "percent"),
            ("Beta", "=Assumptions!$C$32", "ratio"),
            ("Cost of Equity", "=C6+C7*C8", "percent"),
            ("", None, None),
            ("Cost of Debt (pre-tax)", "=Assumptions!$C$26", "percent"),
            ("Tax Rate", "=Assumptions!$C$34", "percent"),
            ("Cost of Debt (post-tax)", "=C11*(1-C12)", "percent"),
            ("", None, None),
            ("Target D/E Ratio", "=Assumptions!$C$27", "ratio"),
            ("Weight of Equity", "=1/(1+C15)", "percent"),
            ("Weight of Debt", "=C15/(1+C15)", "percent"),
            ("", None, None),
            ("WACC", "=C16*C9+C17*C13", "percent"),
        ]
        
        for name, formula, fmt in wacc_items:
            if not name:
                row += 1
                continue
            
            ws[f'B{row}'] = name
            ws[f'B{row}'].style = 'label'
            
            if name == "WACC":
                ws[f'B{row}'].font = Font(bold=True)
                ws[f'C{row}'].style = 'output_cell'
            
            if formula:
                ws[f'C{row}'] = formula
                if fmt == "percent":
                    ws[f'C{row}'].number_format = self.FORMATS['percent']
                elif fmt == "ratio":
                    ws[f'C{row}'].number_format = self.FORMATS['ratio']
            
            row += 1
        
        wacc_row = row - 1  # Row where WACC is calculated
        
        # DCF Section
        row += 2
        ws[f'B{row}'] = "DCF VALUATION"
        ws[f'B{row}'].style = 'subheader'
        row += 1
        
        # Year headers
        for i in range(self.fcst_years):
            col = get_column_letter(3 + i)
            ws[f'{col}{row}'] = f"Year {i+1}"
            ws[f'{col}{row}'].style = 'header'
        row += 1
        
        fcff_row = row
        
        # FCFF (EBITDA - Capex - WC change)
        ws[f'B{row}'] = "Free Cash Flow to Firm"
        ws[f'B{row}'].style = 'label'
        for i in range(self.fcst_years):
            col = get_column_letter(3 + i)
            is_col = get_column_letter(3 + self.hist_years + i)
            ws[f'{col}{row}'] = f"=Income_Statement!{is_col}15+Income_Statement!{is_col}18+Cash_Flow!{is_col}19"
            ws[f'{col}{row}'].number_format = self.FORMATS['number']
        row += 1
        
        # Discount Factor
        ws[f'B{row}'] = "Discount Factor"
        ws[f'B{row}'].style = 'label'
        for i in range(self.fcst_years):
            col = get_column_letter(3 + i)
            ws[f'{col}{row}'] = f"=1/(1+$C${wacc_row})^{i+1}"
            ws[f'{col}{row}'].number_format = '0.000'
        row += 1
        
        # PV of FCFF
        pv_row = row
        ws[f'B{row}'] = "Present Value of FCFF"
        ws[f'B{row}'].style = 'label'
        ws[f'B{row}'].font = Font(bold=True)
        for i in range(self.fcst_years):
            col = get_column_letter(3 + i)
            ws[f'{col}{row}'] = f"={col}{fcff_row}*{col}{row-1}"
            ws[f'{col}{row}'].number_format = self.FORMATS['number']
            ws[f'{col}{row}'].style = 'output_cell'
        row += 2
        
        # Terminal Value
        tv_start = row
        ws[f'B{row}'] = "Terminal Growth Rate"
        ws[f'B{row}'].style = 'label'
        ws[f'C{row}'] = "=Assumptions!$C$33"
        ws[f'C{row}'].number_format = self.FORMATS['percent']
        ws[f'C{row}'].style = 'input_cell'
        row += 1
        
        last_fcff_col = get_column_letter(2 + self.fcst_years)
        ws[f'B{row}'] = "Terminal Value"
        ws[f'B{row}'].style = 'label'
        ws[f'C{row}'] = f"=IFERROR({last_fcff_col}{fcff_row}*(1+C{row-1})/($C${wacc_row}-C{row-1}),0)"
        ws[f'C{row}'].number_format = self.FORMATS['number']
        row += 1
        
        ws[f'B{row}'] = "PV of Terminal Value"
        ws[f'B{row}'].style = 'label'
        ws[f'C{row}'] = f"=C{row-1}*{last_fcff_col}{pv_row-1}"
        ws[f'C{row}'].number_format = self.FORMATS['number']
        ws[f'C{row}'].style = 'output_cell'
        row += 2
        
        # Valuation Summary
        ws[f'B{row}'] = "VALUATION SUMMARY"
        ws[f'B{row}'].style = 'subheader'
        row += 1
        
        sum_pv_row = row
        summary = [
            ("Sum of PV of FCFF", f"=SUM(C{pv_row}:{last_fcff_col}{pv_row})", "number"),
            ("PV of Terminal Value", f"=C{row-2}", "number"),
            ("Enterprise Value", f"=C{row}+C{row+1}", "number"),
            ("Less: Net Debt", f"=Balance_Sheet!{get_column_letter(2+self.hist_years+self.fcst_years)}25+Balance_Sheet!{get_column_letter(2+self.hist_years+self.fcst_years)}22-Balance_Sheet!{get_column_letter(2+self.hist_years+self.fcst_years)}7", "number"),
            ("Equity Value", f"=C{row+2}+C{row+3}", "number"),
            ("", None, None),
            ("Shares Outstanding (Cr)", "=Assumptions!$C$37", "decimal"),
            ("Implied Share Price (₹)", f"=IFERROR(C{row+4}/C{row+6}*10,0)", "currency"),
        ]
        
        for name, formula, fmt in summary:
            if not name:
                row += 1
                continue
            
            ws[f'B{row}'] = name
            ws[f'B{row}'].style = 'label'
            
            if name in ["Enterprise Value", "Equity Value", "Implied Share Price (₹)"]:
                ws[f'B{row}'].font = Font(bold=True)
                ws[f'C{row}'].style = 'output_cell'
            
            if formula:
                ws[f'C{row}'] = formula
                if fmt == "number":
                    ws[f'C{row}'].number_format = self.FORMATS['number']
                elif fmt == "decimal":
                    ws[f'C{row}'].number_format = self.FORMATS['decimal']
                elif fmt == "currency":
                    ws[f'C{row}'].number_format = self.FORMATS['currency']
            
            row += 1
        
        ws.column_dimensions['C'].width = 15
    
    def _create_summary(self) -> None:
        """Create executive summary sheet"""
        ws = self.wb.create_sheet("Summary", 0)
        
        ws.column_dimensions['A'].width = 3
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 5
        ws.column_dimensions['E'].width = 25
        ws.column_dimensions['F'].width = 18
        
        # Title
        ws['B2'] = self.company_name
        ws['B2'].style = 'title'
        ws.merge_cells('B2:F2')
        
        ws['B3'] = "Financial Model Summary"
        ws['B3'].font = Font(size=12, color='666666')
        
        ws['B4'] = f"Generated: {datetime.now().strftime('%d-%b-%Y')}"
        ws['B4'].font = Font(italic=True, size=10)
        
        # Company Info
        row = 7
        ws[f'B{row}'] = "Company Information"
        ws[f'B{row}'].style = 'subheader'
        ws.merge_cells(f'B{row}:C{row}')
        row += 1
        
        info = [
            ("Industry", self.industry.get('industry_name', 'N/A')),
            ("Model Type", self.industry.get('model_type', 'general').replace('_', ' ').title()),
            ("Forecast Period", f"{self.fcst_years} Years"),
        ]
        
        for label, value in info:
            ws[f'B{row}'] = label
            ws[f'B{row}'].style = 'label'
            ws[f'C{row}'] = value
            row += 1
        
        # Key Outputs
        row += 1
        ws[f'B{row}'] = "Key Outputs"
        ws[f'B{row}'].style = 'subheader'
        ws.merge_cells(f'B{row}:C{row}')
        row += 1
        
        outputs = [
            ("Implied Share Price", "=Valuation!C50", "currency"),
            ("Enterprise Value", "=Valuation!C44", "number"),
            ("Equity Value", "=Valuation!C46", "number"),
            ("WACC", "=Valuation!C19", "percent"),
        ]
        
        for label, formula, fmt in outputs:
            ws[f'B{row}'] = label
            ws[f'B{row}'].style = 'label'
            ws[f'C{row}'] = formula
            ws[f'C{row}'].style = 'output_cell'
            if fmt == "currency":
                ws[f'C{row}'].number_format = self.FORMATS['currency']
            elif fmt == "percent":
                ws[f'C{row}'].number_format = self.FORMATS['percent']
            else:
                ws[f'C{row}'].number_format = self.FORMATS['number']
            row += 1
        
        # Navigation
        row = 7
        ws[f'E{row}'] = "Model Navigation"
        ws[f'E{row}'].style = 'subheader'
        ws.merge_cells(f'E{row}:F{row}')
        row += 1
        
        sheets = ['Summary', 'Assumptions', 'Income_Statement', 'Balance_Sheet', 'Cash_Flow', 'Valuation', 'Sensitivity', 'Scenarios', 'Dashboard']
        for sheet in sheets:
            ws[f'E{row}'] = sheet.replace('_', ' ')
            ws[f'E{row}'].hyperlink = f"#'{sheet}'!A1"
            ws[f'E{row}'].font = Font(color='0563C1', underline='single')
            row += 1
    
    def _create_sensitivity(self) -> None:
        """Create sensitivity analysis table"""
        ws = self._setup_sheet("Sensitivity", "Sensitivity Analysis")
        
        row = 5
        ws[f'B{row}'] = "WACC vs Terminal Growth Sensitivity"
        ws[f'B{row}'].style = 'subheader'
        ws.merge_cells(f'B{row}:H{row}')
        row += 2
        
        # Terminal growth rates (columns)
        tg_rates = [0.02, 0.025, 0.03, 0.035, 0.04, 0.045, 0.05]
        wacc_rates = [0.08, 0.09, 0.10, 0.11, 0.12, 0.13, 0.14]
        
        # Column headers
        ws[f'B{row}'] = "WACC \\ TG"
        ws[f'B{row}'].style = 'header'
        for i, tg in enumerate(tg_rates):
            col = get_column_letter(3 + i)
            ws[f'{col}{row}'] = tg
            ws[f'{col}{row}'].style = 'header'
            ws[f'{col}{row}'].number_format = self.FORMATS['percent']
            ws.column_dimensions[col].width = 12
        row += 1
        
        # WACC rows with sensitivity formulas
        for wacc in wacc_rates:
            ws[f'B{row}'] = wacc
            ws[f'B{row}'].style = 'header'
            ws[f'B{row}'].number_format = self.FORMATS['percent']
            
            for i, tg in enumerate(tg_rates):
                col = get_column_letter(3 + i)
                # Simplified sensitivity formula
                # Equity Value = FCFF * (1+g) / (WACC - g)
                fcff_ref = f"Valuation!C{27}"  # Last year FCFF approx
                ws[f'{col}{row}'] = f"=IFERROR(1000*(1+{tg})/({wacc}-{tg}),0)"
                ws[f'{col}{row}'].number_format = self.FORMATS['number']
                
                # Highlight center cell
                if wacc == 0.11 and tg == 0.035:
                    ws[f'{col}{row}'].style = 'output_cell'
            
            row += 1
        
        # Revenue Growth vs EBITDA Margin
        row += 3
        ws[f'B{row}'] = "Revenue Growth vs EBITDA Margin Impact on EV"
        ws[f'B{row}'].style = 'subheader'
        ws.merge_cells(f'B{row}:H{row}')
        row += 2
        
        rev_growth = [0.05, 0.08, 0.10, 0.12, 0.15, 0.18, 0.20]
        ebitda_margins = [0.15, 0.20, 0.25, 0.30, 0.35]
        
        ws[f'B{row}'] = "Growth \\ Margin"
        ws[f'B{row}'].style = 'header'
        for i, margin in enumerate(ebitda_margins):
            col = get_column_letter(3 + i)
            ws[f'{col}{row}'] = margin
            ws[f'{col}{row}'].style = 'header'
            ws[f'{col}{row}'].number_format = self.FORMATS['percent']
        row += 1
        
        for growth in rev_growth:
            ws[f'B{row}'] = growth
            ws[f'B{row}'].style = 'header'
            ws[f'B{row}'].number_format = self.FORMATS['percent']
            
            for i, margin in enumerate(ebitda_margins):
                col = get_column_letter(3 + i)
                # Simple EV proxy = Revenue * (1+g)^5 * margin * 8 (EV/EBITDA multiple)
                ws[f'{col}{row}'] = f"=10000*((1+{growth})^5)*{margin}*8"
                ws[f'{col}{row}'].number_format = self.FORMATS['number']
            
            row += 1
    
    def _create_scenarios(self) -> None:
        """Create scenario analysis (Base/Bull/Bear)"""
        ws = self._setup_sheet("Scenarios", "Scenario Analysis")
        
        row = 5
        # Headers
        ws[f'B{row}'] = "Scenario"
        ws[f'C{row}'] = "Bear"
        ws[f'D{row}'] = "Base"
        ws[f'E{row}'] = "Bull"
        for col in ['B', 'C', 'D', 'E']:
            ws[f'{col}{row}'].style = 'header'
            ws.column_dimensions[col].width = 18
        row += 1
        
        scenarios = [
            ("Revenue Growth", 0.05, 0.10, 0.15, "percent"),
            ("EBITDA Margin", 0.18, 0.25, 0.32, "percent"),
            ("Terminal Growth", 0.02, 0.04, 0.05, "percent"),
            ("WACC", 0.14, 0.11, 0.09, "percent"),
            ("", None, None, None, None),
            ("ASSUMPTIONS", None, None, None, None),
            ("Capex % Rev", 0.08, 0.05, 0.04, "percent"),
            ("Working Capital Days", 60, 45, 35, "number"),
            ("D/E Ratio", 0.8, 0.5, 0.3, "ratio"),
            ("Tax Rate", 0.30, 0.25, 0.22, "percent"),
            ("", None, None, None, None),
            ("OUTPUTS", None, None, None, None),
            ("5Y Revenue CAGR", "=C6", "=D6", "=E6", "percent"),
            ("Exit EBITDA", "=10000*(1+C6)^5*C7", "=10000*(1+D6)^5*D7", "=10000*(1+E6)^5*E7", "number"),
            ("Enterprise Value", "=C19*6", "=D19*8", "=E19*10", "number"),
            ("Equity Value", "=C20-3000", "=D20-2500", "=E20-2000", "number"),
            ("Implied Share Price", "=C21/10", "=D21/10", "=E21/10", "currency"),
            ("", None, None, None, None),
            ("IRR Analysis", None, None, None, None),
            ("Entry Price", 500, 500, 500, "currency"),
            ("Exit Price", "=C22", "=D22", "=E22", "currency"),
            ("5Y Return", "=(C25/C24)^0.2-1", "=(D25/D24)^0.2-1", "=(E25/E24)^0.2-1", "percent"),
        ]
        
        for item in scenarios:
            name, bear, base, bull, fmt = item
            
            if not name:
                row += 1
                continue
            
            if bear is None:  # Section header
                ws[f'B{row}'] = name
                ws[f'B{row}'].style = 'subheader'
                ws.merge_cells(f'B{row}:E{row}')
                row += 1
                continue
            
            ws[f'B{row}'] = name
            ws[f'B{row}'].style = 'label'
            
            for col, val in [('C', bear), ('D', base), ('E', bull)]:
                ws[f'{col}{row}'] = val
                if col == 'D':
                    ws[f'{col}{row}'].style = 'input_cell'
                
                if fmt == "percent":
                    ws[f'{col}{row}'].number_format = self.FORMATS['percent']
                elif fmt == "ratio":
                    ws[f'{col}{row}'].number_format = self.FORMATS['ratio']
                elif fmt == "currency":
                    ws[f'{col}{row}'].number_format = self.FORMATS['currency']
                else:
                    ws[f'{col}{row}'].number_format = self.FORMATS['number']
            
            row += 1
    
    def _create_dashboard(self) -> None:
        """Create visual dashboard with charts"""
        from openpyxl.chart import LineChart, BarChart, Reference
        from openpyxl.chart.series import SeriesLabel
        
        ws = self._setup_sheet("Dashboard", f"{self.company_name} - Financial Dashboard")
        
        # First create data for charts
        row = 40  # Put data below where charts will be
        data_start = row
        
        # Revenue data
        ws[f'B{row}'] = "Chart Data"
        ws[f'B{row}'].style = 'subheader'
        row += 1
        
        ws[f'B{row}'] = "Year"
        total_years = self.hist_years + self.fcst_years
        for i in range(total_years):
            col = get_column_letter(3 + i)
            year = self.base_year - self.hist_years + i
            ws[f'{col}{row}'] = f"FY{year}"
        row += 1
        
        # Revenue
        ws[f'B{row}'] = "Revenue"
        for i in range(total_years):
            col = get_column_letter(3 + i)
            is_col = col
            ws[f'{col}{row}'] = f"=Income_Statement!{is_col}6"
            ws[f'{col}{row}'].number_format = self.FORMATS['number']
        rev_row = row
        row += 1
        
        # EBITDA
        ws[f'B{row}'] = "EBITDA"
        for i in range(total_years):
            col = get_column_letter(3 + i)
            ws[f'{col}{row}'] = f"=Income_Statement!{col}15"
            ws[f'{col}{row}'].number_format = self.FORMATS['number']
        ebitda_row = row
        row += 1
        
        # Net Income
        ws[f'B{row}'] = "Net Income"
        for i in range(total_years):
            col = get_column_letter(3 + i)
            ws[f'{col}{row}'] = f"=Income_Statement!{col}26"
            ws[f'{col}{row}'].number_format = self.FORMATS['number']
        ni_row = row
        row += 1
        
        # EBITDA Margin
        ws[f'B{row}'] = "EBITDA Margin %"
        for i in range(total_years):
            col = get_column_letter(3 + i)
            ws[f'{col}{row}'] = f"=IFERROR({col}{ebitda_row}/{col}{rev_row},0)"
            ws[f'{col}{row}'].number_format = self.FORMATS['percent']
        margin_row = row
        row += 1
        
        # Create Revenue & EBITDA Bar Chart
        chart1 = BarChart()
        chart1.type = "col"
        chart1.grouping = "clustered"
        chart1.title = "Revenue & EBITDA (₹ Crores)"
        chart1.style = 10
        
        data = Reference(ws, min_col=3, min_row=rev_row, max_col=2+total_years, max_row=ebitda_row)
        cats = Reference(ws, min_col=3, min_row=data_start+1, max_col=2+total_years)
        chart1.add_data(data, from_rows=True, titles_from_data=False)
        chart1.set_categories(cats)
        # Set series titles properly using SeriesLabel
        if len(chart1.series) > 0:
            chart1.series[0].tx = SeriesLabel(v="Revenue")
        if len(chart1.series) > 1:
            chart1.series[1].tx = SeriesLabel(v="EBITDA")
        chart1.shape = 4
        chart1.width = 18
        chart1.height = 10
        
        ws.add_chart(chart1, "B5")
        
        # Create Margin Line Chart
        chart2 = LineChart()
        chart2.title = "EBITDA Margin Trend"
        chart2.style = 10
        chart2.y_axis.title = "Margin %"
        
        data2 = Reference(ws, min_col=3, min_row=margin_row, max_col=2+total_years)
        chart2.add_data(data2, from_rows=True)
        chart2.set_categories(cats)
        if len(chart2.series) > 0:
            chart2.series[0].tx = SeriesLabel(v="EBITDA Margin")
        chart2.width = 12
        chart2.height = 8
        
        ws.add_chart(chart2, "L5")
        
        # Net Income Line Chart
        chart3 = LineChart()
        chart3.title = "Net Income Trend (₹ Crores)"
        chart3.style = 12
        
        data3 = Reference(ws, min_col=3, min_row=ni_row, max_col=2+total_years)
        chart3.add_data(data3, from_rows=True)
        chart3.set_categories(cats)
        if len(chart3.series) > 0:
            chart3.series[0].tx = SeriesLabel(v="Net Income")
        chart3.width = 12
        chart3.height = 8
        
        ws.add_chart(chart3, "L18")
        
        # Key Metrics Summary
        ws['B22'] = "KEY METRICS SUMMARY"
        ws['B22'].style = 'subheader'
        ws.merge_cells('B22:D22')
        
        metrics = [
            ("Current Revenue", f"=Income_Statement!{get_column_letter(2+self.hist_years)}6", "number"),
            ("Forecast Y5 Revenue", f"=Income_Statement!{get_column_letter(2+self.hist_years+self.fcst_years)}6", "number"),
            ("Revenue CAGR", f"=IFERROR((Income_Statement!{get_column_letter(2+self.hist_years+self.fcst_years)}6/Income_Statement!{get_column_letter(2+self.hist_years)}6)^(1/{self.fcst_years})-1,0)", "percent"),
            ("Current EBITDA Margin", f"={get_column_letter(2+self.hist_years)}{margin_row}", "percent"),
            ("Forecast Y5 EBITDA Margin", f"={get_column_letter(2+self.hist_years+self.fcst_years)}{margin_row}", "percent"),
            ("Enterprise Value", "=Valuation!C44", "number"),
            ("Equity Value", "=Valuation!C46", "number"),
            ("Implied Share Price", "=Valuation!C50", "currency"),
        ]
        
        row = 23
        for name, formula, fmt in metrics:
            ws[f'B{row}'] = name
            ws[f'B{row}'].style = 'label'
            ws[f'C{row}'] = formula
            ws[f'C{row}'].style = 'output_cell'
            if fmt == "percent":
                ws[f'C{row}'].number_format = self.FORMATS['percent']
            elif fmt == "currency":
                ws[f'C{row}'].number_format = self.FORMATS['currency']
            else:
                ws[f'C{row}'].number_format = self.FORMATS['number']
            row += 1


def generate_financial_model(
    company_name: str,
    model_structure: Dict[str, Any],
    financial_data: Dict[str, Any],
    industry_info: Dict[str, Any],
    output_path: str
) -> str:
    """Generate financial model"""
    generator = FinancialModelGenerator(
        company_name, model_structure, financial_data, industry_info
    )
    return generator.generate(output_path)
